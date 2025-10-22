"""
Management portal routes for Moulya College Management System
Handles all management functionality
"""

from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify
from routes.auth import login_required
from services.management_service import ManagementService
from services.auth_service import AuthService
from services.reporting_service import ReportingService
from services.excel_export_service import ExcelExportService
from services.mongodb_service import mongodb_service
from models.user import Lecturer
from models.academic import Course, Subject
from models.student import Student
from database import db
from utils.validators import validate_username, validate_password
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
from io import BytesIO
from datetime import datetime

def is_ajax_request():
    """Check if the current request is an AJAX request"""
    return (request.headers.get('X-Requested-With') == 'XMLHttpRequest' or 
            request.headers.get('Content-Type') == 'application/json' or
            request.is_json)

management_bp = Blueprint('management', __name__)

@management_bp.route('/dashboard')
@login_required('management')
def dashboard():
    """Management dashboard with overview statistics"""
    try:
        # Get overview statistics
        stats = ManagementService.get_dashboard_stats()
        return render_template('management/dashboard.html', stats=stats)
    except Exception as e:
        flash(f'Error loading dashboard: {str(e)}', 'error')
        return render_template('management/dashboard.html', stats={})

@management_bp.route('/tracking')
@login_required('management')
def tracking_dashboard():
    """Tracking Lecturer Updates (Marks & Attendance) page."""
    try:
        # Provide basic lookups for filters/selects
        courses = Course.query.filter_by(is_active=True).order_by(Course.name).all()
        subjects = Subject.query.filter_by(is_active=True).order_by(Subject.name).all()
        lecturers = Lecturer.query.filter_by(is_active=True).order_by(Lecturer.name).all()
        from datetime import datetime
        now_dt = datetime.now()
        return render_template(
            'management/tracking.html',
            courses=courses,
            subjects=subjects,
            lecturers=lecturers,
            current_month=now_dt.month,
            current_year=now_dt.year,
        )
    except Exception as e:
        flash(f'Error loading tracking page: {str(e)}', 'error')
        from datetime import datetime
        now_dt = datetime.now()
        return render_template(
            'management/tracking.html',
            courses=[],
            subjects=[],
            lecturers=[],
            current_month=now_dt.month,
            current_year=now_dt.year,
        )

@management_bp.route('/tracking/marks')
@login_required('management')
def tracking_marks():
    """Return updated vs pending lecturers for marks by assessment/filters."""
    try:
        assessment_type = request.args.get('assessment_type')
        course_id = request.args.get('course_id', type=int)
        subject_id = request.args.get('subject_id', type=int)
        lecturer_id = request.args.get('lecturer_id', type=int)

        result = ManagementService.get_marks_tracking(assessment_type, course_id, subject_id, lecturer_id)
        return jsonify({'success': True, **result})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error fetching marks tracking: {str(e)}'}), 500

@management_bp.route('/tracking/attendance')
@login_required('management')
def tracking_attendance():
    """Return updated vs pending lecturers for attendance by month/year/filters."""
    try:
        # Support 'overall' and numeric months
        month_param = request.args.get('month')
        if month_param and str(month_param).lower() == 'overall':
            month = 'overall'
        else:
            try:
                month = int(month_param) if month_param is not None else None
            except Exception:
                month = None
        year = request.args.get('year', type=int)
        course_id = request.args.get('course_id', type=int)
        subject_id = request.args.get('subject_id', type=int)
        lecturer_id = request.args.get('lecturer_id', type=int)
        deputation = request.args.get('deputation')

        result = ManagementService.get_attendance_tracking(
            month, year, course_id, subject_id, lecturer_id, bool(deputation)
        )
        return jsonify({'success': True, 'deputation': bool(deputation), **result})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error fetching attendance tracking: {str(e)}'}), 500

@management_bp.route('/tracking/export/marks')
@login_required('management')
def tracking_export_marks():
    """Export tracking list (marks) to Excel."""
    try:
        from flask import make_response
        assessment_type = request.args.get('assessment_type')
        course_id = request.args.get('course_id', type=int)
        subject_id = request.args.get('subject_id', type=int)
        lecturer_id = request.args.get('lecturer_id', type=int)
        status = request.args.get('status')  # 'updated' or 'pending'

        data = ManagementService.get_marks_tracking(assessment_type, course_id, subject_id, lecturer_id)
        wb = ExcelExportService.create_workbook()
        ws = wb.active
        ws.title = 'Marks Tracking'

        # Compute top labels for Assessment and Status
        if status in ['updated', 'pending']:
            status_display = 'Updated' if status == 'updated' else 'Pending'
            title_suffix = status_display
        else:
            status_display = 'All Status'
            title_suffix = None

        at = (assessment_type or '').lower()
        pretty_map = {
            'internal1': 'Internal 1',
            'internal2': 'Internal 2',
            'assignment': 'Assignment',
            'project': 'Project',
            'any': 'Any'
        }
        assessment_display = pretty_map.get(at, (assessment_type or ''))

        # Write top info row
        ws.cell(row=1, column=1, value=f"Assessment: {assessment_display}")
        ws.cell(row=1, column=2, value=f"Status: {status_display}")
        header_font = Font(bold=True, size=12)
        for col in [1, 2]:
            cell = ws.cell(row=1, column=col)
            cell.font = header_font

        # Updated headers: remove Assessment and Status columns; keep single Class
        headers = ['Lecturer', 'Subject', 'Code', 'Class']
        ExcelExportService.style_header_row(ws, 3, headers)

        row = 4
        keys_to_iterate = [status] if status in ['updated', 'pending'] else ['updated', 'pending']
        for status_key in keys_to_iterate:
            for item in data.get(status_key, []):
                col = 1
                ws.cell(row=row, column=col, value=item.get('lecturer_name')); col += 1
                ws.cell(row=row, column=col, value=item.get('subject_name')); col += 1
                ws.cell(row=row, column=col, value=item.get('subject_code')); col += 1
                # Build pretty class from course_code; remove underscores
                class_text = None
                try:
                    cc = item.get('course_code') or ''
                    class_text = str(cc).replace('_', ' ').strip() if cc else ''
                except Exception:
                    class_text = item.get('class_display') or ''
                ws.cell(row=row, column=col, value=class_text); col += 1
                row += 1

        ExcelExportService.auto_adjust_columns(ws)
        excel_data = ExcelExportService.workbook_to_bytes(wb)
        response = make_response(excel_data)
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        fname = 'marks_tracking'
        if title_suffix:
            fname += f'_{title_suffix.lower()}'
        fname += '.xlsx'
        response.headers['Content-Disposition'] = f'attachment; filename={fname}'
        return response
    except Exception as e:
        flash(f'Error exporting marks tracking: {str(e)}', 'error')
        return redirect(url_for('management.tracking_dashboard'))

@management_bp.route('/tracking/export/attendance')
@login_required('management')
def tracking_export_attendance():
    """Export tracking list (attendance) to Excel."""
    try:
        from flask import make_response
        import calendar
        # Support 'overall' and numeric months
        month_param = request.args.get('month')
        if month_param and str(month_param).lower() == 'overall':
            month = 'overall'
        else:
            try:
                month = int(month_param) if month_param is not None else None
            except Exception:
                month = None
        year = request.args.get('year', type=int)
        course_id = request.args.get('course_id', type=int)
        subject_id = request.args.get('subject_id', type=int)
        lecturer_id = request.args.get('lecturer_id', type=int)
        status = request.args.get('status')  # 'updated' or 'pending'

        data = ManagementService.get_attendance_tracking(month, year, course_id, subject_id, lecturer_id)
        wb = ExcelExportService.create_workbook()
        ws = wb.active
        ws.title = 'Attendance Tracking'

        # Get month name
        month_name = calendar.month_name[month] if month else 'All Months'
        
        # Determine status display
        if status in ['updated', 'pending']:
            status_display = 'Updated' if status == 'updated' else 'Pending'
            title_suffix = status_display
        else:
            status_display = 'All Status'
            title_suffix = None

        # Add header information at the top
        ws.cell(row=1, column=1, value=f"Month: {month_name} {year}")
        ws.cell(row=1, column=2, value=f"Status: {status_display}")
        
        # Style the header row
        header_font = Font(bold=True, size=12)
        for col in [1, 2]:
            cell = ws.cell(row=1, column=col)
            cell.font = header_font

        # Updated headers for attendance export: single Class column (no underscores), drop Status column
        headers = ['Lecturer', 'Subject', 'Code', 'Class']
        ExcelExportService.style_header_row(ws, 3, headers)  # Start table headers at row 3

        row = 4  # Start data at row 4
        keys_to_iterate = [status] if status in ['updated', 'pending'] else ['updated', 'pending']
        for status_key in keys_to_iterate:
            for item in data.get(status_key, []):
                col = 1
                ws.cell(row=row, column=col, value=item.get('lecturer_name')); col += 1
                ws.cell(row=row, column=col, value=item.get('subject_name')); col += 1
                ws.cell(row=row, column=col, value=item.get('subject_code')); col += 1
                # Build pretty class from course_code (e.g., I_BCA_A -> I BCA A)
                class_text = None
                try:
                    cc = item.get('course_code') or ''
                    class_text = str(cc).replace('_', ' ').strip() if cc else ''
                except Exception:
                    class_text = item.get('class_display') or ''
                ws.cell(row=row, column=col, value=class_text); col += 1
                # No Status column in export anymore
                row += 1

        ExcelExportService.auto_adjust_columns(ws)
        excel_data = ExcelExportService.workbook_to_bytes(wb)
        response = make_response(excel_data)
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        fname = 'attendance_tracking'
        if title_suffix:
            fname += f'_{title_suffix.lower()}'
        fname += '.xlsx'
        response.headers['Content-Disposition'] = f'attachment; filename={fname}'
        return response
    except Exception as e:
        flash(f'Error exporting attendance tracking: {str(e)}', 'error')
        return redirect(url_for('management.tracking_dashboard'))
@management_bp.route('/lecturers')
@login_required('management')
def lecturers():
    """List all lecturers"""
    try:
        page = request.args.get('page', 1, type=int)
        search = request.args.get('search', '', type=str)
        
        lecturers_data = ManagementService.get_lecturers_paginated(page, search)
        subjects = Subject.query.filter_by(is_active=True).order_by(Subject.name).all()
        return render_template('management/lecturers.html', 
                             lecturers=lecturers_data['lecturers'],
                             pagination=lecturers_data['pagination'],
                             search=search,
                             subjects=subjects)
    except Exception as e:
        flash(f'Error loading lecturers: {str(e)}', 'error')
        subjects = Subject.query.filter_by(is_active=True).order_by(Subject.name).all()
        return render_template('management/lecturers.html', lecturers=[], pagination=None, subjects=subjects)

@management_bp.route('/students')
@login_required('management')
def students():
    """List all students"""
    try:
        page = request.args.get('page', 1, type=int)
        search = request.args.get('search', '', type=str)
        course_id = request.args.get('course_id', type=int)
        
        students_data = ManagementService.get_students_paginated(page, search, course_id)
        courses = Course.query.filter_by(is_active=True).all()
        
        return render_template('management/students.html',
                             students=students_data['students'],
                             pagination=students_data['pagination'],
                             courses=courses,
                             search=search,
                             selected_course=course_id)
    except Exception as e:
        flash(f'Error loading students: {str(e)}', 'error')
        return render_template('management/students.html', students=[], pagination=None, courses=[])

@management_bp.route('/courses')
@login_required('management')
def courses():
    """Manage courses"""
    try:
        page = request.args.get('page', 1, type=int)
        search = request.args.get('search', '', type=str)
        
        query = Course.query.filter_by(is_active=True)
        
        if search:
            query = query.filter(
                db.or_(
                    Course.name.contains(search),
                    Course.code.contains(search)
                )
            )
        
        pagination = query.order_by(Course.name).paginate(
            page=page, per_page=20, error_out=False
        )
        
        return render_template('management/courses.html', 
                             courses=pagination.items,
                             pagination=pagination,
                             search=search)
    except Exception as e:
        flash(f'Error loading courses: {str(e)}', 'error')
        return render_template('management/courses.html', courses=[], pagination=None, search='')

@management_bp.route('/subjects')
@login_required('management')
def subjects():
    """Manage subjects"""
    try:
        course_id = request.args.get('course_id', type=int)
        subjects_data = ManagementService.get_subjects_by_course(course_id)
        courses = Course.query.filter_by(is_active=True).all()
        
        return render_template('management/subjects.html',
                             subjects=subjects_data,
                             courses=courses,
                             selected_course=course_id)
    except Exception as e:
        flash(f'Error loading subjects: {str(e)}', 'error')
        return render_template('management/subjects.html', subjects=[], courses=[])



@management_bp.route('/lecturers/add', methods=['POST'])
@login_required('management')
def add_lecturer():
    """Add single lecturer"""
    try:
        lecturer_data = {
            'lecturer_id': request.form.get('lecturer_id', '').strip(),
            'name': request.form.get('name', '').strip(),
            'subject_ids': request.form.getlist('subject_ids')
        }
        
        success, message = ManagementService.add_lecturer(lecturer_data)
        
        # Handle AJAX requests
        if is_ajax_request():
            return jsonify({'success': success, 'message': message})
        
        if success:
            flash(message, 'success')
        else:
            flash(message, 'error')
            
    except Exception as e:
        error_msg = f'Error adding lecturer: {str(e)}'
        if is_ajax_request():
            return jsonify({'success': False, 'message': error_msg})
        flash(error_msg, 'error')
    
    return redirect(url_for('management.lecturers'))

@management_bp.route('/lecturers/bulk', methods=['POST'])
@login_required('management')
def bulk_add_lecturers():
    """Bulk add lecturers from Excel file"""
    try:
        if 'file' not in request.files:
            error_msg = 'No file selected'
            if is_ajax_request():
                return jsonify({'success': False, 'message': error_msg})
            flash(error_msg, 'error')
            return redirect(url_for('management.lecturers'))

        file = request.files['file']
        if file.filename == '':
            error_msg = 'No file selected'
            if is_ajax_request():
                return jsonify({'success': False, 'message': error_msg})
            flash(error_msg, 'error')
            return redirect(url_for('management.lecturers'))

        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            error_msg = 'Please upload an Excel file (.xlsx or .xls)'
            if is_ajax_request():
                return jsonify({'success': False, 'message': error_msg})
            flash(error_msg, 'error')
            return redirect(url_for('management.lecturers'))

        success, message, credentials, errors = ManagementService.bulk_add_lecturers(file.read())

        # Handle AJAX requests
        if is_ajax_request():
            return jsonify({'success': success, 'message': message, 'errors': errors})

        if success:
            flash(message, 'success')
            if errors:
                flash(f'Some errors occurred: {"; ".join(errors[:5])}', 'warning')
        else:
            flash(message, 'error')
            if errors:
                flash(f'Errors: {"; ".join(errors[:5])}', 'error')

    except Exception as e:
        error_msg = f'Error processing file: {str(e)}'
        if is_ajax_request():
            return jsonify({'success': False, 'message': error_msg})
        flash(error_msg, 'error')

    return redirect(url_for('management.lecturers'))

@management_bp.route('/lecturers/<int:lecturer_id>/toggle-status', methods=['POST'])
@login_required('management')
def toggle_lecturer_status(lecturer_id):
    """On deactivate: permanently delete the lecturer and dependencies; on inactive -> active: reactivate.

    This preserves the existing button/JS flow while ensuring a hard delete when switching from
    active to inactive.
    """
    try:
        lecturer = Lecturer.query.get_or_404(lecturer_id)

        if lecturer.is_active:
            # Deactivation request -> permanently delete
            success, message = ManagementService.delete_lecturer_permanently(lecturer_id)
            if not success:
                if is_ajax_request():
                    return jsonify({'success': False, 'message': message})
                flash(message, 'error')
                return redirect(url_for('management.lecturers'))

            if is_ajax_request():
                return jsonify({'success': True, 'message': message})
            flash(message, 'success')
            return redirect(url_for('management.lecturers'))
        else:
            # Inactive -> Active toggle retains record
            lecturer.is_active = True
            db.session.commit()
            message = f'Lecturer {lecturer.name} has been activated'
            if is_ajax_request():
                return jsonify({'success': True, 'message': message})
            flash(message, 'success')
            return redirect(url_for('management.lecturers'))

    except Exception as e:
        error_msg = f'Error updating lecturer status: {str(e)}'
        if is_ajax_request():
            return jsonify({'success': False, 'message': error_msg})
        flash(error_msg, 'error')
    
    return redirect(url_for('management.lecturers'))

@management_bp.route('/lecturers/<int:lecturer_id>/reset-password', methods=['POST'])
@login_required('management')
def reset_lecturer_password(lecturer_id):
    """Reset lecturer password. If username/password provided, set custom values with validation."""
    try:
        lecturer = Lecturer.query.get_or_404(lecturer_id)

        payload = {}
        if request.is_json:
            payload = request.get_json(silent=True) or {}
        username = (payload.get('username') or request.form.get('username') or '').strip()
        password = (payload.get('password') or request.form.get('password') or '').strip()

        # If either field present, treat as custom update
        if username or password:
            # Validate fields
            if username:
                valid, msg = validate_username(username)
                if not valid:
                    if is_ajax_request():
                        return jsonify({'success': False, 'message': msg}), 400
                    flash(msg, 'error')
                    return redirect(url_for('management.lecturers'))
                # Uniqueness check excluding current lecturer
                existing = Lecturer.query.filter(Lecturer.username == username, Lecturer.id != lecturer.id).first()
                if existing:
                    msg = 'Username already exists. Please choose a different one.'
                    if is_ajax_request():
                        return jsonify({'success': False, 'message': msg}), 400
                    flash(msg, 'error')
                    return redirect(url_for('management.lecturers'))

            if password:
                valid, msg = validate_password(password)
                if not valid:
                    if is_ajax_request():
                        return jsonify({'success': False, 'message': msg}), 400
                    flash(msg, 'error')
                    return redirect(url_for('management.lecturers'))

            # Apply updates
            if username:
                lecturer.username = username
            if password:
                lecturer.set_password(password)
            db.session.commit()

            message = 'Credentials updated successfully'
            if is_ajax_request():
                return jsonify({'success': True, 'message': message})
            flash(message, 'success')
            return redirect(url_for('management.lecturers'))

        # Default behavior: auto-generate a new password
        success, new_password, message = AuthService.reset_lecturer_password(lecturer.lecturer_id)
        if success:
            response_message = f'Password reset for {lecturer.name}. New password: {new_password}'
            if is_ajax_request():
                return jsonify({'success': True, 'message': response_message, 'new_password': new_password})
            flash(response_message, 'success')
        else:
            if is_ajax_request():
                return jsonify({'success': False, 'message': message})
            flash(message, 'error')

    except Exception as e:
        error_msg = f'Error resetting password: {str(e)}'
        if is_ajax_request():
            return jsonify({'success': False, 'message': error_msg})
        flash(error_msg, 'error')
    
    return redirect(url_for('management.lecturers'))

@management_bp.route('/lecturers/reset-passwords-all', methods=['POST'])
@login_required('management')
def reset_passwords_all_lecturers():
    """Set a custom password for ALL lecturers"""
    try:
        # Accept JSON or form data
        password = None
        if request.is_json:
            payload = request.get_json(silent=True) or {}
            password = (payload.get('password') or '').strip()
        if not password:
            password = (request.form.get('password') or '').strip()

        if not password:
            message = 'Password is required'
            if is_ajax_request():
                return jsonify({'success': False, 'message': message}), 400
            flash(message, 'error')
            return redirect(url_for('management.lecturers'))

        # Basic validation
        if len(password) < 6:
            message = 'Password must be at least 6 characters'
            if is_ajax_request():
                return jsonify({'success': False, 'message': message}), 400
            flash(message, 'error')
            return redirect(url_for('management.lecturers'))

        from models.user import Lecturer
        from database import db

        lecturers = Lecturer.query.all()
        for lecturer in lecturers:
            lecturer.set_password(password)
        db.session.commit()

        message = f'Password updated for {len(lecturers)} lecturer(s).'
        if is_ajax_request():
            return jsonify({'success': True, 'message': message})
        flash(message, 'success')
    except Exception as e:
        try:
            from database import db as _db
            _db.session.rollback()
        except Exception:
            pass
        error_msg = f'Error resetting passwords: {str(e)}'
        if is_ajax_request():
            return jsonify({'success': False, 'message': error_msg}), 500
        flash(error_msg, 'error')

    return redirect(url_for('management.lecturers'))

@management_bp.route('/lecturers/credentials')
@login_required('management')
def view_bulk_credentials():
    """View bulk upload credentials"""
    credentials = session.pop('bulk_credentials', [])
    return render_template('management/bulk_credentials.html', credentials=credentials)

@management_bp.route('/lecturers/<int:lecturer_id>/password')
@login_required('management')
def get_lecturer_password(lecturer_id):
    """Get lecturer's decrypted password for management"""
    try:
        lecturer = Lecturer.query.get_or_404(lecturer_id)
        password = lecturer.get_decrypted_password()
        
        if password:
            return jsonify({'success': True, 'password': password})
        else:
            return jsonify({'success': False, 'message': 'Password not available'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error retrieving password: {str(e)}'})

@management_bp.route('/lecturers/<int:lecturer_id>/assign-subjects', methods=['POST'])
@login_required('management')
def assign_subjects_to_lecturer(lecturer_id):
    """Assign subjects to a lecturer"""
    try:
        subject_ids = request.form.getlist('subject_ids')
        
        if not subject_ids:
            if is_ajax_request():
                return jsonify({'success': False, 'message': 'Please select at least one subject'})
            flash('Please select at least one subject', 'error')
            return redirect(url_for('management.lecturers'))
        
        success, message = ManagementService.assign_subjects_to_lecturer(lecturer_id, subject_ids)
        
        if is_ajax_request():
            return jsonify({'success': success, 'message': message})
        
        if success:
            flash(message, 'success')
        else:
            flash(message, 'error')
        
        return redirect(url_for('management.lecturers'))
        
    except Exception as e:
        error_msg = f'Error assigning subjects: {str(e)}'
        if is_ajax_request():
            return jsonify({'success': False, 'message': error_msg})
        flash(error_msg, 'error')
        return redirect(url_for('management.lecturers'))

@management_bp.route('/lecturers/<int:lecturer_id>/unassign-subject', methods=['POST'])
@login_required('management')
def unassign_subject_from_lecturer(lecturer_id):
    """Unassign a subject from a lecturer (current academic year)"""
    try:
        subject_id = request.form.get('subject_id')
        if not subject_id:
            if is_ajax_request():
                return jsonify({'success': False, 'message': 'Subject is required'})
            flash('Subject is required', 'error')
            return redirect(url_for('management.lecturers'))

        success, message = ManagementService.unassign_subject_from_lecturer(lecturer_id, subject_id)

        if is_ajax_request():
            return jsonify({'success': success, 'message': message})

        if success:
            flash(message, 'success')
        else:
            flash(message, 'error')

    except Exception as e:
        error_msg = f'Error unassigning subject: {str(e)}'
        if is_ajax_request():
            return jsonify({'success': False, 'message': error_msg})
        flash(error_msg, 'error')

    return redirect(url_for('management.lecturers'))

@management_bp.route('/lecturers/credentials/export')
@login_required('management')
def export_lecturer_credentials():
    """Export lecturer credentials to Excel"""
    try:
        from flask import make_response
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
        from io import BytesIO
        
        # Get all active lecturers with their credentials
        lecturers = Lecturer.query.filter_by(is_active=True).all()
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lecturer Credentials"
        
        # Headers
        headers = ['Lecturer ID', 'Name', 'Username', 'Password', 'Assigned Subjects', 'Created Date']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows
        for row, lecturer in enumerate(lecturers, 2):
            ws.cell(row=row, column=1, value=lecturer.lecturer_id)
            ws.cell(row=row, column=2, value=lecturer.name)
            ws.cell(row=row, column=3, value=lecturer.username)
            ws.cell(row=row, column=4, value=lecturer.get_decrypted_password() or 'N/A')
            assigned_subjects = [subject.name for subject in lecturer.get_assigned_subjects()]
            ws.cell(row=row, column=5, value=', '.join(assigned_subjects) if assigned_subjects else 'No subjects assigned')
            ws.cell(row=row, column=6, value=lecturer.created_at.strftime('%Y-%m-%d') if lecturer.created_at else '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=lecturer_credentials.xlsx'
        
        return response
        
    except Exception as e:
        flash(f'Error exporting credentials: {str(e)}', 'error')
        return redirect(url_for('management.lecturers'))

@management_bp.route('/students/add', methods=['POST'])
@login_required('management')
def add_student():
    """Add single student"""
    try:
        student_data = {
            'roll_number': request.form.get('roll_number', '').strip(),
            'name': request.form.get('name', '').strip(),
            'course_id': int(request.form.get('course_id')),
            'academic_year': int(request.form.get('academic_year')),
            'current_semester': int(request.form.get('current_semester', 1)),
            'email': request.form.get('email', '').strip() or None
        }
        
        success, message = ManagementService.add_student(student_data)
        
        # Handle AJAX requests
        if is_ajax_request():
            return jsonify({'success': success, 'message': message})
        
        if success:
            flash(message, 'success')
        else:
            flash(message, 'error')
            
    except Exception as e:
        error_msg = f'Error adding student: {str(e)}'
        if is_ajax_request():
            return jsonify({'success': False, 'message': error_msg})
        flash(error_msg, 'error')
    
    return redirect(url_for('management.students'))

@management_bp.route('/students/bulk', methods=['POST'])
@login_required('management')
def bulk_add_students():
    """Bulk add students from Excel file"""
    try:
        if 'file' not in request.files:
            error_msg = 'No file selected'
            if is_ajax_request():
                return jsonify({'success': False, 'message': error_msg})
            flash(error_msg, 'error')
            return redirect(url_for('management.students'))
        
        file = request.files['file']
        if file.filename == '':
            error_msg = 'No file selected'
            if is_ajax_request():
                return jsonify({'success': False, 'message': error_msg})
            flash(error_msg, 'error')
            return redirect(url_for('management.students'))
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            error_msg = 'Please upload an Excel file (.xlsx or .xls)'
            if is_ajax_request():
                return jsonify({'success': False, 'message': error_msg})
            flash(error_msg, 'error')
            return redirect(url_for('management.students'))
        
        success, message, errors = ManagementService.bulk_add_students(file.read())
        
        # Handle AJAX requests
        if is_ajax_request():
            return jsonify({'success': success, 'message': message, 'errors': errors})
        
        if success:
            flash(message, 'success')
            if errors:
                flash(f'Some errors occurred: {"; ".join(errors[:5])}', 'warning')
        else:
            flash(message, 'error')
            if errors:
                flash(f'Errors: {"; ".join(errors[:5])}', 'error')
                
    except Exception as e:
        error_msg = f'Error processing file: {str(e)}'
        if is_ajax_request():
            return jsonify({'success': False, 'message': error_msg})
        flash(error_msg, 'error')
    
    return redirect(url_for('management.students'))

@management_bp.route('/students/<int:student_id>/toggle-status', methods=['POST'])
@login_required('management')
def toggle_student_status(student_id):
    """On deactivate: permanently delete the student; on inactive -> active: reactivate."""
    try:
        student = Student.query.get_or_404(student_id)

        if student.is_active:
            success, message = ManagementService.delete_student_permanently(student_id)
            if not success:
                if is_ajax_request():
                    return jsonify({'success': False, 'message': message})
                flash(message, 'error')
                return redirect(url_for('management.students'))

            if is_ajax_request():
                return jsonify({'success': True, 'message': message})
            flash(message, 'success')
            return redirect(url_for('management.students'))
        else:
            student.is_active = True
            db.session.commit()
            message = f'Student {student.name} has been activated'
            if is_ajax_request():
                return jsonify({'success': True, 'message': message})
            flash(message, 'success')
            return redirect(url_for('management.students'))

    except Exception as e:
        error_msg = f'Error updating student status: {str(e)}'
        if is_ajax_request():
            return jsonify({'success': False, 'message': error_msg})
        flash(error_msg, 'error')
    
    return redirect(url_for('management.students'))

@management_bp.route('/courses/add', methods=['POST'])
@login_required('management')
def add_course():
    """Add new course"""
    try:
        course_data = {
            'name': request.form.get('name', '').strip(),
            'code': request.form.get('code', '').strip(),
            'description': request.form.get('description', '').strip() or None,
            'duration_years': int(request.form.get('duration_years', 3)),
            'total_semesters': int(request.form.get('total_semesters', 6))
        }
        
        success, message = ManagementService.create_course(course_data)
        
        # Handle AJAX requests
        if is_ajax_request():
            return jsonify({'success': success, 'message': message})
        
        if success:
            flash(message, 'success')
        else:
            flash(message, 'error')
            
    except Exception as e:
        error_msg = f'Error adding course: {str(e)}'
        if is_ajax_request():
            return jsonify({'success': False, 'message': error_msg})
        flash(error_msg, 'error')
    
    return redirect(url_for('management.courses'))

@management_bp.route('/courses/<int:course_id>/toggle-status', methods=['POST'])
@login_required('management')
def toggle_course_status(course_id):
    """On deactivate: permanently delete the course; on inactive -> active: reactivate."""
    try:
        course = Course.query.get_or_404(course_id)

        if course.is_active:
            success, message = ManagementService.delete_course_permanently(course_id)
            if not success:
                if is_ajax_request():
                    return jsonify({'success': False, 'message': message})
                flash(message, 'error')
                return redirect(url_for('management.courses'))

            if is_ajax_request():
                return jsonify({'success': True, 'message': message})
            flash(message, 'success')
            return redirect(url_for('management.courses'))
        else:
            course.is_active = True
            db.session.commit()
            message = f'Course {course.name} has been activated'
            if is_ajax_request():
                return jsonify({'success': True, 'message': message})
            flash(message, 'success')
            return redirect(url_for('management.courses'))

    except Exception as e:
        error_msg = f'Error updating course status: {str(e)}'
        if is_ajax_request():
            return jsonify({'success': False, 'message': error_msg})
        flash(error_msg, 'error')
    
    return redirect(url_for('management.courses'))

@management_bp.route('/subjects/add', methods=['POST'])
@login_required('management')
def add_subject():
    """Add new subject"""
    try:
        subject_data = {
            'name': request.form.get('name', '').strip(),
            'code': request.form.get('code', '').strip(),
            'course_id': int(request.form.get('course_id')),
            'year': int(request.form.get('year')),
            'semester': int(request.form.get('semester')),
            'description': request.form.get('description', '').strip() or None
        }
        
        success, message = ManagementService.create_subject(subject_data)
        
        # Handle AJAX requests
        if is_ajax_request():
            return jsonify({'success': success, 'message': message})
        
        if success:
            flash(message, 'success')
        else:
            flash(message, 'error')
            
    except Exception as e:
        error_msg = f'Error adding subject: {str(e)}'
        if is_ajax_request():
            return jsonify({'success': False, 'message': error_msg})
        flash(error_msg, 'error')
    
    return redirect(url_for('management.subjects'))

@management_bp.route('/subjects/<int:subject_id>/toggle-status', methods=['POST'])
@login_required('management')
def toggle_subject_status(subject_id):
    """On deactivate: permanently delete the subject; on inactive -> active: reactivate."""
    try:
        subject = Subject.query.get_or_404(subject_id)

        if subject.is_active:
            success, message = ManagementService.delete_subject_permanently(subject_id)
            if not success:
                if is_ajax_request():
                    return jsonify({'success': False, 'message': message})
                flash(message, 'error')
                return redirect(url_for('management.subjects'))

            if is_ajax_request():
                return jsonify({'success': True, 'message': message})
            flash(message, 'success')
            return redirect(url_for('management.subjects'))
        else:
            subject.is_active = True
            db.session.commit()
            message = f'Subject {subject.name} has been activated'
            if is_ajax_request():
                return jsonify({'success': True, 'message': message})
            flash(message, 'success')
            return redirect(url_for('management.subjects'))

    except Exception as e:
        error_msg = f'Error updating subject status: {str(e)}'
        if is_ajax_request():
            return jsonify({'success': False, 'message': error_msg})
        flash(error_msg, 'error')
    
    return redirect(url_for('management.subjects'))

@management_bp.route('/subjects/<int:subject_id>/update', methods=['POST'])
@login_required('management')
def update_subject(subject_id):
    """Update a subject's name and code (AJAX-friendly)."""
    try:
        payload = {}
        if request.is_json:
            payload = request.get_json(silent=True) or {}
        name = (payload.get('name') or request.form.get('name') or '').strip()
        code = (payload.get('code') or request.form.get('code') or '').strip()

        success, message = ManagementService.update_subject(subject_id, {
            'name': name,
            'code': code
        })

        if is_ajax_request():
            return jsonify({'success': success, 'message': message})

        if success:
            flash(message, 'success')
        else:
            flash(message, 'error')
    except Exception as e:
        error_msg = f'Error updating subject: {str(e)}'
        if is_ajax_request():
            return jsonify({'success': False, 'message': error_msg})
        flash(error_msg, 'error')
    return redirect(url_for('management.subjects'))

# ============================================================================
# REPORTING ROUTES
# ============================================================================

@management_bp.route('/reports')
@login_required('management')
def reports_dashboard():
    """Reports dashboard"""
    try:
        subjects = ReportingService.get_subjects_for_reporting()
        courses = ReportingService.get_courses_for_reporting()
        students = ReportingService.get_students_for_reporting()
        
        return render_template('management/reports.html', 
                             subjects=subjects, 
                             courses=courses,
                             students=students)
    except Exception as e:
        flash(f'Error loading reports dashboard: {str(e)}', 'error')
        return render_template('management/reports.html', subjects=[], courses=[], students=[])
        return render_template('management/reports.html', subjects=[], courses=[], students=[])

@management_bp.route('/reports/students/filter')
@login_required('management')
def filter_students_for_reports():
    """Filter students by course for reports"""
    try:
        course_id = request.args.get('course_id', type=int)
        students = ReportingService.get_students_for_reporting(course_id)
        return jsonify({'success': True, 'students': students})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error filtering students: {str(e)}'})

@management_bp.route('/reports/student/<int:student_id>')
@login_required('management')
def student_report(student_id):
    """Individual student report"""
    try:
        report = ReportingService.get_student_detailed_report(student_id)
        if not report:
            flash('Student not found', 'error')
            return redirect(url_for('management.reports_dashboard'))
        
        return render_template('management/student_report.html', report=report)
    except Exception as e:
        flash(f'Error generating student report: {str(e)}', 'error')
        return redirect(url_for('management.reports_dashboard'))

@management_bp.route('/reports/class/marks/<int:subject_id>')
@login_required('management')
def class_marks_report(subject_id):
    """Class marks report"""
    try:
        assessment_type = request.args.get('assessment_type')
        report = ReportingService.get_class_marks_report(subject_id, assessment_type)
        
        if not report:
            flash('Subject not found', 'error')
            return redirect(url_for('management.reports_dashboard'))
        
        return render_template('management/class_marks_report.html', report=report)
    except Exception as e:
        flash(f'Error generating class marks report: {str(e)}', 'error')
        return redirect(url_for('management.reports_dashboard'))

@management_bp.route('/reports/class/attendance/<int:subject_id>')
@login_required('management')
def class_attendance_report(subject_id):
    """Class attendance report"""
    try:
        raw_month = request.args.get('month')
        print(f"[DEBUG] Raw month from request: {raw_month}")
        
        # Support "overall" to show cumulative attendance across all months
        if raw_month and str(raw_month).lower() == 'overall':
            month = 'overall'
        else:
            month = request.args.get('month', type=int)
        year = request.args.get('year', type=int)
        
        print(f"[DEBUG] Processed month: {month}, year: {year}, subject_id: {subject_id}")
        
        report = ReportingService.get_class_attendance_report(subject_id, month, year)
        
        print(f"[DEBUG] Report returned: {type(report)}, is None: {report is None}")
        if report:
            print(f"[DEBUG] Report keys: {list(report.keys()) if isinstance(report, dict) else 'Not a dict'}")
        
        # Only flash error if report is None (service couldn't create even empty structure)
        if report is None:
            print(f"[DEBUG] Report is None, flashing error")
            flash('Subject not found', 'error')
            return redirect(url_for('management.reports_dashboard'))
        
        print(f"[DEBUG] Rendering template with report")
        return render_template('management/class_attendance_report.html', report=report)
    except Exception as e:
        print(f"[DEBUG] Exception in route: {str(e)}")
        flash(f'Error generating class attendance report: {str(e)}', 'error')
        return redirect(url_for('management.reports_dashboard'))

@management_bp.route('/reports/course/<int:course_id>')
@login_required('management')
def course_overview_report(course_id):
    """Course overview report"""
    try:
        report = ReportingService.get_course_overview_report(course_id)
        
        if not report:
            flash('Course not found', 'error')
            return redirect(url_for('management.reports_dashboard'))
        
        return render_template('management/course_overview_report.html', report=report)
    except Exception as e:
        flash(f'Error generating course overview report: {str(e)}', 'error')
        return redirect(url_for('management.reports_dashboard'))

# ============================================================================
# EXCEL EXPORT ROUTES
# ============================================================================

@management_bp.route('/reports/export/student/<int:student_id>')
@login_required('management')
def export_student_report(student_id):
    """Export student report to Excel"""
    try:
        from flask import make_response
        
        report = ReportingService.get_student_detailed_report(student_id)
        if not report:
            flash('Student not found', 'error')
            return redirect(url_for('management.reports_dashboard'))
        
        workbook = ExcelExportService.export_student_report(report)
        if not workbook:
            flash('Error generating Excel file', 'error')
            return redirect(url_for('management.student_report', student_id=student_id))
        
        excel_data = ExcelExportService.workbook_to_bytes(workbook)
        
        response = make_response(excel_data)
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=student_report_{report["student"]["roll_number"]}.xlsx'
        
        return response
        
    except Exception as e:
        flash(f'Error exporting student report: {str(e)}', 'error')
        return redirect(url_for('management.student_report', student_id=student_id))

# ---------------- PDF Export ----------------
@management_bp.route('/reports/export/student/<int:student_id>/pdf')
@login_required('management')
def export_student_report_pdf(student_id):
    """Export student report to PDF"""
    try:
        from flask import make_response
        report = ReportingService.get_student_detailed_report(student_id)
        if not report:
            flash('Student not found', 'error')
            return redirect(url_for('management.reports_dashboard'))

        pdf_bytes = ReportingService.generate_student_report_pdf(report)
        response = make_response(pdf_bytes)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=student_report_{report["student"]["roll_number"]}.pdf'
        return response
    except Exception as e:
        flash(f'Error exporting PDF: {str(e)}', 'error')
        return redirect(url_for('management.student_report', student_id=student_id))


@management_bp.route('/reports/export/class/attendance/<int:subject_id>/pdf')
@login_required('management')
def export_class_attendance_report_pdf(subject_id):
    """Export class attendance report to PDF"""
    try:
        from flask import make_response
        # Support 'overall' and numeric months, like the view route
        month_param = request.args.get('month')
        if month_param and str(month_param).lower() == 'overall':
            month = 'overall'
        else:
            try:
                month = int(month_param) if month_param is not None else None
            except Exception:
                month = None
        year = request.args.get('year', type=int)

        report = ReportingService.get_class_attendance_report(subject_id, month, year)
        if not report:
            flash('Subject not found', 'error')
            return redirect(url_for('management.reports_dashboard'))
        pdf_bytes = ReportingService.generate_class_attendance_report_pdf(report)
        filename = f"class_attendance_{report['subject']['code']}_{report['month']}_{report['year']}.pdf"
        response = make_response(pdf_bytes)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        return response
    except Exception as e:
        flash(f'Error exporting class attendance PDF: {str(e)}', 'error')
        return redirect(url_for('management.class_attendance_report', subject_id=subject_id))

@management_bp.route('/reports/export/course/<int:course_id>/pdf')
@login_required('management')
def export_course_overview_report_pdf(course_id):
    """Export course overview report to PDF"""
    try:
        from flask import make_response
        report = ReportingService.get_course_overview_report(course_id)
        if not report:
            flash('Course not found', 'error')
            return redirect(url_for('management.reports_dashboard'))
        pdf_bytes = ReportingService.generate_course_overview_report_pdf(report)
        filename = f"course_overview_{report['course']['code']}.pdf"
        response = make_response(pdf_bytes)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        return response
    except Exception as e:
        flash(f'Error exporting course overview PDF: {str(e)}', 'error')
        return redirect(url_for('management.course_overview_report', course_id=course_id))

@management_bp.route('/reports/export/class/marks/<int:subject_id>')
@login_required('management')
def export_class_marks_report(subject_id):
    """Export class marks report to Excel"""
    try:
        from flask import make_response
        
        assessment_type = request.args.get('assessment_type')
        report = ReportingService.get_class_marks_report(subject_id, assessment_type)
        
        if not report:
            flash('Subject not found', 'error')
            return redirect(url_for('management.reports_dashboard'))
        
        workbook = ExcelExportService.export_class_marks_report(report)
        if not workbook:
            flash('Error generating Excel file', 'error')
            return redirect(url_for('management.class_marks_report', subject_id=subject_id))
        
        excel_data = ExcelExportService.workbook_to_bytes(workbook)
        
        filename = f"class_marks_{report['subject']['code']}"
        if assessment_type:
            filename += f"_{assessment_type}"
        filename += ".xlsx"
        
        response = make_response(excel_data)
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        flash(f'Error exporting class marks report: {str(e)}', 'error')
        return redirect(url_for('management.class_marks_report', subject_id=subject_id))

@management_bp.route('/reports/export/class/marks/<int:subject_id>/pdf')
@login_required('management')
def export_class_marks_report_pdf(subject_id):
    """Export class marks report to PDF"""
    try:
        from flask import make_response
        assessment_type = request.args.get('assessment_type')
        report = ReportingService.get_class_marks_report(subject_id, assessment_type)
        if not report:
            flash('Subject not found', 'error')
            return redirect(url_for('management.reports_dashboard'))
        pdf_bytes = ReportingService.generate_class_marks_report_pdf(report)
        filename = f"class_marks_{report['subject']['code']}"
        if assessment_type:
            filename += f"_{assessment_type}"
        filename += ".pdf"
        response = make_response(pdf_bytes)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        return response
    except Exception as e:
        flash(f'Error exporting class marks PDF: {str(e)}', 'error')
        return redirect(url_for('management.class_marks_report', subject_id=subject_id))

@management_bp.route('/reports/export/class/attendance/<int:subject_id>')
@login_required('management')
def export_class_attendance_report(subject_id):
    """Export class attendance report to Excel"""
    try:
        from flask import make_response
        # Support 'overall' and numeric months, like the view route
        month_param = request.args.get('month')
        if month_param and str(month_param).lower() == 'overall':
            month = 'overall'
        else:
            try:
                month = int(month_param) if month_param is not None else None
            except Exception:
                month = None
        year = request.args.get('year', type=int)

        report = ReportingService.get_class_attendance_report(subject_id, month, year)
        
        if not report:
            flash('Subject not found', 'error')
            return redirect(url_for('management.reports_dashboard'))
        
        workbook = ExcelExportService.export_class_attendance_report(report)
        if not workbook:
            flash('Error generating Excel file', 'error')
            return redirect(url_for('management.class_attendance_report', subject_id=subject_id))
        
        excel_data = ExcelExportService.workbook_to_bytes(workbook)
        
        filename = f"class_attendance_{report['subject']['code']}_{report['month']}_{report['year']}.xlsx"
        
        response = make_response(excel_data)
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        flash(f'Error exporting class attendance report: {str(e)}', 'error')
        return redirect(url_for('management.class_attendance_report', subject_id=subject_id))

@management_bp.route('/reports/export/course/<int:course_id>')
@login_required('management')
def export_course_overview_report(course_id):
    """Export course overview report to Excel"""
    try:
        from flask import make_response
        
        report = ReportingService.get_course_overview_report(course_id)
        
        if not report:
            flash('Course not found', 'error')
            return redirect(url_for('management.reports_dashboard'))
        
        workbook = ExcelExportService.export_course_overview_report(report)
        if not workbook:
            flash('Error generating Excel file', 'error')
            return redirect(url_for('management.course_overview_report', course_id=course_id))
        
        excel_data = ExcelExportService.workbook_to_bytes(workbook)
        
        filename = f"course_overview_{report['course']['code']}.xlsx"
        
        response = make_response(excel_data)
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        flash(f'Error exporting course overview report: {str(e)}', 'error')
        return redirect(url_for('management.course_overview_report', course_id=course_id))

@management_bp.route('/reports/comprehensive-class/<int:course_id>')
@login_required('management')
def comprehensive_class_report(course_id):
    """Comprehensive class report for all subjects in a course"""
    try:
        report_type = request.args.get('type', 'attendance')  # 'attendance' or 'marks'
        assessment_type = request.args.get('assessment_type')  # for marks report
        
        # For marks reports, assessment_type is required
        if report_type == 'marks' and not assessment_type:
            flash('Assessment type is required for marks reports', 'error')
            return redirect(url_for('management.reports_dashboard'))
        
        report = ReportingService.get_comprehensive_class_report(course_id, report_type, assessment_type)
        
        if not report:
            flash('Course not found or no data available', 'error')
            return redirect(url_for('management.reports_dashboard'))
        
        return render_template('management/comprehensive_class_report.html', report=report)
        
    except Exception as e:
        flash(f'Error generating comprehensive class report: {str(e)}', 'error')
        return redirect(url_for('management.reports_dashboard'))

@management_bp.route('/reports/export/comprehensive-class/<int:course_id>/excel')
@login_required('management')
def export_comprehensive_class_report_excel(course_id):
    """Export comprehensive class report to Excel"""
    try:
        from flask import make_response
        
        report_type = request.args.get('type', 'attendance')
        assessment_type = request.args.get('assessment_type')
        
        # For marks reports, assessment_type is required
        if report_type == 'marks' and not assessment_type:
            flash('Assessment type is required for marks reports', 'error')
            return redirect(url_for('management.reports_dashboard'))
        
        report = ReportingService.get_comprehensive_class_report(course_id, report_type, assessment_type)
        
        if not report:
            flash('Course not found or no data available', 'error')
            return redirect(url_for('management.reports_dashboard'))
        
        excel_bytes = ExcelExportService.export_comprehensive_class_report(report)
        
        response = make_response(excel_bytes)
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=comprehensive_class_report_{course_id}_{report_type}.xlsx'
        
        return response
        
    except Exception as e:
        flash(f'Error exporting comprehensive class report: {str(e)}', 'error')
        return redirect(url_for('management.reports_dashboard'))

@management_bp.route('/reports/export/comprehensive-class/<int:course_id>/pdf')
@login_required('management')
def export_comprehensive_class_report_pdf(course_id):
    """Export comprehensive class report to PDF"""
    try:
        from flask import make_response
        
        report_type = request.args.get('type', 'attendance')
        assessment_type = request.args.get('assessment_type')
        
        # For marks reports, assessment_type is required
        if report_type == 'marks' and not assessment_type:
            flash('Assessment type is required for marks reports', 'error')
            return redirect(url_for('management.reports_dashboard'))
        
        report = ReportingService.get_comprehensive_class_report(course_id, report_type, assessment_type)
        
        if not report:
            flash('Course not found or no data available', 'error')
            return redirect(url_for('management.reports_dashboard'))
        
        pdf_bytes = ReportingService.generate_comprehensive_class_report_pdf(report)
        
        response = make_response(pdf_bytes)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=comprehensive_class_report_{course_id}_{report_type}.pdf'
        
        return response
        
    except Exception as e:
        flash(f'Error exporting comprehensive class report: {str(e)}', 'error')
        return redirect(url_for('management.reports_dashboard'))

# Student Credentials Management Routes
@management_bp.route('/students_credentials')
@login_required('management')
def students_credentials():
    """Manage students credentials page"""
    return render_template('student/manage_students_credencials.html')

@management_bp.route('/add_student_credential', methods=['POST'])
@login_required('management')
def add_student_credential():
    """Add a new student credential"""
    try:
        data = request.get_json()
        
        roll_number = data.get('roll_number', '').strip()
        name = data.get('name', '').strip()
        password = data.get('password', '').strip()
        class_name = data.get('class_name', '').strip()
        
        # Validate required fields
        if not roll_number or not name or not password:
            return jsonify({'success': False, 'message': 'Roll number, name, and password are required'})
        
        # Add student to MongoDB with class_name if provided
        result = mongodb_service.add_student_credential(roll_number, name, password, class_name)
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error adding student: {str(e)}'})

@management_bp.route('/reset_all_passwords', methods=['POST'])
@login_required('management')
def reset_all_passwords():
    """Reset all student passwords to a provided value"""
    try:
        data = request.get_json() or {}
        new_password = (data.get('new_password') or '').strip()
        if not new_password:
            return jsonify({'success': False, 'message': 'new_password is required'})
        result = mongodb_service.reset_all_passwords(new_password)
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error resetting passwords: {str(e)}'})

@management_bp.route('/bulk_upload_students', methods=['POST'])
@login_required('management')
def bulk_upload_students():
    """Bulk upload students from Excel file"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file uploaded'})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'})
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'Please upload an Excel file (.xlsx or .xls)'})
        
        # Read Excel file
        excel_data = pd.read_excel(file)
        
        # Validate required columns
        required_columns = ['Roll Number', 'Student Name', 'Password']
        missing_columns = [col for col in required_columns if col not in excel_data.columns]
        
        if missing_columns:
            return jsonify({
                'success': False, 
                'message': f'Missing required columns: {", ".join(missing_columns)}'
            })
        
        # Upload to MongoDB
        result = mongodb_service.bulk_upload_students(excel_data)
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error uploading file: {str(e)}'})

@management_bp.route('/get_students_credentials')
@login_required('management')
def get_students_credentials():
    """Get all students credentials"""
    try:
        result = mongodb_service.get_all_students()
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error fetching students: {str(e)}'})

@management_bp.route('/search_students_credentials', methods=['POST'])
@login_required('management')
def search_students_credentials():
    """Search students credentials"""
    try:
        data = request.get_json()
        search_term = data.get('searchTerm', '').strip()
        
        if not search_term:
            result = mongodb_service.get_all_students()
        else:
            result = mongodb_service.search_students(search_term)
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error searching students: {str(e)}'})

@management_bp.route('/delete_student_credential/<student_id>', methods=['DELETE'])
@login_required('management')
def delete_student_credential(student_id):
    """Delete a student credential"""
    try:
        result = mongodb_service.delete_student(student_id)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error deleting student: {str(e)}'})

@management_bp.route('/update_student_credential/<student_id>', methods=['PUT'])
@login_required('management')
def update_student_credential(student_id):
    """Update a student credential"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'message': 'No data provided'}), 400
        
        # Validate required fields
        roll_number = data.get('roll_number', '').strip()
        name = data.get('name', '').strip()
        password = data.get('password', '').strip()
        class_name = data.get('class_name', '').strip()
        
        print(f"DEBUG: Received data - roll_number: '{roll_number}', name: '{name}', password: '{password}', class_name: '{class_name}'")
        print(f"DEBUG: Raw request data: {data}")
        
        if not roll_number:
            return jsonify({'success': False, 'message': 'Roll number is required'}), 400
        if not name:
            return jsonify({'success': False, 'message': 'Student name is required'}), 400
        if not password:
            return jsonify({'success': False, 'message': 'Password is required'}), 400
        
        # Prepare update data
        update_data = {
            'roll_number': roll_number,
            'name': name,
            'password': password,
            'username': roll_number.lower(),
            'updatedAt': datetime.now()
        }
        
        # Always update class_name if provided in the request
        if 'class_name' in data:
            if class_name and class_name.strip():
                update_data['class_name'] = class_name.strip()
                # Auto-generate course_code from class_name by replacing spaces with underscores
                course_code = class_name.strip().replace(' ', '_')
                update_data['course_code'] = course_code
                print(f"DEBUG: Added class_name to update_data: '{class_name.strip()}'")
                print(f"DEBUG: Auto-generated course_code: '{course_code}'")
            else:
                # If class_name is explicitly set to empty, we'll handle this in MongoDB service
                update_data['class_name'] = ''
                update_data['course_code'] = ''
                print(f"DEBUG: Setting class_name and course_code to empty string")
        else:
            print(f"DEBUG: class_name not provided in request, not modifying it")
        
        print(f"DEBUG: Final update data prepared: {update_data}")
        
        result = mongodb_service.update_student(student_id, update_data)
        print(f"DEBUG: MongoDB result: {result}")
        
        return jsonify(result)
        
    except Exception as e:
        print(f"Error in update_student_credential: {e}")
        return jsonify({'success': False, 'message': f'Error updating student: {str(e)}'}), 500

@management_bp.route('/export_students_credentials')
@login_required('management')
def export_students_credentials():
    """Export students credentials to Excel with proper formatting"""
    try:
        result = mongodb_service.get_all_students()
        
        if not result['success']:
            return jsonify({'success': False, 'message': result['message']})
        
        students = result['data']
        
        # Create DataFrame with only add student fields
        export_data = []
        for student in students:
            export_data.append({
                'Roll Number': student.get('roll_number', ''),
                'Student Name': student.get('name', ''),
                'Class Name': student.get('class_name', ''),
                'Password': student.get('password', '')
            })
        
        df = pd.DataFrame(export_data)
        
        # Create Excel file in memory with proper formatting
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Students Credentials', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Students Credentials']
            
            # Define styles
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # White border for header row
            header_border = Border(
                left=Side(style='thin', color='FFFFFF'),
                right=Side(style='thin', color='FFFFFF'),
                top=Side(style='thin', color='FFFFFF'),
                bottom=Side(style='thin', color='FFFFFF')
            )
            
            data_font = Font(name='Calibri', size=11)
            data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Regular border for data rows
            data_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply header formatting with white borders
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border
            
            # Apply data formatting
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = data_border
            
            # Set column widths
            column_widths = {
                'A': 15,  # Roll Number
                'B': 30,  # Student Name (increased for long names)
                'C': 20,  # Class Name
                'D': 15   # Password
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Set row heights
            for row in range(1, worksheet.max_row + 1):
                worksheet.row_dimensions[row].height = 25
        
        output.seek(0)
        
        from flask import Response
        response = Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=students_credentials.xlsx'}
        )
        
        return response
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error exporting students: {str(e)}'})