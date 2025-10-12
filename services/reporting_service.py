"""
Reporting service for Moulya College Management System
Handles marks and attendance reports for management
"""

from models.student import Student
from models.academic import Course, Subject
from models.marks import StudentMarks
from models.attendance import AttendanceRecord, MonthlyAttendanceSummary
from models.user import Lecturer
from database import db
from datetime import datetime, date
from sqlalchemy import func, and_, or_
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from xml.sax.saxutils import escape as xml_escape
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from services.excel_export_service import ExcelExportService

class ReportingService:
    """Service for generating reports"""
    
    @staticmethod
    def _format_number(value):
        """Format number: whole numbers without decimals, fractional numbers with 2 decimal places."""
        try:
            if value is None:
                return None
            num = float(value)
            if num == int(num):
                return int(num)  # 32.0 -> 32
            else:
                return round(num, 2)  # 32.43 -> 32.43, 32.05 -> 32.05
        except (ValueError, TypeError):
            return value

    @staticmethod
    def _extract_year_from_class(class_name):
        """Extract year from class name like 'I BCA B' -> 1, 'II BCA B' -> 2, 'III BCA B' -> 3"""
        if not class_name:
            return None
        
        class_name = str(class_name).strip().upper()
        
        # Check for Roman numerals
        if class_name.startswith('I '):
            return 1
        elif class_name.startswith('II '):
            return 2
        elif class_name.startswith('III '):
            return 3
        elif class_name.startswith('IV '):
            return 4
        
        # Check for Arabic numerals
        if class_name.startswith('1 '):
            return 1
        elif class_name.startswith('2 '):
            return 2
        elif class_name.startswith('3 '):
            return 3
        elif class_name.startswith('4 '):
            return 4
        
        # Check for words
        if class_name.startswith('FIRST '):
            return 1
        elif class_name.startswith('SECOND '):
            return 2
        elif class_name.startswith('THIRD '):
            return 3
        elif class_name.startswith('FOURTH '):
            return 4
        
        return None
    
    @staticmethod
    def _get_paragraph_style():
        """Return a compact cell Paragraph style to enable auto word-wrap in table cells."""
        styles = getSampleStyleSheet()
        from reportlab.lib.styles import ParagraphStyle
        return ParagraphStyle(
            'Cell',
            parent=styles['Normal'],
            fontSize=9,
            leading=11,
            spaceAfter=0,
            spaceBefore=0,
        )

    @staticmethod
    def _get_header_paragraph_style():
        styles = getSampleStyleSheet()
        from reportlab.lib.styles import ParagraphStyle
        return ParagraphStyle(
            'HeaderCell',
            parent=styles['Normal'],
            fontSize=9,
            leading=11,
            textColor=colors.white,
            spaceAfter=0,
            spaceBefore=0,
        )

    @staticmethod
    def _to_paragraph(value):
        """Convert any value to a Paragraph so ReportLab wraps text within cell width."""
        try:
            if value is None:
                return Paragraph('', ReportingService._get_paragraph_style())
            # Keep numbers as plain strings too; Paragraph will handle fine
            # Escape text but preserve explicit line breaks by converting \n -> <br/>
            text = xml_escape(str(value)).replace('\n', '<br/>')
            return Paragraph(text, ReportingService._get_paragraph_style())
        except Exception:
            return Paragraph(xml_escape(str(value)), ReportingService._get_paragraph_style())

    @staticmethod
    def _wrap_table_data(rows, skip_header=True, header_text_white=False, no_wrap_cols=None):
        """Map table cells to Paragraphs for word-wrap.
        If skip_header=True, the first row is left as-is so TableStyle header
        text color/background rules still apply.
        """
        if not rows:
            return rows
        no_wrap_set = set(no_wrap_cols or [])
        start_idx = 1 if skip_header and len(rows) > 0 else 0
        wrapped_rows = []
        # Keep header as-is if requested
        if start_idx == 1:
            if header_text_white:
                wrapped_rows.append([Paragraph(xml_escape(str(c)), ReportingService._get_header_paragraph_style()) for c in rows[0]])
            else:
                wrapped_rows.append(rows[0])
        for row in rows[start_idx:]:
            wrapped = []
            for idx, cell in enumerate(row):
                if idx in no_wrap_set:
                    wrapped.append(xml_escape(str(cell)) if cell is not None else '')
                else:
                    wrapped.append(ReportingService._to_paragraph(cell))
            wrapped_rows.append(wrapped)
        return wrapped_rows
    
    @staticmethod
    def _full_width_colwidths(total_width, num_columns):
        """Return equal column widths that use the full available width.
        This keeps every table consistent and maximized to the page width.
        """
        if num_columns <= 0:
            return []
        col = total_width / float(num_columns)
        return [col] * num_columns

    @staticmethod
    def _parse_course_and_section(course_name: str):
        """Split a user-entered course string into course and section.
        Examples:
        - "III bca b" -> ("bca", "b")
        - "I bcom" -> ("bcom", None)
        - "bsc" -> ("bsc", None)
        We consider the last token as section only when there are at least 3 tokens.
        The course becomes the last token when there are 1-2 tokens, otherwise the second last.
        """
        if not course_name:
            return None, None
        parts = [p for p in course_name.strip().split(' ') if p]
        if not parts:
            return None, None
        if len(parts) >= 3:
            return parts[-2], parts[-1]
        # len == 1 or 2 -> course is last token, no section
        return parts[-1], None
    
    @staticmethod
    def get_student_detailed_report(student_id):
        """Get detailed report for a specific student"""
        try:
            student = Student.query.get(student_id)
            if not student:
                return None
            
            # Get all subjects the student is enrolled in
            from models.student import StudentEnrollment
            subjects = Subject.query.join(StudentEnrollment).filter(
                StudentEnrollment.student_id == student_id,
                StudentEnrollment.is_active == True
            ).all()
            
            course_display, section = ReportingService._parse_course_and_section(student.course.name if student.course else None)

            report = {
                'student': {
                    'id': student.id,
                    'name': student.name,
                    'roll_number': student.roll_number,
                    'course': student.course.name if student.course else None,
                    'course_display': course_display,
                    'section': section,
                    'academic_year': student.academic_year,
                    'current_semester': student.current_semester
                },
                'subjects': []
            }
            
            for subject in subjects:
                # Get marks for this subject
                from models.marks import StudentMarks
                marks = StudentMarks.query.filter_by(
                    student_id=student_id,
                    subject_id=subject.id
                ).all()
                
                # Get overall attendance for this subject - sum all monthly data
                from models.attendance import MonthlyStudentAttendance, MonthlyAttendanceSummary
                from datetime import datetime
                
                # Get all monthly attendance records for this student and subject
                monthly_attendance_records = MonthlyStudentAttendance.query.filter_by(
                    student_id=student_id,
                    subject_id=subject.id
                ).all()
                
                if monthly_attendance_records:
                    # Calculate cumulative attendance across all months
                    total_classes = 0
                    present_classes = 0
                    
                    for record in monthly_attendance_records:
                        # Get total classes for this month
                        monthly_summary = MonthlyAttendanceSummary.query.filter_by(
                            subject_id=subject.id,
                            month=record.month,
                            year=record.year
                        ).first()
                        
                        if monthly_summary:
                            total_classes += monthly_summary.total_classes
                            present_classes += record.present_count
                    
                    attendance_percentage = round((present_classes / total_classes) * 100, 2) if total_classes > 0 else 0
                else:
                    # Fallback to daily attendance records
                    attendance_records = AttendanceRecord.query.filter_by(
                        student_id=student_id,
                        subject_id=subject.id
                    ).all()
                    
                    total_classes = len(attendance_records)
                    present_classes = len([r for r in attendance_records if r.status == 'present'])
                    attendance_percentage = round((present_classes / total_classes) * 100, 2) if total_classes > 0 else 0
                
                # Calculate overall marks percentage for this subject
                if marks:
                    total_obtained = sum(mark.marks_obtained for mark in marks)
                    total_max = sum(mark.max_marks for mark in marks)
                    overall_percentage = round((total_obtained / total_max) * 100, 2) if total_max > 0 else 0
                else:
                    overall_percentage = 0
                
                # Format marks data with proper assessment type names
                marks_data = []
                for mark in marks:
                    mark_dict = mark.to_dict()
                    # Ensure assessment_type is properly formatted
                    assessment_type = mark.assessment_type
                    if assessment_type == 'internal1':
                        mark_dict['assessment_type'] = 'Internal 1'
                    elif assessment_type == 'internal2':
                        mark_dict['assessment_type'] = 'Internal 2'
                    elif assessment_type == 'assignment':
                        mark_dict['assessment_type'] = 'Assignment'
                    elif assessment_type == 'project':
                        mark_dict['assessment_type'] = 'Project'
                    else:
                        mark_dict['assessment_type'] = assessment_type.title()
                    
                    # Add percentage calculation
                    if mark.max_marks > 0:
                        mark_dict['percentage'] = round((mark.marks_obtained / mark.max_marks) * 100, 2)
                    else:
                        mark_dict['percentage'] = 0
                    
                    marks_data.append(mark_dict)
                
                subject_data = {
                    'subject_id': subject.id,
                    'subject_name': subject.name,
                    'subject_code': subject.code,
                    'year': subject.year,
                    'semester': subject.semester,
                    'marks': marks_data,
                    'overall_marks_percentage': overall_percentage,
                    'attendance': {
                        'total_classes': total_classes,
                        'present_classes': present_classes,
                        'absent_classes': total_classes - present_classes,
                        'attendance_percentage': attendance_percentage
                    }
                }
                
                report['subjects'].append(subject_data)
            
            return report
            
        except Exception as e:
            print(f"Error generating student report: {e}")
            return None
    
    @staticmethod
    def get_class_marks_report(subject_id, assessment_type=None):
        """Get marks report for entire class in a subject"""
        try:
            subject = Subject.query.get(subject_id)
            if not subject:
                # Gracefully return an empty report shell so UI can still render
                is_overall = (str(month).lower() == 'overall') if isinstance(month, str) or month is not None else False
                if not is_overall:
                    if not month:
                        month = datetime.now().month
                    if not year:
                        year = datetime.now().year
                month_name = 'Overall' if (is_overall) else (
                    ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December'][month]
                    if isinstance(month, int) and 1 <= month <= 12 else str(month)
                )
                return {
                    'subject': {
                        'id': subject_id,
                        'name': 'Subject',
                        'code': '',
                        'course': None,
                        'course_display': None,
                        'section': None,
                        'year': None,
                        'semester': None,
                        'lecturers': []
                    },
                    'month': month_name,
                    'year': year,
                    'statistics': {
                        'total_students': 0,
                        'total_classes_conducted': 0,
                        'class_average_attendance': 0,
                        'students_with_good_attendance': 0,
                        'students_with_poor_attendance': 0
                    },
                    'student_attendance': []
                }
            
            from models.marks import StudentMarks
            
            # Get all enrolled students for this subject
            enrolled_students = subject.get_enrolled_students()
            
            # Initialize student_marks with all enrolled students
            student_marks = {}
            for student in enrolled_students:
                student_marks[student.id] = {
                    'student': student,
                    'student_name': student.name,
                    'roll_number': student.roll_number,
                        'marks': []
                    }
                
            # Get marks - if assessment_type is specified, filter by it
            if assessment_type:
                # Get marks for the specific assessment type
                marks = StudentMarks.query.filter_by(subject_id=subject_id, assessment_type=assessment_type).all()
            else:
                # Get all marks for this subject
                marks = StudentMarks.query.filter_by(subject_id=subject_id).all()
            
            # Add marks to student_marks
            for mark in marks:
                student_id = mark.student_id
                if student_id in student_marks:
                    mark_dict = mark.to_dict()
                    # Format assessment type for display
                    assessment_type_display = mark.assessment_type
                    if assessment_type_display == 'internal1':
                        mark_dict['assessment_type'] = 'Internal 1'
                    elif assessment_type_display == 'internal2':
                        mark_dict['assessment_type'] = 'Internal 2'
                    elif assessment_type_display == 'assignment':
                        mark_dict['assessment_type'] = 'Assignment'
                    elif assessment_type_display == 'project':
                        mark_dict['assessment_type'] = 'Project'
                    else:
                        mark_dict['assessment_type'] = assessment_type_display.title()
                    
                    student_marks[student_id]['marks'].append(mark_dict)
            
            # Calculate statistics
            all_percentages = []
            for mark in marks:
                if mark.max_marks > 0:
                    percentage = round((mark.marks_obtained / mark.max_marks) * 100, 2)
                    all_percentages.append(percentage)
            
            # Calculate passing/failing assessments (individual assessments, not students)
            passing_assessments = len([p for p in all_percentages if p >= 35])
            failing_assessments = len([p for p in all_percentages if p < 35])
            
            # Calculate passing/failing students (unique students based on their average)
            passing_students_count = 0
            failing_students_count = 0
            
            for student_id, student_data in student_marks.items():
                # Get all percentages for this student
                student_percentages = []
                for mark in marks:
                    if mark.student_id == student_id and mark.max_marks > 0:
                        percentage = round((mark.marks_obtained / mark.max_marks) * 100, 2)
                        student_percentages.append(percentage)
                
                # A student passes if their overall average is >= 35%
                if student_percentages:
                    student_average = sum(student_percentages) / len(student_percentages)
                    if student_average >= 35:
                        passing_students_count += 1
                    else:
                        failing_students_count += 1
            
            statistics = {
                'total_students': len(student_marks),
                'total_assessments': len(marks),
                'class_average': round(sum(all_percentages) / len(all_percentages), 2) if all_percentages else 0,
                'highest_score': max(all_percentages) if all_percentages else 0,
                'lowest_score': min(all_percentages) if all_percentages else 0,
                'passing_students': passing_assessments,  # This will be used for Class Pass Rate
                'failing_students': failing_assessments
            }
            
            # Format assessment type for display
            assessment_type_display = assessment_type
            if assessment_type == 'internal1':
                assessment_type_display = 'Internal 1'
            elif assessment_type == 'internal2':
                assessment_type_display = 'Internal 2'
            elif assessment_type == 'assignment':
                assessment_type_display = 'Assignment'
            elif assessment_type == 'project':
                assessment_type_display = 'Project'
            elif assessment_type:
                assessment_type_display = assessment_type.title()
            
            course_display, section = ReportingService._parse_course_and_section(subject.course.name if subject.course else None)

            lecturers_list = []
            try:
                lecturers_list = [lec.name for lec in subject.get_assigned_lecturers() if lec]
            except Exception:
                lecturers_list = []

            report = {
                'subject': {
                    'id': subject.id,
                    'name': subject.name,
                    'code': subject.code,
                    'course': subject.course.name if subject.course else None,
                    'course_display': course_display,
                    'section': section,
                    'year': subject.year,
                    'semester': subject.semester,
                    'lecturers': lecturers_list
                },
                'assessment_type': assessment_type_display,
                'statistics': statistics,
                'student_marks': list(student_marks.values())
            }
            
            return report
            
        except Exception as e:
            print(f"Error generating class marks report: {e}")
            return None
    
    @staticmethod
    def get_class_attendance_report(subject_id, month=None, year=None):
        """Get attendance report for entire class in a subject"""
        try:
            print(f"[DEBUG] Service called with subject_id={subject_id}, month={month}, year={year}")
            
            subject = Subject.query.get(subject_id)
            print(f"[DEBUG] Subject found: {subject is not None}")
            if subject:
                print(f"[DEBUG] Subject name: {subject.name}, code: {subject.code}")
            
            if not subject:
                print(f"[DEBUG] Subject not found, returning empty report structure")
                # Gracefully return an empty report shell so UI can still render
                is_overall = (str(month).lower() == 'overall') if isinstance(month, str) or month is not None else False
                if not is_overall:
                    if not month:
                        month = datetime.now().month
                    if not year:
                        year = datetime.now().year
                month_name = 'Overall' if (is_overall) else (
                    ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December'][month]
                    if isinstance(month, int) and 1 <= month <= 12 else str(month)
                )
                empty_report = {
                    'subject': {
                        'id': subject_id,
                        'name': 'Subject',
                        'code': '',
                        'course': None,
                        'course_display': None,
                        'section': None,
                        'year': None,
                        'semester': None,
                        'lecturers': []
                    },
                    'month': month_name,
                    'year': year,
                    'statistics': {
                        'total_students': 0,
                        'total_classes_conducted': 0,
                        'class_average_attendance': 0,
                        'students_with_good_attendance': 0,
                        'students_with_poor_attendance': 0
                    },
                    'student_attendance': []
                }
                print(f"[DEBUG] Returning empty report with month_name: {month_name}")
                return empty_report
            
            # If month is 'overall', compute cumulative across all months/years
            is_overall = (str(month).lower() == 'overall') if month is not None else False
            # For specific month requests coming as strings (e.g., '10'), coerce to int
            if not is_overall and month is not None and not isinstance(month, int):
                try:
                    month = int(month)
                except Exception:
                    month = None
            # Default to current month/year if not specified and not overall
            if not is_overall:
                if not month:
                    month = datetime.now().month
                if not year:
                    year = datetime.now().year
            
            # Get all students enrolled in this subject
            from models.student import StudentEnrollment
            students = Student.query.join(StudentEnrollment).filter(
                StudentEnrollment.subject_id == subject_id,
                StudentEnrollment.is_active == True
            ).all()
            
            student_attendance = []
            total_classes_conducted = 0
            
            # First, try to get data from monthly_attendance_summary table
            from models.attendance import MonthlyAttendanceSummary, MonthlyStudentAttendance
            if is_overall:
                monthly_summary = None
            else:
                monthly_summary = MonthlyAttendanceSummary.query.filter_by(
                    subject_id=subject_id,
                    month=month,
                    year=year
                ).first()
            
            if monthly_summary:
                total_classes_conducted = monthly_summary.total_classes
                
                # Get data from monthly_student_attendance table
                from models.attendance import MonthlyStudentAttendance
                monthly_attendance_records = MonthlyStudentAttendance.query.filter_by(
                    subject_id=subject_id,
                    month=month,
                    year=year
                ).all()
                
                # Create a mapping of student_id to present_count
                attendance_map = {record.student_id: record.present_count for record in monthly_attendance_records}
                
                for student in students:
                    present_classes = attendance_map.get(student.id, 0)
                    absent_classes = max(total_classes_conducted - present_classes, 0)
                    
                    # Calculate percentage
                    attendance_percentage = round((present_classes / total_classes_conducted) * 100, 2) if total_classes_conducted > 0 else 0
                    
                    student_data = {
                        'student_id': student.id,
                        'student_name': student.name,
                        'roll_number': student.roll_number,
                        'total_classes': total_classes_conducted,
                        'present_classes': present_classes,
                        'absent_classes': absent_classes,
                        'attendance_percentage': attendance_percentage,
                        'status': 'Good' if attendance_percentage >= 75 else 'Poor' if attendance_percentage < 50 else 'Average'
                    }
                    
                    student_attendance.append(student_data)
            else:
                # If overall, compute from monthly aggregates; otherwise, fallback to daily records for the specific month
                if is_overall:
                    # Compute using MonthlyStudentAttendance + MonthlyAttendanceSummary across all months/years
                    from models.attendance import MonthlyAttendanceSummary, MonthlyStudentAttendance
                    summaries = MonthlyAttendanceSummary.query.filter_by(subject_id=subject_id).all()
                    total_classes_conducted = sum(s.total_classes for s in summaries)
                    # Map student -> present (including deputation)
                    student_present_map = {s.id: 0 for s in students}
                    monthly_attendance_records = MonthlyStudentAttendance.query.filter_by(subject_id=subject_id).all()
                    for rec in monthly_attendance_records:
                        student_present_map[rec.student_id] = student_present_map.get(rec.student_id, 0) + int(rec.present_count or 0) + int(rec.deputation_count or 0)
                    for student in students:
                        present_classes = student_present_map.get(student.id, 0)
                        absent_classes = max(total_classes_conducted - present_classes, 0)
                        attendance_percentage = round((present_classes / total_classes_conducted) * 100, 2) if total_classes_conducted > 0 else 0
                        student_attendance.append({
                            'student_id': student.id,
                            'student_name': student.name,
                            'roll_number': student.roll_number,
                            'total_classes': total_classes_conducted,
                            'present_classes': present_classes,
                            'absent_classes': absent_classes,
                            'attendance_percentage': attendance_percentage,
                            'status': 'Good' if attendance_percentage >= 75 else 'Poor' if attendance_percentage < 50 else 'Average'
                        })
                else:
                    # Daily records for a specific month/year
                    all_attendance_records = AttendanceRecord.query.filter(
                        AttendanceRecord.subject_id == subject_id,
                        db.extract('month', AttendanceRecord.date) == month,
                        db.extract('year', AttendanceRecord.date) == year
                    ).all()

                    # Get unique dates to count total classes
                    unique_dates = set(record.date for record in all_attendance_records)
                    total_classes_conducted = len(unique_dates)

                    # If no classes were conducted, still show students with 0 attendance
                    if total_classes_conducted == 0:
                        for student in students:
                            student_data = {
                                'student_id': student.id,
                                'student_name': student.name,
                                'roll_number': student.roll_number,
                                'total_classes': 0,
                                'present_classes': 0,
                                'absent_classes': 0,
                                'attendance_percentage': 0,
                                'status': 'Poor'
                            }
                            student_attendance.append(student_data)
                    else:
                        # Calculate attendance for each student for the specific month
                        for student in students:
                            # Get attendance records for this student in this subject for this month
                            student_records = AttendanceRecord.query.filter(
                                AttendanceRecord.student_id == student.id,
                                AttendanceRecord.subject_id == subject_id,
                                db.extract('month', AttendanceRecord.date) == month,
                                db.extract('year', AttendanceRecord.date) == year
                            ).all()

                            present_classes = len([r for r in student_records if r.status == 'present'])
                            absent_classes = len([r for r in student_records if r.status == 'absent'])

                            # Calculate percentage
                            attendance_percentage = round((present_classes / total_classes_conducted) * 100, 2) if total_classes_conducted > 0 else 0

                            student_data = {
                                'student_id': student.id,
                                'student_name': student.name,
                                'roll_number': student.roll_number,
                                'total_classes': total_classes_conducted,
                                'present_classes': present_classes,
                                'absent_classes': absent_classes,
                                'attendance_percentage': attendance_percentage,
                                'status': 'Good' if attendance_percentage >= 75 else 'Poor' if attendance_percentage < 50 else 'Average'
                            }

                            student_attendance.append(student_data)
            
            # Sort students by last 3 digits of roll number
            def get_roll_sort_key(student):
                roll_number = student.get('roll_number', '')
                # Extract last 3 digits from roll number (e.g., BCA25001 -> 001)
                if len(roll_number) >= 3:
                    last_three = roll_number[-3:]
                    try:
                        return int(last_three)
                    except ValueError:
                        return 999  # Put non-numeric at end
                return 999  # Put short roll numbers at end
            
            student_attendance.sort(key=get_roll_sort_key)
            
            # Calculate class statistics
            all_percentages = [s['attendance_percentage'] for s in student_attendance]
            valid_percentages = [p for p in all_percentages if p > 0]
            
            statistics = {
                'total_students': len(students),
                'total_classes_conducted': total_classes_conducted,
                'class_average_attendance': round(sum(valid_percentages) / len(valid_percentages), 2) if valid_percentages else 0,
                'students_with_good_attendance': len([s for s in student_attendance if s['attendance_percentage'] >= 75]),
                'students_with_poor_attendance': len([s for s in student_attendance if s['attendance_percentage'] < 50])
            }
            
            # Get month name for display
            if is_overall:
                month_name = 'Overall'
            else:
                month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                              'July', 'August', 'September', 'October', 'November', 'December']
                try:
                    month_name = month_names[int(month)] if 1 <= int(month) <= 12 else str(month)
                except Exception:
                    month_name = str(month)
            
            course_display, section = ReportingService._parse_course_and_section(subject.course.name if subject.course else None)

            lecturers_list = []
            try:
                lecturers_list = [lec.name for lec in subject.get_assigned_lecturers() if lec]
            except Exception:
                lecturers_list = []

            report = {
                'subject': {
                    'id': subject.id,
                    'name': subject.name,
                    'code': subject.code,
                    'course': subject.course.name if subject.course else None,
                    'course_display': course_display,
                    'section': section,
                    'year': subject.year,
                    'semester': subject.semester,
                    'lecturers': lecturers_list
                },
                'month': month_name,
                'month_num': month if not is_overall else None,
                'year': year,
                'statistics': statistics,
                'student_attendance': student_attendance
            }
            
            print(f"[DEBUG] Returning valid report with {len(student_attendance)} students")
            return report
            
        except Exception as e:
            print(f"[DEBUG] Exception in service: {str(e)}")
            print(f"Error generating class attendance report: {e}")
            return None
    
    @staticmethod
    def get_course_overview_report(course_id):
        """Get overview report for entire course"""
        try:
            course = Course.query.get(course_id)
            if not course:
                return None
            
            # Get all subjects in this course
            subjects = Subject.query.filter_by(course_id=course_id, is_active=True).all()
            
            # Get all students in this course
            students = Student.query.filter_by(course_id=course_id, is_active=True).all()
            
            subject_reports = []
            for subject in subjects:
                # Get marks statistics
                marks = StudentMarks.query.filter_by(subject_id=subject.id).all()
                marks_percentages = [mark.percentage for mark in marks]
                
                # Calculate passing rate correctly for unique students
                passing_students_count = 0
                total_students_with_marks = 0
                
                # Group marks by student to calculate per-student averages
                student_marks_dict = {}
                for mark in marks:
                    if mark.student_id not in student_marks_dict:
                        student_marks_dict[mark.student_id] = []
                    student_marks_dict[mark.student_id].append(mark.percentage)
                
                # Calculate passing rate based on student averages
                for student_id, student_percentages in student_marks_dict.items():
                    if student_percentages:
                        student_average = sum(student_percentages) / len(student_percentages)
                        total_students_with_marks += 1
                        if student_average >= 35:
                            passing_students_count += 1
                
                passing_rate = round((passing_students_count / total_students_with_marks) * 100, 2) if total_students_with_marks > 0 else 0
                
                # Get attendance statistics - use monthly data first, then fallback to daily records
                from models.attendance import MonthlyStudentAttendance, MonthlyAttendanceSummary
                
                # Try to get monthly attendance data
                monthly_attendance_records = MonthlyStudentAttendance.query.filter_by(subject_id=subject.id).all()
                
                if monthly_attendance_records:
                    # Calculate cumulative attendance across all months
                    total_classes = 0
                    total_present = 0
                    
                    for record in monthly_attendance_records:
                        # Get total classes for this month
                        monthly_summary = MonthlyAttendanceSummary.query.filter_by(
                            subject_id=subject.id,
                            month=record.month,
                            year=record.year
                        ).first()
                        
                        if monthly_summary:
                            total_classes += monthly_summary.total_classes
                            total_present += record.present_count
                    
                    total_records = total_classes
                    present_records = total_present
                else:
                    # Fallback to daily attendance records
                    attendance_records = AttendanceRecord.query.filter_by(subject_id=subject.id).all()
                    total_records = len(attendance_records)
                    present_records = len([r for r in attendance_records if r.status == 'present'])
                
                subject_data = {
                    'subject_id': subject.id,
                    'subject_name': subject.name,
                    'subject_code': subject.code,
                    'year': subject.year,
                    'semester': subject.semester,
                    'enrolled_students': subject.get_enrolled_students_count(),
                    'marks_statistics': {
                        'total_assessments': len(marks),
                        'average_marks': round(sum(marks_percentages) / len(marks_percentages), 2) if marks_percentages else 0,
                        'passing_rate': passing_rate
                    },
                    'attendance_statistics': {
                        'total_classes': total_records,
                        'average_attendance': round((present_records / total_records) * 100, 2) if total_records > 0 else 0
                    }
                }
                
                subject_reports.append(subject_data)
            
            report = {
                'course': {
                    'id': course.id,
                    'name': course.name,
                    'code': course.code,
                    'duration_years': course.duration_years,
                    'total_semesters': course.total_semesters
                },
                'total_students': len(students),
                'total_subjects': len(subjects),
                'subjects': subject_reports
            }
            
            return report
            
        except Exception as e:
            print(f"Error generating course overview report: {e}")
            return None
    
    @staticmethod
    def get_subjects_for_reporting():
        """Get all subjects available for reporting"""
        try:
            subjects = Subject.query.filter_by(is_active=True).all()
            return [{
                'id': subject.id,
                'name': subject.name,
                'code': subject.code,
                'course_name': subject.course.name if subject.course else None,
                'year': subject.year,
                'semester': subject.semester,
                'enrolled_students_count': subject.get_enrolled_students_count()
            } for subject in subjects]
        except Exception as e:
            print(f"Error getting subjects for reporting: {e}")
            return []
    
    @staticmethod
    def get_courses_for_reporting():
        """Get all courses available for reporting"""
        try:
            courses = Course.query.filter_by(is_active=True).all()
            return [{
                'id': course.id,
                'name': course.name,
                'code': course.code,
                'total_students': course.get_active_students_count(),
                'total_subjects': course.subjects.filter_by(is_active=True).count()
            } for course in courses]
        except Exception as e:
            print(f"Error getting courses for reporting: {e}")
            return []
    
    @staticmethod
    def get_students_for_reporting(course_id=None):
        """Get all students available for reporting, optionally filtered by course"""
        try:
            query = Student.query.filter_by(is_active=True)
            if course_id:
                query = query.filter_by(course_id=course_id)
            
            students = query.all()
            return [{
                'id': student.id,
                'name': student.name,
                'roll_number': student.roll_number,
                'course_name': student.course.name if student.course else None,
                'course_id': student.course_id,
                'academic_year': student.academic_year,
                'current_semester': student.current_semester
            } for student in students]
        except Exception as e:
            print(f"Error getting students for reporting: {e}")
            return []

    # ======================== PDF GENERATION ========================
    @staticmethod
    def generate_student_report_pdf(report):
        """Generate a PDF for the student detailed report and return bytes."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm, topMargin=18*mm, bottomMargin=18*mm)
        elements = []
        styles = getSampleStyleSheet()
        from reportlab.lib.styles import ParagraphStyle
        title_center = ParagraphStyle('TitleCenter', parent=styles['Title'], alignment=1)
        subtitle_center = ParagraphStyle('SubtitleCenter', parent=styles['Normal'], alignment=1)
        header_title = ParagraphStyle('HeaderTitle', parent=styles['Title'], alignment=0, fontSize=16, leading=19)
        header_sub = ParagraphStyle('HeaderSub', parent=styles['Normal'], alignment=0, fontSize=10, leading=12)

        # Header with logo and college name
        # College-style header (logo + text, underline)
        try:
            from flask import current_app
            logo_path = current_app.root_path + '/static/img/logo-removebg-preview.png'
            logo_img = Image(logo_path)
            logo_img._restrictSize(26*mm, 26*mm)
        except Exception:
            logo_img = ''
        header_text = [
            Paragraph('Dr. B. B. Hegde First Grade College, Kundapura', header_title),
            Paragraph('A Unit of Coondapur Education Society (R)', header_sub)
        ]
        header_table = Table([[logo_img, header_text]], colWidths=[26*mm, (A4[0] - (18*mm + 18*mm) - 26*mm)])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LINEBELOW', (0,0), (-1,0), 0.75, colors.lightgrey),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        elements.append(header_table)

        # Main report title
        elements.append(Spacer(1, 6))
        elements.append(Paragraph('Student Performance Report', title_center))
        elements.append(Spacer(1, 8))

        # Student info table
        s = report['student']
        course_line = s.get('course_display') or s.get('course') or ''
        data = [
            ['Name', s.get('name', '')],
            ['Roll Number', s.get('roll_number', '')],
            ['Course', course_line],
        ]
        if s.get('section'):
            data.append(['Section', str(s.get('section')).upper()])
        data.extend([
            ['Academic Year', s.get('academic_year', '')],
            ['Current Semester', s.get('current_semester', '')]
        ])

        table = Table(data, colWidths=[45*mm, 115*mm])
        table.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
            ('INNERGRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
            # no header row now
            ('ALIGN', (0,0), (0,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        elements.extend([Spacer(1, 6), table, Spacer(1, 12)])

        # Marks table
        elements.append(Paragraph('Marks Report', styles['Heading2']))
        # Removed Percent and Grade columns for a cleaner layout
        marks_headers = ['Subject', 'Code', 'Assessment', 'Marks', 'Max', 'Status']
        marks_rows = [marks_headers]
        for subj in report.get('subjects', []):
            for m in subj.get('marks', []):
                marks_rows.append([
                    subj.get('subject_name',''), (subj.get('subject_code','') or '').replace(' ', '\u00A0'), m.get('assessment_type',''),
                    ReportingService._format_number(m.get('marks_obtained')), ReportingService._format_number(m.get('max_marks')),
                    m.get('performance_status','')
                ])
        if len(marks_rows) == 1:
            marks_rows.append(['No data'] + ['']*5)
        # Build standardized marks table with consistent font size
        page_width = A4[0] - (18*mm + 18*mm)
        # Make Subject narrower and widen Status so it fits on one line
        marks_col_fracs = [0.34, 0.18, 0.16, 0.10, 0.10, 0.12]
        marks_table = ReportingService._build_table(
            marks_rows,
            page_width,
            marks_col_fracs,
            no_wrap_cols={1},          # Code must not wrap
            center_cols={3,4},         # Marks and Max
            header_bg=colors.black,
        )
        elements.extend([marks_table, Spacer(1, 12)])

        # Attendance table
        elements.append(Paragraph('Attendance Report', styles['Heading2']))
        # Removed Absent column for a cleaner layout
        att_headers = ['Subject', 'Code', 'Total', 'Present', 'Percent', 'Status']
        att_rows = [att_headers]
        for subj in report.get('subjects', []):
            a = subj.get('attendance', {})
            att_rows.append([
                subj.get('subject_name',''), (subj.get('subject_code','') or '').replace(' ', '\u00A0'), ReportingService._format_number(a.get('total_classes')),
                ReportingService._format_number(a.get('present_classes')), ReportingService._format_number(a.get('attendance_percentage')),
                'Good' if a.get('attendance_percentage',0) >= 75 else 'Average' if a.get('attendance_percentage',0) >= 50 else 'Poor'
            ])
        if len(att_rows) == 1:
            att_rows.append(['No data'] + ['']*5)
        # Wrap text in table data
        # Prevent wrapping in Code column (index 1)
        att_rows_wrapped = ReportingService._wrap_table_data(att_rows, skip_header=True, header_text_white=True, no_wrap_cols={1})
        # Build standardized attendance table
        # Columns: Subject, Code, Total, Present, Percent, Status
        # Make Subject narrower and give Status more width to keep it on one line
        att_col_fracs = [0.34, 0.18, 0.12, 0.12, 0.12, 0.12]
        att_table = ReportingService._build_table(
            att_rows,
            page_width,
            att_col_fracs,
            no_wrap_cols={1},              # Code
            center_cols={2,3,4},
            header_bg=colors.black,
        )
        elements.append(att_table)

        doc.build(elements)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        return pdf_bytes

    # ======================== LECTURER PDFS ========================
    @staticmethod
    def generate_subject_marks_report_pdf(subject, marks_report):
        """Generate a PDF for a subject's marks report (lecturer view)."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm, topMargin=18*mm, bottomMargin=18*mm)
        elements = []
        styles = getSampleStyleSheet()
        from reportlab.lib.styles import ParagraphStyle
        header_title = ParagraphStyle('HeaderTitle', parent=styles['Title'], alignment=0, fontSize=16, leading=19)
        header_sub = ParagraphStyle('HeaderSub', parent=styles['Normal'], alignment=0, fontSize=10, leading=12)

        # Header (logo + college text)
        try:
            from flask import current_app
            logo_path = current_app.root_path + '/static/img/logo-removebg-preview.png'
            logo_img = Image(logo_path)
            logo_img._restrictSize(26*mm, 26*mm)
        except Exception:
            logo_img = ''
        header_text = [
            Paragraph('Dr. B. B. Hegde First Grade College, Kundapura', header_title),
            Paragraph('A Unit of Coondapur Education Society (R)', header_sub)
        ]
        header_table = Table([[logo_img, header_text]], colWidths=[26*mm, (A4[0] - (18*mm + 18*mm) - 26*mm)])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LINEBELOW', (0,0), (-1,0), 0.75, colors.lightgrey),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        elements.append(header_table)

        # Get faculty name
        from models.assignments import SubjectAssignment
        assignment = SubjectAssignment.query.filter_by(subject_id=subject.id, is_active=True).first()
        faculty_name = assignment.lecturer.name if assignment and assignment.lecturer else 'N/A'
        
        # Parse course name to extract course code and section
        course_name = subject.course.name if subject.course else 'N/A'
        course_parts = course_name.split() if course_name != 'N/A' else []
        
        # Determine if section exists (if there are 3+ parts, last part is section)
        has_section = len(course_parts) >= 3
        course_code = course_parts[-2] if has_section else (course_parts[-1] if course_parts else 'N/A')
        section = course_parts[-1] if has_section else None
        
        # Subject summary box
        subj_rows = [
            ['Subject', subject.name],
            ['Code', subject.code],
            ['Course', course_code]
        ]
        if has_section:
            subj_rows.append(['Section', section])
        subj_rows.append(['Faculty', faculty_name])
        # Year/Semester from subject
        try:
            ys_display = f"{getattr(subject, 'year', 'N/A')}/{getattr(subject, 'semester', 'N/A')}"
        except Exception:
            ys_display = "N/A/N/A"
        subj_rows.append(['Year/Semester', ys_display])
        
        subj_table = Table(subj_rows, colWidths=[35*mm, (A4[0] - (18*mm + 18*mm) - 35*mm)])
        subj_table.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
            ('INNERGRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        elements.extend([Spacer(1, 10), Paragraph('Marks Report', styles['Heading2']), Spacer(1, 6), subj_table, Spacer(1, 10)])

        # Marks table
        def _fmt_mark_pair(obt, mx):
            try:
                if obt is None or mx is None:
                    return ''
                fo = ReportingService._format_number(obt)
                fm = ReportingService._format_number(mx)
                return f"{fo}/{fm}"
            except Exception:
                try:
                    return f"{obt}/{mx}"
                except Exception:
                    return ''

        # Decide which assessment components to include based on actual recorded values
        comp_keys = ['internal1', 'internal2', 'assignment', 'project']
        include = {k: False for k in comp_keys}
        for record in (marks_report or []):
            ms = record.get('marks_summary', {})
            for k in comp_keys:
                a = getattr(ms, k, None) if hasattr(ms, k) else (ms.get(k) if isinstance(ms, dict) else None)
                if not a:
                    continue
                obt = getattr(a, 'obtained', None) if hasattr(a, 'obtained') else (a.get('obtained') if isinstance(a, dict) else None)
                mx = getattr(a, 'max', None) if hasattr(a, 'max') else (a.get('max') if isinstance(a, dict) else None)
                try:
                    obt_num = float(obt or 0)
                    mx_num = float(mx or 0)
                    if obt_num > 0 or mx_num > 0:
                        include[k] = True
                except Exception:
                    pass

        ordered_components = [k for k in comp_keys if include[k]]
        comp_to_header = {'internal1':'Internal 1','internal2':'Internal 2','assignment':'Assignment','project':'Project'}

        header = ['Student'] + [comp_to_header[k] for k in ordered_components] + ['Overall %', 'Status']
        rows = [header]
        for record in marks_report or []:
            student = record['student']
            ms = record.get('marks_summary', {})
            def _cell(assess):
                a = getattr(ms, assess, None) if hasattr(ms, assess) else (ms.get(assess) if isinstance(ms, dict) else None)
                if not a:
                    return ''
                obtained = getattr(a, 'obtained', None) if hasattr(a, 'obtained') else a.get('obtained') if isinstance(a, dict) else None
                max_marks = getattr(a, 'max', None) if hasattr(a, 'max') else a.get('max') if isinstance(a, dict) else None
                if obtained in (None, 0, '0', '0.0') and max_marks in (None, 0, '0', '0.0'):
                    return ''
                return _fmt_mark_pair(obtained, max_marks)
            # Build row with only included components
            overall = record.get('overall_percentage') or 0
            status = 'Good' if overall >= 50 else 'Deficient'
            # Combine name and roll in one cell (two lines)
            combined_student = f"{student.name}\n{getattr(student, 'roll_number', '') or ''}".strip()
            row = [combined_student]
            for k in ordered_components:
                row.append(_cell(k))
            row.extend([f"{ReportingService._format_number(overall)}%", status])
            rows.append(row)
        if len(rows) == 1:
            rows.append(['No data', '', '', '', '', '', ''])
        # Wrap text in table data
        rows_wrapped = ReportingService._wrap_table_data(rows, skip_header=True, header_text_white=True)

        page_width = A4[0] - (18*mm + 18*mm)
        table = Table(rows_wrapped, repeatRows=1, colWidths=ReportingService._full_width_colwidths(page_width, len(rows[0])))
        table.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
            ('INNERGRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
            ('BACKGROUND', (0,0), (-1,0), colors.black),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold')
        ]))
        elements.append(table)

        doc.build(elements)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        return pdf_bytes

    @staticmethod
    def generate_subject_attendance_report_pdf(subject, attendance_report):
        """Generate a PDF for a subject's attendance report (lecturer view)."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm, topMargin=18*mm, bottomMargin=18*mm)
        elements = []
        styles = getSampleStyleSheet()
        from reportlab.lib.styles import ParagraphStyle
        header_title = ParagraphStyle('HeaderTitle', parent=styles['Title'], alignment=0, fontSize=16, leading=19)
        header_sub = ParagraphStyle('HeaderSub', parent=styles['Normal'], alignment=0, fontSize=10, leading=12)

        try:
            from flask import current_app
            logo_path = current_app.root_path + '/static/img/logo-removebg-preview.png'
            logo_img = Image(logo_path)
            logo_img._restrictSize(26*mm, 26*mm)
        except Exception:
            logo_img = ''
        header_text = [
            Paragraph('Dr. B. B. Hegde First Grade College, Kundapura', header_title),
            Paragraph('A Unit of Coondapur Education Society (R)', header_sub)
        ]
        header_table = Table([[logo_img, header_text]], colWidths=[26*mm, 148*mm])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LINEBELOW', (0,0), (-1,0), 0.75, colors.lightgrey),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        elements.append(header_table)

        # Get faculty name
        from models.assignments import SubjectAssignment
        assignment = SubjectAssignment.query.filter_by(subject_id=subject.id, is_active=True).first()
        faculty_name = assignment.lecturer.name if assignment and assignment.lecturer else 'N/A'
        
        # Parse course name to extract course code and section
        course_name = subject.course.name if subject.course else 'N/A'
        course_parts = course_name.split() if course_name != 'N/A' else []
        
        # Determine if section exists (if there are 3+ parts, last part is section)
        has_section = len(course_parts) >= 3
        course_code = course_parts[-2] if has_section else (course_parts[-1] if course_parts else 'N/A')
        section = course_parts[-1] if has_section else None
        
        # Subject summary box
        subj_rows = [
            ['Subject', subject.name],
            ['Code', subject.code],
            ['Course', course_code]
        ]
        if has_section:
            subj_rows.append(['Section', section])
        subj_rows.append(['Faculty', faculty_name])
        # Year/Semester directly from subject
        ys_display = f"{getattr(subject, 'year', 'N/A')}/{getattr(subject, 'semester', 'N/A')}"
        subj_rows.append(['Year/Semester', ys_display])
        
        subj_table = Table(subj_rows, colWidths=[35*mm, (A4[0] - (18*mm + 18*mm) - 35*mm)])
        subj_table.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
            ('INNERGRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        elements.extend([Spacer(1, 10), Paragraph('Attendance Report', styles['Heading2']), Spacer(1, 6), subj_table, Spacer(1, 10)])

        # Attendance table (Student | Present | Total | % | Status)
        rows = [['Student', 'Roll Number', 'Present', 'Total', '%', 'Status']]
        for record in attendance_report or []:
            student = record['student']
            percent = record.get('attendance_percentage') or 0
            status = 'Good' if percent >= 75 else 'Shortage'
            rows.append([
                student.name,
                student.roll_number,
                record.get('present_classes') or 0,
                record.get('total_classes') or 0,
                f"{ReportingService._format_number(percent)}%",
                status
            ])
        if len(rows) == 1:
            rows.append(['No data', '', '', '', '', ''])
        # Wrap text in table data
        rows_wrapped = ReportingService._wrap_table_data(rows, skip_header=True, header_text_white=True)

        page_width = A4[0] - (18*mm + 18*mm)
        table = Table(rows_wrapped, repeatRows=1, colWidths=ReportingService._full_width_colwidths(page_width, len(rows[0])))
        table.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
            ('INNERGRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
            ('BACKGROUND', (0,0), (-1,0), colors.black),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold')
        ]))
        elements.append(table)

        doc.build(elements)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        return pdf_bytes

    # ======================== EXCEL EXPORT FUNCTIONS ========================
    @staticmethod
    def generate_subject_marks_report_excel(subject, marks_report):
        """Generate an Excel file for a subject's marks report (lecturer view)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Marks Report"
        
        # Get faculty name
        from models.assignments import SubjectAssignment
        assignment = SubjectAssignment.query.filter_by(subject_id=subject.id, is_active=True).first()
        faculty_name = assignment.lecturer.name if assignment and assignment.lecturer else 'N/A'
        
        # Get students and marks data using the same logic as HTML view
        from models.student import Student
        from models.marks import StudentMarks
        
        # Get enrolled students
        students = subject.get_enrolled_students()
        
        # Get existing marks for students (same as HTML view)
        existing_marks = {}
        for student in students:
            marks = StudentMarks.query.filter_by(
                student_id=student.id,
                subject_id=subject.id
            ).all()
            # map assessment_type -> mark row (same as HTML view)
            existing_marks[student.id] = {mark.assessment_type: mark for mark in marks}
        
        # Parse course name to extract course code and section
        course_name = subject.course.name if subject.course else 'N/A'
        course_parts = course_name.split() if course_name != 'N/A' else []
        
        # Determine if section exists (if there are 3+ parts, last part is section)
        has_section = len(course_parts) >= 3
        course_code = course_parts[-2] if has_section else (course_parts[-1] if course_parts else 'N/A')
        section = course_parts[-1] if has_section else None
        
        # Title
        ws['A1'] = 'Marks Report'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Subject info
        ws['A3'] = 'Subject'
        ws['B3'] = subject.name
        ws['A4'] = 'Code'
        ws['B4'] = subject.code
        ws['A5'] = 'Course'
        ws['B5'] = course_code
        if has_section:
            ws['A6'] = 'Section'
            ws['B6'] = section
            ws['A7'] = 'Faculty'
            ws['B7'] = faculty_name
            # Year/Semester directly from subject
            ys_display = f"{getattr(subject, 'year', 'N/A')}/{getattr(subject, 'semester', 'N/A')}"
            ws['A8'] = 'Year/Semester'
            ws['B8'] = ys_display
        else:
            ws['A6'] = 'Faculty'
            ws['B6'] = faculty_name
            ys_display = f"{getattr(subject, 'year', 'N/A')}/{getattr(subject, 'semester', 'N/A')}"
            ws['A7'] = 'Year/Semester'
            ws['B7'] = ys_display
        # Determine where to place spacer row and headers
        last_info_row = 8 if has_section else 7
        spacer_row = last_info_row + 1
        header_row = spacer_row + 1
        data_start_row = header_row + 1

        # Spacer row above the table
        ws.merge_cells(start_row=spacer_row, start_column=1, end_row=spacer_row, end_column=4)
        spacer_cell = ws.cell(row=spacer_row, column=1, value='')
        spacer_cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

        # Decide which assessment components to include based on actual recorded values
        comp_keys = ['internal1', 'internal2', 'assignment', 'project']
        include = {k: False for k in comp_keys}
        
        # Check which components have marks recorded (same logic as HTML view)
        for student in students:
            student_marks = existing_marks.get(student.id, {})
            for k in comp_keys:
                if k in student_marks and student_marks[k].max_marks > 0:
                    include[k] = True
        ordered_components = [k for k in comp_keys if include[k]]
        comp_to_header = {'internal1': 'Internal 1', 'internal2': 'Internal 2', 'assignment': 'Assignment', 'project': 'Project'}

        # Table headers
        headers = ['Student', 'Roll Number'] + [comp_to_header[k] for k in ordered_components] + ['Overall %', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        row = data_start_row
        for student in students:
            student_marks = existing_marks.get(student.id, {})
            
            def _pair(assess):
                if assess in student_marks:
                    mark = student_marks[assess]
                    obtained = mark.marks_obtained
                    max_marks = mark.max_marks
                    
                    if max_marks > 0:
                        fo_formatted = ExcelExportService.format_number(obtained)
                        fm_formatted = ExcelExportService.format_number(max_marks)
                        return f"{fo_formatted}/{fm_formatted}"
                return ''
            
            # Calculate overall percentage (same logic as HTML view)
            if student_marks:
                total_obtained = sum(mark.marks_obtained for mark in student_marks.values())
                total_max = sum(mark.max_marks for mark in student_marks.values())
                overall = round((total_obtained / total_max) * 100, 2) if total_max > 0 else 0.0
            else:
                overall = 0.0
            
            status = 'Good' if overall >= 50 else 'Deficient'

            col = 1
            ws.cell(row=row, column=col, value=student.name); col += 1
            ws.cell(row=row, column=col, value=student.roll_number); col += 1
            for k in ordered_components:
                ws.cell(row=row, column=col, value=_pair(k)); col += 1
            ws.cell(row=row, column=col, value=f"{ReportingService._format_number(overall)}%"); col += 1
            ws.cell(row=row, column=col, value=status)
            row += 1
        
        if not students:
            ws.cell(row=data_start_row, column=1, value='No data')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def generate_subject_attendance_report_excel(subject, attendance_report):
        """Generate an Excel file for a subject's attendance report (lecturer view)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Get faculty name
        from models.assignments import SubjectAssignment
        assignment = SubjectAssignment.query.filter_by(subject_id=subject.id, is_active=True).first()
        faculty_name = assignment.lecturer.name if assignment and assignment.lecturer else 'N/A'
        
        # Parse course name to extract course code and section
        course_name = subject.course.name if subject.course else 'N/A'
        course_parts = course_name.split() if course_name != 'N/A' else []
        
        # Determine if section exists (if there are 3+ parts, last part is section)
        has_section = len(course_parts) >= 3
        course_code = course_parts[-2] if has_section else (course_parts[-1] if course_parts else 'N/A')
        section = course_parts[-1] if has_section else None
        
        # Title
        ws['A1'] = 'Attendance Report'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Subject info
        ws['A3'] = 'Subject'
        ws['B3'] = subject.name
        ws['A4'] = 'Code'
        ws['B4'] = subject.code
        ws['A5'] = 'Course'
        ws['B5'] = course_code
        if has_section:
            ws['A6'] = 'Section'
            ws['B6'] = section
            ws['A7'] = 'Faculty'
            ws['B7'] = faculty_name
            # Year/Semester directly from subject
            ys_display = f"{getattr(subject, 'year', 'N/A')}/{getattr(subject, 'semester', 'N/A')}"
            ws['A8'] = 'Year/Semester'
            ws['B8'] = ys_display
        else:
            ws['A6'] = 'Faculty'
            ws['B6'] = faculty_name
            ys_display = f"{getattr(subject, 'year', 'N/A')}/{getattr(subject, 'semester', 'N/A')}"
            ws['A7'] = 'Year/Semester'
            ws['B7'] = ys_display
        # Determine where to place spacer row and headers
        last_info_row = 8 if has_section else 7
        spacer_row = last_info_row + 1
        header_row = spacer_row + 1
        data_start_row = header_row + 1

        # Spacer row above the table
        ws.merge_cells(start_row=spacer_row, start_column=1, end_row=spacer_row, end_column=6)
        spacer_cell = ws.cell(row=spacer_row, column=1, value='')
        spacer_cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

        # Table headers
        headers = ['Student', 'Roll Number', 'Present', 'Total', '%', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        row = data_start_row
        for record in attendance_report or []:
            student = record['student']
            percent = record.get('attendance_percentage') or 0
            status = 'Good' if percent >= 75 else 'Shortage'
            
            ws.cell(row=row, column=1, value=student.name)
            ws.cell(row=row, column=2, value=student.roll_number)
            ws.cell(row=row, column=3, value=record.get('present_classes') or 0)
            ws.cell(row=row, column=4, value=record.get('total_classes') or 0)
            ws.cell(row=row, column=5, value=f"{percent}%")
            ws.cell(row=row, column=6, value=status)
            row += 1
        
        if not attendance_report:
            ws.cell(row=data_start_row, column=1, value='No data')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    # ======================== ADDITIONAL PDF GENERATORS ========================
    @staticmethod
    def generate_class_marks_report_pdf(report):
        """Generate a PDF for the class marks report and return bytes."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm, topMargin=18*mm, bottomMargin=18*mm)
        elements = []
        styles = getSampleStyleSheet()
        # Header (same format as student PDF)
        try:
            from flask import current_app
            logo_path = current_app.root_path + '/static/img/logo-removebg-preview.png'
            logo_img = Image(logo_path)
            logo_img._restrictSize(26*mm, 26*mm)
        except Exception:
            logo_img = ''
        header_title = styles['Title']
        from reportlab.lib.styles import ParagraphStyle
        header_title = ParagraphStyle('HeaderTitle', parent=styles['Title'], alignment=0, fontSize=16, leading=19)
        header_sub = ParagraphStyle('HeaderSub', parent=styles['Normal'], alignment=0, fontSize=10, leading=12)
        header_text = [
            Paragraph('Dr. B. B. Hegde First Grade College, Kundapura', header_title),
            Paragraph('A Unit of Coondapur Education Society (R)', header_sub)
        ]
        header_table = Table([[logo_img, header_text]], colWidths=[26*mm, 148*mm])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LINEBELOW', (0,0), (-1,0), 0.75, colors.lightgrey),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 6))
        elements.append(Paragraph('Class Marks Report', styles['Title']))
        s = report.get('subject', {})
        meta_rows = [
            ['Subject', f"{s.get('name','')} ({s.get('code','')})"],
            ['Course', s.get('course_display') or s.get('course') or '']
        ]
        if s.get('section'):
            meta_rows.append(['Section', str(s.get('section')).upper()])
        meta_rows.append(['Year/Sem', f"{s.get('year','')}/{s.get('semester','')}"])
        # Faculty line if available
        try:
            lecturers = s.get('lecturers') or []
            if lecturers:
                meta_rows.append(['Faculty', ', '.join(lecturers)])
        except Exception:
            pass
        if report.get('assessment_type'):
            meta_rows.append(['Assessment Type', report.get('assessment_type')])
        meta_table = Table(meta_rows, colWidths=[40*mm, 120*mm])
        meta_table.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
            ('INNERGRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
        ]))
        elements.extend([Spacer(1, 6), meta_table, Spacer(1, 8)])

        # Statistics
        stats = report.get('statistics', {})
        stats_rows = [
            ['Total Students', stats.get('total_students', 0)],
            ['Class Average (%)', stats.get('class_average', 0)],
            ['Highest (%)', stats.get('highest_score', 0)],
            ['Lowest (%)', stats.get('lowest_score', 0)],
            ['Passing Students', stats.get('passing_students', 0)],
            ['Failing Students', stats.get('failing_students', 0)],
        ]
        stats_table = Table(stats_rows, colWidths=[55*mm, 105*mm])
        stats_table.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
            ('INNERGRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
        ]))
        elements.extend([Paragraph('Statistics', styles['Heading2']), stats_table, Spacer(1, 8)])

        # Student marks table
        # Check if specific assessment type is selected
        assessment_type = report.get('assessment_type')
        if assessment_type and assessment_type != 'All Assessments':
            # Remove Assessment column when specific assessment type is selected
            header = ['Student', 'Roll', 'Marks', 'Max', 'Percent']
            rows = [header]
            # Sort students by roll number ascending for readability
            student_marks_sorted = sorted(report.get('student_marks', []), key=lambda sm: (sm.get('roll_number') or ''))
            for sm in student_marks_sorted:
                for m in sm.get('marks', []):
                    percent = m.get('percentage') if 'percentage' in m else (
                        round((m.get('marks_obtained', 0) / m.get('max_marks', 1)) * 100, 2) if m.get('max_marks') else 0
                    )
                    rows.append([
                        sm.get('student_name',''), sm.get('roll_number',''),
                        ReportingService._format_number(m.get('marks_obtained')), ReportingService._format_number(m.get('max_marks')), ReportingService._format_number(percent)
                    ])
            if len(rows) == 1:
                rows.append(['No data', '', '', '', ''])
            # Build standardized table that fits the page width
            page_width = A4[0] - (18*mm + 18*mm)
            col_fracs = [0.45, 0.15, 0.15, 0.10, 0.15]  # Adjusted for 5 columns
            tbl = ReportingService._build_table(
                rows,
                page_width,
                col_fracs,
                no_wrap_cols={1,2,3,4},   # only Student may wrap
                center_cols={2,3,4},
                header_bg=colors.black,
            )
        else:
            # Include Assessment column when "All Assessments" is selected
            header = ['Student', 'Roll', 'Assessment', 'Marks', 'Max', 'Percent']
            rows = [header]
            # Sort students by roll number ascending for readability
            student_marks_sorted = sorted(report.get('student_marks', []), key=lambda sm: (sm.get('roll_number') or ''))
            for sm in student_marks_sorted:
                for m in sm.get('marks', []):
                    percent = m.get('percentage') if 'percentage' in m else (
                        round((m.get('marks_obtained', 0) / m.get('max_marks', 1)) * 100, 2) if m.get('max_marks') else 0
                    )
                    rows.append([
                        sm.get('student_name',''), sm.get('roll_number',''), m.get('assessment_type',''),
                        ReportingService._format_number(m.get('marks_obtained')), ReportingService._format_number(m.get('max_marks')), ReportingService._format_number(percent)
                    ])
        
        if len(rows) == 1:
            rows.append(['No data', '', '', '', '', ''])
        
        # Build standardized table that fits the page width
        page_width = A4[0] - (18*mm + 18*mm)
        col_fracs = [0.40, 0.14, 0.16, 0.10, 0.10, 0.10]
        tbl = ReportingService._build_table(
            rows,
            page_width,
            col_fracs,
            no_wrap_cols={1,2,3,4,5},   # only Student may wrap
            center_cols={3,4,5},
            header_bg=colors.black,
            )
        elements.append(tbl)

        doc.build(elements)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        return pdf_bytes

    @staticmethod
    def generate_class_attendance_report_pdf(report):
        """Generate a PDF for the class attendance report and return bytes."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm, topMargin=18*mm, bottomMargin=18*mm)
        elements = []
        styles = getSampleStyleSheet()
        # Header (same format as student PDF)
        try:
            from flask import current_app
            logo_path = current_app.root_path + '/static/img/logo-removebg-preview.png'
            logo_img = Image(logo_path)
            logo_img._restrictSize(26*mm, 26*mm)
        except Exception:
            logo_img = ''
        from reportlab.lib.styles import ParagraphStyle
        header_title = ParagraphStyle('HeaderTitle', parent=styles['Title'], alignment=0, fontSize=16, leading=19)
        header_sub = ParagraphStyle('HeaderSub', parent=styles['Normal'], alignment=0, fontSize=10, leading=12)
        header_text = [
            Paragraph('Dr. B. B. Hegde First Grade College, Kundapura', header_title),
            Paragraph('A Unit of Coondapur Education Society (R)', header_sub)
        ]
        header_table = Table([[logo_img, header_text]], colWidths=[26*mm, 148*mm])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LINEBELOW', (0,0), (-1,0), 0.75, colors.lightgrey),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 6))
        elements.append(Paragraph('Class Attendance Report', styles['Title']))
        s = report.get('subject', {})
        meta_rows = [
            ['Subject', f"{s.get('name','')} ({s.get('code','')})"],
            ['Course', s.get('course_display') or s.get('course') or '']
        ]
        if s.get('section'):
            meta_rows.append(['Section', str(s.get('section')).upper()])
        meta_rows.append(['Year/Sem', f"{s.get('year','')}/{s.get('semester','')}"])
        # Faculty line if available
        try:
            lecturers = s.get('lecturers') or []
            if lecturers:
                meta_rows.append(['Faculty', ', '.join(lecturers)])
        except Exception:
            pass
        # Period shown only when a specific month is selected (not Overall)
        if str(report.get('month','')).lower() != 'overall':
            meta_rows.append(['Period', f"{report.get('month','')} {report.get('year','')}"])
        meta_table = Table(meta_rows, colWidths=[40*mm, 120*mm])
        meta_table.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
            ('INNERGRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
        ]))
        elements.extend([Spacer(1, 6), meta_table, Spacer(1, 8)])

        stats = report.get('statistics', {})
        stats_rows = [
            ['Total Students', ReportingService._format_number(stats.get('total_students', 0))],
            ['Classes Conducted', ReportingService._format_number(stats.get('total_classes_conducted', 0))],
            ['Class Average (%)', ReportingService._format_number(stats.get('class_average_attendance', 0))],
            ['Good Attendance (≥75%)', ReportingService._format_number(stats.get('students_with_good_attendance', 0))],
            ['Poor Attendance (<50%)', ReportingService._format_number(stats.get('students_with_poor_attendance', 0))],
        ]
        stats_table = Table(stats_rows, colWidths=[60*mm, 100*mm])
        stats_table.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
            ('INNERGRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
        ]))
        elements.extend([Paragraph('Statistics', styles['Heading2']), stats_table, Spacer(1, 8)])

        header = ['Student', 'Roll', 'Total', 'Present', 'Absent', 'Percent', 'Status']
        rows = [header]
        for st in report.get('student_attendance', []):
            rows.append([
                st.get('student_name',''), st.get('roll_number',''), ReportingService._format_number(st.get('total_classes')),
                ReportingService._format_number(st.get('present_classes')), ReportingService._format_number(st.get('absent_classes')), ReportingService._format_number(st.get('attendance_percentage')), st.get('status','')
            ])
        if len(rows) == 1:
            rows.append(['No data', '', '', '', '', '', ''])
        # Wrap text in table data
        # Prevent wrapping for Roll (1)
        rows_wrapped = ReportingService._wrap_table_data(rows, skip_header=True, header_text_white=True, no_wrap_cols={1})
        page_width = A4[0] - (18*mm + 18*mm)
        # 7 columns: Student | Roll | Total | Present | Absent | Percent | Status
        # Use proportional widths for a clean, consistent layout
        # Student gets 38%, Roll 16%, each numeric column 9%, Status 10%
        proportions = [0.38, 0.16, 0.09, 0.09, 0.09, 0.09, 0.10]
        col_widths = [page_width * p for p in proportions]
        tbl = Table(rows_wrapped, repeatRows=1, colWidths=col_widths)
        tbl.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
            ('INNERGRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
            ('BACKGROUND', (0,0), (-1,0), colors.black),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold')
        ]))
        elements.append(tbl)

        doc.build(elements)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        return pdf_bytes

    @staticmethod
    def generate_course_overview_report_pdf(report):
        """Generate a PDF for the course overview report and return bytes."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm, topMargin=18*mm, bottomMargin=18*mm)
        elements = []
        styles = getSampleStyleSheet()
        # Header (same format as student PDF)
        try:
            from flask import current_app
            logo_path = current_app.root_path + '/static/img/logo-removebg-preview.png'
            logo_img = Image(logo_path)
            logo_img._restrictSize(26*mm, 26*mm)
        except Exception:
            logo_img = ''
        from reportlab.lib.styles import ParagraphStyle
        header_title = ParagraphStyle('HeaderTitle', parent=styles['Title'], alignment=0, fontSize=16, leading=19)
        header_sub = ParagraphStyle('HeaderSub', parent=styles['Normal'], alignment=0, fontSize=10, leading=12)
        header_text = [
            Paragraph('Dr. B. B. Hegde First Grade College, Kundapura', header_title),
            Paragraph('A Unit of Coondapur Education Society (R)', header_sub)
        ]
        header_table = Table([[logo_img, header_text]], colWidths=[26*mm, 148*mm])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LINEBELOW', (0,0), (-1,0), 0.75, colors.lightgrey),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 6))
        elements.append(Paragraph('Course Overview Report', styles['Title']))
        c = report.get('course', {})
        meta_rows = [
            ['Class', c.get('name','')],
            ['Duration (years)', ReportingService._format_number(c.get('duration_years'))],
            ['Total Semesters', ReportingService._format_number(c.get('total_semesters'))],
            ['Total Students', ReportingService._format_number(report.get('total_students'))],
            ['Total Subjects', ReportingService._format_number(report.get('total_subjects'))],
        ]
        meta_table = Table(meta_rows, colWidths=[55*mm, 105*mm])
        meta_table.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
            ('INNERGRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
        ]))
        elements.extend([Spacer(1, 6), meta_table, Spacer(1, 8)])

        # Subjects table
        header = ['Subject', 'Code', 'Enrolled', 'Avg Marks %', 'Pass Rate %', 'Avg Attendance %']
        rows = [header]
        for s in report.get('subjects', []):
            rows.append([
                s.get('subject_name',''), s.get('subject_code',''),
                ReportingService._format_number(s.get('enrolled_students')), ReportingService._format_number(s.get('marks_statistics',{}).get('average_marks',0)),
                ReportingService._format_number(s.get('marks_statistics',{}).get('passing_rate',0)), ReportingService._format_number(s.get('attendance_statistics',{}).get('average_attendance',0))
            ])
        if len(rows) == 1:
            rows.append(['No data', '', '', '', '', ''])
        
        # Build standardized table with proper column widths
        page_width = A4[0] - (18*mm + 18*mm)
        col_fracs = [0.25, 0.15, 0.10, 0.12, 0.12, 0.18]  # Increased Avg Attendance column width to fit full header
        tbl = ReportingService._build_table(
            rows,
            page_width,
            col_fracs,
            no_wrap_cols={2,3,4,5},  # Allow Subject and Code columns to wrap, prevent others
            center_cols={0,1,2,3,4,5}     # Center all columns including wrapped text
        )
        elements.append(tbl)

        doc.build(elements)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        return pdf_bytes

    # ======================== LECTURER SHORTAGE/DEFICIENCY PDFS ========================
    @staticmethod
    def generate_attendance_shortage_pdf(threshold, shortage_data, lecturer_name=None):
        """Generate a PDF for Attendance Shortage (lecturer view)."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm, topMargin=18*mm, bottomMargin=18*mm)
        elements = []
        styles = getSampleStyleSheet()
        from reportlab.lib.styles import ParagraphStyle
        header_title = ParagraphStyle('HeaderTitle', parent=styles['Title'], alignment=0, fontSize=16, leading=19)
        header_sub = ParagraphStyle('HeaderSub', parent=styles['Normal'], alignment=0, fontSize=10, leading=12)

        # Header (logo + college text)
        try:
            from flask import current_app
            logo_path = current_app.root_path + '/static/img/logo-removebg-preview.png'
            logo_img = Image(logo_path)
            logo_img._restrictSize(26*mm, 26*mm)
        except Exception:
            logo_img = ''
        header_text = [
            Paragraph('Dr. B. B. Hegde First Grade College, Kundapura', header_title),
            Paragraph('A Unit of Coondapur Education Society (R)', header_sub)
        ]
        header_table = Table([[logo_img, header_text]], colWidths=[26*mm, (A4[0] - (18*mm + 18*mm) - 26*mm)])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LINEBELOW', (0,0), (-1,0), 0.75, colors.lightgrey),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        elements.append(header_table)

        # Title and subheading (threshold)
        from reportlab.lib.styles import ParagraphStyle
        title_center = ParagraphStyle('TitleCenter', parent=styles['Title'], alignment=1)
        sub_center = ParagraphStyle('SubCenter', parent=styles['Normal'], alignment=1, fontSize=10)
        elements.extend([Spacer(1, 8), Paragraph('Attendance Shortage Report', title_center)])
        elements.append(Paragraph(f"Below {ReportingService._format_number(threshold)}%", sub_center))
        elements.append(Spacer(1, 10))

        # Create separate tables for each subject
        sorted_shortage_data = sorted(shortage_data or [], key=lambda x: getattr(x.get('subject'), 'name', '') if x.get('subject') else '')
        
        for block_idx, block in enumerate(sorted_shortage_data):
            subj = block.get('subject')
            if not subj:
                continue
                
            course_name = subj.course.name if getattr(subj, 'course', None) else ''
            
            # If only one subject in export, place a single info table at the top
            if len(sorted_shortage_data) == 1 and block_idx == 0:
                info_rows = [
                    ['Subject', subj.name],
                    ['Class', course_name],
                    ['Code', subj.code],
                    ['Faculty', lecturer_name or 'N/A']
                ]
                info_table = Table(info_rows, colWidths=[25*mm, 135*mm])
                info_table.setStyle(TableStyle([
                    ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
                    ('INNERGRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
                ]))
                elements.extend([info_table, Spacer(1, 8)])
            else:
                # For multiple subjects, keep a compact subject header above each table
                if block_idx > 0:
                    elements.append(Spacer(1, 12))
                subject_rows = [
                    ['Subject', subj.name],
                    ['Class', course_name],
                    ['Code', subj.code],
                    ['Faculty', lecturer_name or 'N/A']
                ]
                subject_table = Table(subject_rows, colWidths=[25*mm, 135*mm])
                subject_table.setStyle(TableStyle([
                    ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
                    ('INNERGRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
                ]))
                elements.extend([subject_table, Spacer(1, 8)])

            # Table for this subject - remove Shortage column
            headers = ['Student', 'Roll', 'Present', 'Total', '%']
            rows = [headers]
            
            # Sort students by roll number (last 3 digits)
            shortage_students = block.get('shortage_students') or []
            def get_roll_sort_key(rec):
                roll_number = rec['student'].roll_number
                if len(roll_number) >= 3:
                    last_three = roll_number[-3:]
                    try:
                        return int(last_three)
                    except ValueError:
                        return 999
                return 999
            
            sorted_students = sorted(shortage_students, key=get_roll_sort_key)
            
            for rec in sorted_students:
                pct = rec.get('attendance_percentage') or 0
                rows.append([
                    rec['student'].name,
                    rec['student'].roll_number,
                    ReportingService._format_number(rec.get('present_classes')),
                    ReportingService._format_number(rec.get('total_classes')),
                    ReportingService._format_number(pct),
                ])
            
            if len(rows) == 1:
                rows.append(['No data', '', '', '', ''])

            # Only Student column (0) may wrap; all others should not wrap
            rows_wrapped = ReportingService._wrap_table_data(rows, skip_header=True, header_text_white=True, no_wrap_cols={1,2,3,4})
            page_width = A4[0] - (18*mm + 18*mm)
            # Set widths for new column order: 0 Student | 1 Roll | 2 Present | 3 Total | 4 % | 5 Shortage
            # Calculate proper widths that fit within page boundaries
            # A4 width is 210mm, minus margins (36mm) = 174mm available
            col_widths = [70*mm, 30*mm, 22*mm, 22*mm, 18*mm]
            tbl = Table(rows_wrapped, repeatRows=1, colWidths=col_widths)
            tbl.setStyle(TableStyle([
                ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
                ('INNERGRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
                ('BACKGROUND', (0,0), (-1,0), colors.black),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold')
            ]))
            elements.append(tbl)

        doc.build(elements)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        return pdf_bytes

    @staticmethod
    def generate_marks_deficiency_pdf(threshold, deficiency_data, lecturer_name=None):
        """Generate a PDF for Marks Deficiency (lecturer view)."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm, topMargin=18*mm, bottomMargin=18*mm)
        elements = []
        styles = getSampleStyleSheet()
        from reportlab.lib.styles import ParagraphStyle
        header_title = ParagraphStyle('HeaderTitle', parent=styles['Title'], alignment=0, fontSize=16, leading=19)
        header_sub = ParagraphStyle('HeaderSub', parent=styles['Normal'], alignment=0, fontSize=10, leading=12)

        # Header
        try:
            from flask import current_app
            logo_path = current_app.root_path + '/static/img/logo-removebg-preview.png'
            logo_img = Image(logo_path)
            logo_img._restrictSize(26*mm, 26*mm)
        except Exception:
            logo_img = ''
        header_text = [
            Paragraph('Dr. B. B. Hegde First Grade College, Kundapura', header_title),
            Paragraph('A Unit of Coondapur Education Society (R)', header_sub)
        ]
        header_table = Table([[logo_img, header_text]], colWidths=[26*mm, (A4[0] - (18*mm + 18*mm) - 26*mm)])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LINEBELOW', (0,0), (-1,0), 0.75, colors.lightgrey),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        elements.append(header_table)

        # Title + subheading centered
        from reportlab.lib.styles import ParagraphStyle
        title_center = ParagraphStyle('TitleCenter', parent=styles['Title'], alignment=1)
        sub_center = ParagraphStyle('SubCenter', parent=styles['Normal'], alignment=1, fontSize=10)
        elements.extend([Spacer(1, 8), Paragraph('Marks Deficiency Report', title_center)])
        elements.append(Paragraph(f"Below {ReportingService._format_number(threshold)}%", sub_center))
        elements.append(Spacer(1, 10))

        # Create separate tables for each subject
        sorted_deficiency_data = sorted(deficiency_data or [], key=lambda x: getattr(x.get('subject'), 'name', '') if x.get('subject') else '')
        
        # If a single subject is present, render one info table at the top only
        single_subject_mode = len(sorted_deficiency_data) == 1 and sorted_deficiency_data[0].get('subject') is not None
        if single_subject_mode:
            subj = sorted_deficiency_data[0]['subject']
            course_name = subj.course.name if getattr(subj, 'course', None) else ''
            wrap_style = ParagraphStyle('WrapSmall', parent=styles['Normal'], fontSize=10, leading=12)
            info_rows = [
                ['Subject', Paragraph(str(subj.name or ''), wrap_style)],
                ['Class', Paragraph(str(course_name or ''), wrap_style)],
                ['Code', Paragraph(str(subj.code or ''), wrap_style)],
                ['Faculty', Paragraph(str(lecturer_name or 'N/A'), wrap_style)]
            ]
            info_table = Table(info_rows, colWidths=[25*mm, (A4[0] - (18*mm + 18*mm) - 25*mm)])
            info_table.setStyle(TableStyle([
                ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
                ('INNERGRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
                ('LEFTPADDING', (0,0), (-1,-1), 4),
                ('RIGHTPADDING', (0,0), (-1,-1), 4),
                ('TOPPADDING', (0,0), (-1,-1), 2),
                ('BOTTOMPADDING', (0,0), (-1,-1), 2),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ]))
            elements.extend([info_table, Spacer(1, 8)])

        for block_idx, block in enumerate(sorted_deficiency_data):
            subj = block.get('subject')
            if not subj:
                continue
                
            course_name = subj.course.name if getattr(subj, 'course', None) else ''
            
            # For multiple subjects, show compact info table above each table
            if not single_subject_mode:
                if block_idx > 0:
                    elements.append(Spacer(1, 12))
                wrap_style = ParagraphStyle('WrapSmall', parent=styles['Normal'], fontSize=10, leading=12)
                subject_rows = [
                    ['Subject', Paragraph(str(subj.name or ''), wrap_style)],
                    ['Class', Paragraph(str(course_name or ''), wrap_style)],
                    ['Code', Paragraph(str(subj.code or ''), wrap_style)],
                    ['Faculty', Paragraph(str(lecturer_name or 'N/A'), wrap_style)]
                ]
                subject_table = Table(subject_rows, colWidths=[25*mm, (A4[0] - (18*mm + 18*mm) - 25*mm)])
                subject_table.setStyle(TableStyle([
                    ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
                    ('INNERGRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
                    ('LEFTPADDING', (0,0), (-1,-1), 4),
                    ('RIGHTPADDING', (0,0), (-1,-1), 4),
                    ('TOPPADDING', (0,0), (-1,-1), 2),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 2),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ]))
                elements.extend([subject_table, Spacer(1, 8)])

            # Table for this subject - removed Subject, Code, Course columns
            # Dynamic headers: include only components that have recorded (non-zero) values
            headers = ['Student', 'Roll', 'Overall %']
            comp_keys = ['internal1','internal2','assignment','project']
            comp_to_header = {'internal1':'Internal 1','internal2':'Internal 2','assignment':'Assignment','project':'Project'}
            include = {k: False for k in comp_keys}
            for rec in block.get('deficient_students') or []:
                ms = rec.get('marks_summary') or {}
                for k in comp_keys:
                    v = ms.get(k)
                    try:
                        if isinstance(v, dict):
                            obt = float(v.get('obtained') or 0)
                            mx = float(v.get('max') or 0)
                        else:
                            obt = float(getattr(v, 'obtained', 0) or 0)
                            mx = float(getattr(v, 'max', 0) or 0)
                        if obt > 0 or mx > 0:
                            include[k] = True
                    except Exception:
                        pass
            ordered_components = [k for k in comp_keys if include[k]]
            headers += [comp_to_header[k] for k in ordered_components]
            rows = [headers]
            
            # Sort students by roll number (last 3 digits)
            deficient_students = block.get('deficient_students') or []
            def get_roll_sort_key(rec):
                roll_number = rec['student'].roll_number
                if len(roll_number) >= 3:
                    last_three = roll_number[-3:]
                    try:
                        return int(last_three)
                    except ValueError:
                        return 999
                return 999
            
            sorted_students = sorted(deficient_students, key=get_roll_sort_key)
            
            for rec in sorted_students:
                ms = rec.get('marks_summary') or {}
                def _pair(d):
                    try:
                        if isinstance(d, dict):
                            obt = d.get('obtained'); mx = d.get('max')
                        else:
                            obt = getattr(d, 'obtained', None); mx = getattr(d, 'max', None)
                        if obt in (None, 0, '0', '0.0') and mx in (None, 0, '0', '0.0'):
                            return ''
                        return f"{ReportingService._format_number(obt)}/{ReportingService._format_number(mx)}"
                    except Exception:
                        return ''
                row = [
                    rec['student'].name,
                    rec['student'].roll_number,
                    ReportingService._format_number(rec.get('overall_percentage') or 0),
                ]
                for k in ordered_components:
                    row.append(_pair(ms.get(k)))
                rows.append(row)
            
            if len(rows) == 1:
                rows.append(['No data', '', '', '', '', '', ''])

            # Only Student column (0) may wrap; all others should not wrap
            no_wrap_cols = set(range(1, len(headers)))
            rows_wrapped = ReportingService._wrap_table_data(rows, skip_header=True, header_text_white=True, no_wrap_cols=no_wrap_cols)
            page_width = A4[0] - (18*mm + 18*mm)
            base_widths = [70*mm, 30*mm, 20*mm]
            extra_cols = max(0, len(headers) - 3)
            if extra_cols:
                per = max(18*mm, (page_width - sum(base_widths)) / extra_cols)
                col_widths = base_widths + [per for _ in range(extra_cols)]
            else:
                col_widths = base_widths
            tbl = Table(rows_wrapped, repeatRows=1, colWidths=col_widths)
            tbl.setStyle(TableStyle([
                ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
                ('INNERGRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
                ('BACKGROUND', (0,0), (-1,0), colors.black),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold')
            ]))
            elements.append(tbl)

        doc.build(elements)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        return pdf_bytes

    @staticmethod
    def get_comprehensive_class_report(course_id, report_type='attendance', assessment_type=None):
        """Get comprehensive class report for all subjects in a course"""
        try:
            from models.academic import Course, Subject
            from models.student import Student
            from models.attendance import MonthlyAttendanceSummary, MonthlyStudentAttendance
            from models.marks import StudentMarks
            from sqlalchemy import func, and_, or_
            from database import db
            
            # Get course information
            course = Course.query.get(course_id)
            if not course:
                return None
            
            # Get all subjects for this course
            subjects = Subject.query.filter_by(course_id=course_id, is_active=True).order_by(Subject.name).all()
            
            # Get all students for this course
            students = Student.query.filter_by(course_id=course_id, is_active=True).order_by(Student.roll_number).all()
            
            report = {
                'course': {
                    'id': course.id,
                    'name': course.name,
                    'code': course.code
                },
                'report_type': report_type,
                'assessment_type': assessment_type,
                'subjects': [],
                'students': [],
                'data': {}
            }
            
            # Add student information
            for student in students:
                report['students'].append({
                    'id': student.id,
                    'name': student.name,
                    'roll_number': student.roll_number
                })
            
            # Add subject information with max marks
            for subject in subjects:
                # Get max marks for each assessment type for this subject
                max_marks = {
                    'internal1': 0,
                    'internal2': 0,
                    'assignment': 0,
                    'project': 0
                }
                
                # Get max marks from any student's marks for this subject
                sample_marks = StudentMarks.query.filter_by(subject_id=subject.id).first()
                if sample_marks:
                    # Get all unique max marks for each assessment type
                    for assessment_type in ['internal1', 'internal2', 'assignment', 'project']:
                        max_mark_record = StudentMarks.query.filter_by(
                            subject_id=subject.id,
                            assessment_type=assessment_type
                        ).first()
                        if max_mark_record and max_mark_record.max_marks:
                            max_marks[assessment_type] = max_mark_record.max_marks
                
                report['subjects'].append({
                    'id': subject.id,
                    'name': subject.name,
                    'code': subject.code,
                    'year': subject.year,
                    'semester': subject.semester,
                    'max_marks': max_marks
                })
            
            if report_type == 'attendance':
                # Get attendance data for all subjects
                for subject in subjects:
                    subject_data = {
                        'subject_id': subject.id,
                        'subject_name': subject.name,
                        'subject_code': subject.code,
                        'student_attendance': {},
                        'enrolled_students': {}
                    }
                    
                    # Get enrolled students for this subject
                    from models.student import StudentEnrollment
                    enrolled_student_ids = db.session.query(StudentEnrollment.student_id).filter_by(
                        subject_id=subject.id, 
                        is_active=True
                    ).all()
                    enrolled_student_ids = [sid[0] for sid in enrolled_student_ids]
                    
                    # Mark which students are enrolled
                    for student in students:
                        subject_data['enrolled_students'][student.id] = student.id in enrolled_student_ids
                    
                    # Get overall attendance for each student in this subject
                    for student in students:
                        # Get total classes and attendance across all months
                        total_classes_query = db.session.query(
                            func.sum(MonthlyAttendanceSummary.total_classes)
                        ).filter(
                            MonthlyAttendanceSummary.subject_id == subject.id
                        ).scalar() or 0
                        
                        # Get total present classes (including deputation)
                        total_present_query = db.session.query(
                            func.sum(MonthlyStudentAttendance.present_count + MonthlyStudentAttendance.deputation_count)
                        ).filter(
                            and_(
                                MonthlyStudentAttendance.subject_id == subject.id,
                                MonthlyStudentAttendance.student_id == student.id
                            )
                        ).scalar() or 0
                        
                        # Calculate percentage
                        percentage = 0
                        if total_classes_query > 0:
                            percentage = round((total_present_query / total_classes_query) * 100, 2)
                        
                        subject_data['student_attendance'][student.id] = {
                            'total_classes': total_classes_query,
                            'present_classes': total_present_query,
                            'percentage': percentage
                        }
                    
                    report['data'][subject.id] = subject_data
            
            elif report_type == 'marks':
                # Get marks data for all subjects
                for subject in subjects:
                    subject_data = {
                        'subject_id': subject.id,
                        'subject_name': subject.name,
                        'subject_code': subject.code,
                        'student_marks': {},
                        'enrolled_students': {}
                    }
                    
                    # Get enrolled students for this subject
                    from models.student import StudentEnrollment
                    enrolled_student_ids = db.session.query(StudentEnrollment.student_id).filter_by(
                        subject_id=subject.id, 
                        is_active=True
                    ).all()
                    enrolled_student_ids = [sid[0] for sid in enrolled_student_ids]
                    
                    # Mark which students are enrolled
                    for student in students:
                        subject_data['enrolled_students'][student.id] = student.id in enrolled_student_ids
                    
                    for student in students:
                        student_marks = {
                            'internal1': {'obtained': None, 'max': None, 'recorded': False},
                            'internal2': {'obtained': None, 'max': None, 'recorded': False},
                            'assignment': {'obtained': None, 'max': None, 'recorded': False},
                            'project': {'obtained': None, 'max': None, 'recorded': False},
                            'overall_percentage': 0
                        }
                        
                        # Get marks for each assessment type for this student and subject
                        marks_records = StudentMarks.query.filter_by(
                            student_id=student.id,
                            subject_id=subject.id
                        ).all()
                        
                        for marks in marks_records:
                            assessment_type = marks.assessment_type.lower()
                            if assessment_type in student_marks:
                                student_marks[assessment_type]['obtained'] = marks.marks_obtained or 0
                                student_marks[assessment_type]['max'] = marks.max_marks or 0
                                student_marks[assessment_type]['recorded'] = True
                        
                        # Calculate overall percentage (only for recorded marks)
                        total_obtained = sum([m['obtained'] for m in student_marks.values() if isinstance(m, dict) and m.get('recorded', False) and m['obtained'] is not None])
                        total_max = sum([m['max'] for m in student_marks.values() if isinstance(m, dict) and m.get('recorded', False) and m['max'] is not None])
                        
                        # Add total marks for display
                        student_marks['total_obtained'] = total_obtained
                        student_marks['total_max'] = total_max
                        
                        if total_max > 0:
                            student_marks['overall_percentage'] = round((total_obtained / total_max) * 100, 2)
                        
                        subject_data['student_marks'][student.id] = student_marks
                    
                    report['data'][subject.id] = subject_data
            
            return report
            
        except Exception as e:
            print(f"Error in get_comprehensive_class_report: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
    @staticmethod
    def generate_comprehensive_class_report_pdf(report):
        """Generate PDF for comprehensive class report with proper subject grouping"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import mm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            from io import BytesIO
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm, topMargin=18*mm, bottomMargin=18*mm, showBoundary=0)
            elements = []
            styles = getSampleStyleSheet()
            
            # Header
            try:
                from flask import current_app
                logo_path = current_app.root_path + '/static/img/logo-removebg-preview.png'
                logo_img = Image(logo_path)
                logo_img._restrictSize(26*mm, 26*mm)
            except Exception:
                logo_img = ''
            
            from reportlab.lib.styles import ParagraphStyle
            header_title = ParagraphStyle('HeaderTitle', parent=styles['Title'], alignment=0, fontSize=16, leading=19)
            header_sub = ParagraphStyle('HeaderSub', parent=styles['Normal'], alignment=0, fontSize=10, leading=12)
            
            header_text = [
                Paragraph('Dr. B. B. Hegde First Grade College, Kundapura', header_title),
                Paragraph('A Unit of Coondapur Education Society (R)', header_sub)
            ]
            header_table = Table([[logo_img, header_text]], colWidths=[26*mm, 148*mm])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('LINEBELOW', (0,0), (-1,0), 0.75, colors.lightgrey),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('TOPPADDING', (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ]))
            elements.append(header_table)
            elements.append(Spacer(1, 6))
            
            # Title
            report_type = report.get('report_type', 'attendance').title()
            course_name = report.get('course', {}).get('name', 'Unknown Course')
            elements.append(Paragraph(f'Comprehensive {report_type} Report', styles['Title']))
            
            # Course info table with professional styling
            course_info = [
                ['Class', course_name],
                ['Year/Semester', f"{report.get('subjects', [{}])[0].get('year', 'N/A')}/{report.get('subjects', [{}])[0].get('semester', 'N/A')}" if report.get('subjects') and len(report.get('subjects', [])) > 0 else 'N/A/N/A'],
                ['No of Subjects', str(len(report.get('subjects', [])))]
            ]
            
            # Add Assessment Type row for marks reports
            if report['report_type'] == 'marks' and report.get('assessment_type'):
                assessment_display = report['assessment_type'].title().replace('1', ' 1').replace('2', ' 2')
                course_info.append(['Assessment', assessment_display])
            
            # Create custom table for course info without header styling
            page_width = A4[0] - (18*mm + 18*mm)
            col_widths = [page_width * 0.25, page_width * 0.75]
            course_table = Table(course_info, colWidths=col_widths)
            
            # Apply consistent styling without header treatment
            course_table.setStyle(TableStyle([
                ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
                ('INNERGRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
                ('FONTSIZE', (0,0), (-1,-1), 10),
                ('LEFTPADDING', (0,0), (-1,-1), 6),
                ('RIGHTPADDING', (0,0), (-1,-1), 6),
                ('TOPPADDING', (0,0), (-1,-1), 4),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                # Alternating row backgrounds
                ('BACKGROUND', (0,0), (-1,0), colors.white),
                ('BACKGROUND', (0,1), (-1,1), colors.lightgrey),
                ('BACKGROUND', (0,2), (-1,2), colors.white),
                ('BACKGROUND', (0,3), (-1,3), colors.lightgrey),
            ]))
            elements.extend([Spacer(1, 6), course_table, Spacer(1, 12)])
            
            # Group subjects into pages (4 subjects per page)
            subjects_per_page = 4
            subjects = report['subjects']
            students = report['students']
            
            for page_start in range(0, len(subjects), subjects_per_page):
                page_subjects = subjects[page_start:page_start + subjects_per_page]
                
                if page_start > 0:
                    elements.append(Spacer(1, 20))
                
                # Create table for this page
                if report['report_type'] == 'attendance':
                    # Attendance report
                    headers = ['Roll No', 'Student Name'] + [subj['name'] for subj in page_subjects]
                    rows = [headers]
                    
                    for student in students:
                        row_data = [student['roll_number'], student['name']]
                        
                        for subject in page_subjects:
                            subject_data = report['data'].get(subject['id'], {})
                            student_attendance = subject_data.get('student_attendance', {}).get(student['id'], {})
                            
                            # Check if student is enrolled in this subject
                            is_enrolled = subject_data.get('enrolled_students', {}).get(student['id'], False)
                            
                            if not is_enrolled:
                                row_data.append("NA")
                            elif student_attendance and student_attendance.get('total_classes', 0) > 0:
                                percentage = student_attendance.get('percentage', 0)
                                row_data.append(f"{percentage}%")
                            else:
                                row_data.append("-")
                        
                        rows.append(row_data)
                
                elif report['report_type'] == 'marks':
                    # Marks report - spreadsheet format
                    assessment_type = report.get('assessment_type')
                    
                    if assessment_type:
                        # Specific assessment type selected
                        headers = ['Roll No', 'Student Name']
                        for subj in page_subjects:
                            headers.append(subj['name'])
                        rows = [headers]
                        
                        for student in students:
                            row_data = [student['roll_number'], student['name']]
                            
                            for subject in page_subjects:
                                subject_data = report['data'].get(subject['id'], {})
                                student_marks = subject_data.get('student_marks', {}).get(student['id'], {})
                                
                                # Check if student is enrolled in this subject
                                is_enrolled = subject_data.get('enrolled_students', {}).get(student['id'], False)
                                
                                if not is_enrolled:
                                    row_data.append("NA")
                                elif student_marks and assessment_type in student_marks:
                                    assessment_data = student_marks[assessment_type]
                                    if assessment_data.get('recorded', False):
                                        obtained = assessment_data['obtained']
                                        max_marks = assessment_data['max']
                                        if max_marks > 0:
                                            # Format marks: remove .0 from whole numbers
                                            obtained_str = str(int(obtained)) if obtained == int(obtained) else str(obtained)
                                            max_marks_str = str(int(max_marks)) if max_marks == int(max_marks) else str(max_marks)
                                            row_data.append(f"{obtained_str}/{max_marks_str}")
                                        else:
                                            obtained_str = str(int(obtained)) if obtained == int(obtained) else str(obtained)
                                            row_data.append(obtained_str)
                                    else:
                                        row_data.append("-")
                                else:
                                    row_data.append("-")
                            
                            rows.append(row_data)
                
                # Calculate column widths with better proportions
                page_width = A4[0] - (18*mm + 18*mm)
                num_cols = len(headers)
                
                if report['report_type'] == 'marks' and assessment_type:
                    # For marks with specific assessment, give more space to subject columns
                    col_widths = []
                    for i, header in enumerate(headers):
                        if i == 0:  # Roll No
                            col_widths.append(20*mm)
                        elif i == 1:  # Student Name
                            col_widths.append(35*mm)
                        else:  # Subject columns
                            remaining_width = page_width - 55*mm  # Reserve space for Roll No and Name
                            subject_cols = num_cols - 2
                            col_widths.append(remaining_width / subject_cols)
                else:
                    # Default equal width distribution
                    col_width = page_width / num_cols
                    col_widths = [col_width] * num_cols
                
                # Use the same _build_table function as class marks report for consistent styling
                if report['report_type'] == 'marks' and assessment_type and assessment_type != 'all':
                    # For marks with specific assessment, use optimized column fractions
                    col_fracs = [0.12, 0.23] + [0.65 / (num_cols - 2)] * (num_cols - 2)  # Increased Roll No width
                    no_wrap_cols = {0}  # Only allow Student Name to wrap, keep Roll No single line
                    center_cols = {0}  # Center Roll No
                    for i in range(2, num_cols):
                        center_cols.add(i)  # Center all subject columns
                    
                    table = ReportingService._build_table(
                        rows,
                        page_width,
                        col_fracs,
                        no_wrap_cols=no_wrap_cols,
                        center_cols=center_cols,
                        header_bg=colors.black,
                    )
                else:
                    # For other reports, use default styling
                    col_fracs = [1.0 / num_cols] * num_cols
                    table = ReportingService._build_table(
                        rows,
                        page_width,
                        col_fracs,
                        header_bg=colors.black,
                    )
                
                elements.append(table)
                
                # Add page break if not the last page (without page number text)
                if page_start + subjects_per_page < len(subjects):
                    elements.append(Spacer(1, 20))
            
            doc.build(elements)
            pdf_bytes = buffer.getvalue()
            buffer.close()
            return pdf_bytes
            
        except Exception as e:
            print(f"Error in generate_comprehensive_class_report_pdf: {e}")
            return None

    # ------------------------- PDF Helper Utilities -------------------------
    @staticmethod
    def _calc_colwidths_from_fracs(total_width, fracs):
        safe_fracs = fracs or []
        s = float(sum(safe_fracs)) or 1.0
        normalized = [f / s for f in safe_fracs]
        return [total_width * f for f in normalized]

    @staticmethod
    def _build_table(rows, page_width, col_fracs, *, no_wrap_cols=None, center_cols=None, header_bg=colors.black, col_font_sizes=None):
        """Build a standardized table with consistent styling across PDFs.
        - rows: 2D list with header at index 0
        - page_width: available width for table
        - col_fracs: fractions for each column width
        - no_wrap_cols: set of indices that must not wrap
        - center_cols: set of indices to center-align in body
        - col_font_sizes: dict of {col_index: font_size} for custom font sizes
        """
        wrapped = ReportingService._wrap_table_data(rows, skip_header=True, header_text_white=True, no_wrap_cols=no_wrap_cols or set())
        colwidths = ReportingService._calc_colwidths_from_fracs(page_width, col_fracs)
        tbl = Table(wrapped, repeatRows=1, colWidths=colwidths)
        center_cols = center_cols or set()
        # Base style
        base_style = [
            ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
            ('INNERGRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
            ('BACKGROUND', (0,0), (-1,0), header_bg),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('VALIGN', (0,0), (-1,0), 'MIDDLE'),
            # Body paddings and alignment
            ('VALIGN', (0,1), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,1), (-1,-1), 3),
            ('RIGHTPADDING', (0,1), (-1,-1), 3),
            ('TOPPADDING', (0,1), (-1,-1), 2),
            ('BOTTOMPADDING', (0,1), (-1,-1), 2),
        ]
        
        # Add custom font sizes for specific columns
        if col_font_sizes:
            for col_idx, font_size in col_font_sizes.items():
                base_style.append(('FONTSIZE', (col_idx, 0), (col_idx, -1), font_size))
        
        # Center specified columns in body
        if center_cols:
            min_idx = min(center_cols)
            max_idx = max(center_cols)
            base_style.append(('ALIGN', (min_idx,1), (max_idx,-1), 'CENTER'))
        tbl.setStyle(TableStyle(base_style))
        return tbl

    @staticmethod
    def _nbsp(text):
        try:
            return (text or '').replace(' ', '\u00A0')
        except Exception:
            return text
