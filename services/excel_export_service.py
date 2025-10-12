"""
Excel export service for Moulya College Management System
Handles Excel export for reports
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime

class ExcelExportService:
    """Service for exporting reports to Excel"""
    
    @staticmethod
    def create_workbook():
        """Create a new workbook with default styling"""
        wb = openpyxl.Workbook()
        return wb
    
    @staticmethod
    def style_header_row(ws, row_num, columns):
        """Apply styling to header row"""
        header_font = Font(bold=True, color="FFFFFF")
        # Change header background to black across all management exports
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(columns, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    @staticmethod
    def auto_adjust_columns(ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def center_all_cells(ws):
        """Center align all populated cells in the given worksheet."""
        try:
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    if cell.value is not None:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
        except Exception:
            pass

    @staticmethod
    def format_number(value):
        """Format number: whole numbers without decimals, fractional numbers with 2 decimal places."""
        try:
            if value is None:
                return None
            num = float(value)
            if num == int(num):
                return int(num)  # 32.0 -> 32
            else:
                return round(num, 2)  # 32.43 -> 32.43, 32.05 -> 32.05
        except (ValueError, TypeError):
            return value

    @staticmethod
    def set_number(cell, value, align_right=False):
        """Set an integer/float number with alignment preferences."""
        cell.value = ExcelExportService.format_number(value)
        cell.alignment = Alignment(horizontal=("right" if align_right else "left"), vertical="center")
        return cell

    @staticmethod
    def set_percentage(cell, percent_0_to_100, align_left=True):
        """Write a numeric percentage (avoid text with green triangle)."""
        try:
            if percent_0_to_100 is None:
                cell.value = None
            else:
                # Convert 65.88 -> 0.6588 and apply percent format
                cell.value = float(percent_0_to_100) / 100.0
                # Use conditional format: whole numbers without .00, decimals with 2 places
                if percent_0_to_100 == int(percent_0_to_100):
                    cell.number_format = '0%'  # 75% instead of 75.00%
                else:
                    cell.number_format = '0.00%'  # 75.5% for decimals
        except Exception:
            cell.value = None
        cell.alignment = Alignment(horizontal=("left" if align_left else "right"), vertical="center")
        return cell
    
    @staticmethod
    def export_student_report(report_data):
        """Export individual student report to Excel in a single sheet mirroring the PDF layout."""
        try:
            wb = ExcelExportService.create_workbook()

            # Single sheet for the entire report
            ws = wb.active
            ws.title = "Student Report"

            student = report_data['student']

            # Student info table (Field | Value)
            student_headers = ['Field', 'Value']
            ExcelExportService.style_header_row(ws, 1, student_headers)

            ws.cell(row=2, column=1, value="Name")
            ws.cell(row=2, column=2, value=student['name'])
            ws.cell(row=3, column=1, value="Roll Number")
            ws.cell(row=3, column=2, value=student['roll_number'])
            ws.cell(row=4, column=1, value="Course")
            ws.cell(row=4, column=2, value=student.get('course_display') or student.get('course'))
            row = 5
            if student.get('section'):
                ws.cell(row=row, column=1, value="Section")
                ws.cell(row=row, column=2, value=str(student.get('section')).upper())
                row += 1
            ws.cell(row=row, column=1, value="Academic Year")
            ws.cell(row=row, column=2, value=student['academic_year'])
            row += 1
            ws.cell(row=row, column=1, value="Current Semester")
            ws.cell(row=row, column=2, value=student['current_semester'])

            # Gap
            row += 2

            # Marks Report section
            ws.cell(row=row, column=1, value="Marks Report").font = Font(bold=True)
            row += 1
            marks_headers = ['Subject', 'Code', 'Assessment', 'Marks', 'Max', 'Percent', 'Grade', 'Status']
            ExcelExportService.style_header_row(ws, row, marks_headers)
            row += 1

            marks_row_count = 0
            for subject in report_data.get('subjects', []):
                for mark in subject.get('marks', []):
                    ws.cell(row=row, column=1, value=subject.get('subject_name'))
                    ws.cell(row=row, column=2, value=subject.get('subject_code'))
                    ws.cell(row=row, column=3, value=mark.get('assessment_type'))
                    ExcelExportService.set_number(ws.cell(row=row, column=4), mark.get('marks_obtained'), align_right=True)
                    ExcelExportService.set_number(ws.cell(row=row, column=5), mark.get('max_marks'), align_right=True)
                    ExcelExportService.set_percentage(ws.cell(row=row, column=6), mark.get('percentage'), align_left=True)
                    ws.cell(row=row, column=7, value=mark.get('grade'))
                    ws.cell(row=row, column=8, value=mark.get('performance_status'))
                    row += 1
                    marks_row_count += 1
            if marks_row_count == 0:
                ws.cell(row=row, column=1, value="No data")
                row += 1

            # Gap
            row += 2

            # Attendance Report section
            ws.cell(row=row, column=1, value="Attendance Report").font = Font(bold=True)
            row += 1
            attendance_headers = ['Subject', 'Code', 'Total', 'Present', 'Absent', 'Percent', 'Status']
            ExcelExportService.style_header_row(ws, row, attendance_headers)
            row += 1

            att_row_count = 0
            for subject in report_data.get('subjects', []):
                attendance = subject.get('attendance', {})
                ws.cell(row=row, column=1, value=subject.get('subject_name'))
                ws.cell(row=row, column=2, value=subject.get('subject_code'))
                ExcelExportService.set_number(ws.cell(row=row, column=3), attendance.get('total_classes'), align_right=True)
                ExcelExportService.set_number(ws.cell(row=row, column=4), attendance.get('present_classes'), align_right=True)
                ExcelExportService.set_number(ws.cell(row=row, column=5), attendance.get('absent_classes'), align_right=True)
                ExcelExportService.set_percentage(ws.cell(row=row, column=6), attendance.get('attendance_percentage'), align_left=True)
                status = "Good" if (attendance.get('attendance_percentage') or 0) >= 75 else (
                    "Poor" if (attendance.get('attendance_percentage') or 0) < 50 else "Average")
                ws.cell(row=row, column=7, value=status)
                row += 1
                att_row_count += 1
            if att_row_count == 0:
                ws.cell(row=row, column=1, value="No data")
                row += 1

            # Center everything and auto-adjust columns
            ExcelExportService.center_all_cells(ws)
            ExcelExportService.auto_adjust_columns(ws)

            return wb

        except Exception as e:
            print(f"Error exporting student report: {e}")
            return None
    
    @staticmethod
    def export_class_marks_report(report_data):
        """Export class marks report to Excel with clean format (no college headers)"""
        try:
            wb = ExcelExportService.create_workbook()
            ws = wb.active
            ws.title = "Class Marks Report"
            
            # Clean subject info table
            subject = report_data['subject']
            subject_headers = ['Field', 'Value']
            ExcelExportService.style_header_row(ws, 1, subject_headers)
            
            ws.cell(row=2, column=1, value="Subject")
            ws.cell(row=2, column=2, value=f"{subject['name']} ({subject['code']})")
            ws.cell(row=3, column=1, value="Course")
            ws.cell(row=3, column=2, value=subject.get('course_display') or subject.get('course'))
            ws.cell(row=4, column=1, value="Year/Semester")
            ws.cell(row=4, column=2, value=f"{subject['year']}/{subject['semester']}")
            # Optional section line
            if subject.get('section'):
                ws.cell(row=5, column=1, value="Section")
                ws.cell(row=5, column=2, value=str(subject.get('section')).upper())
                next_row = 6
            else:
                next_row = 5
            if subject.get('lecturers'):
                ws.cell(row=next_row, column=1, value="Faculty")
                ws.cell(row=next_row, column=2, value=", ".join(subject.get('lecturers') or []))
            
            if report_data['assessment_type']:
                ws.cell(row=5, column=1, value="Assessment Type")
                ws.cell(row=5, column=2, value=report_data['assessment_type'])
            
            # Clean statistics table
            stats = report_data['statistics']
            stats_headers = ['Statistic', 'Value']
            ExcelExportService.style_header_row(ws, next_row + 2, stats_headers)
            
            base = next_row + 3
            ws.cell(row=base + 0, column=1, value="Total Students")
            ws.cell(row=base + 0, column=2, value=stats['total_students'])
            ws.cell(row=base + 1, column=1, value="Class Average")
            ExcelExportService.set_percentage(ws.cell(row=base + 1, column=2), stats['class_average'], align_left=True)
            ws.cell(row=base + 2, column=1, value="Highest Score")
            ExcelExportService.set_percentage(ws.cell(row=base + 2, column=2), stats['highest_score'], align_left=True)
            ws.cell(row=base + 3, column=1, value="Lowest Score")
            ExcelExportService.set_percentage(ws.cell(row=base + 3, column=2), stats['lowest_score'], align_left=True)
            ws.cell(row=base + 4, column=1, value="Passing Students")
            ws.cell(row=base + 4, column=2, value=stats['passing_students'])
            ws.cell(row=base + 5, column=1, value="Failing Students")
            ws.cell(row=base + 5, column=2, value=stats['failing_students'])
            
            # Student marks table
            # Include Assessment Type column only when "All Assessments" is selected
            if report_data.get('assessment_type') == 'all' or not report_data.get('assessment_type'):
                headers = ['Roll Number', 'Student Name', 'Assessment Type', 'Marks Obtained', 
                          'Max Marks', 'Percentage', 'Grade', 'Status']
            else:
                headers = ['Roll Number', 'Student Name', 'Marks Obtained', 
                          'Max Marks', 'Percentage', 'Grade', 'Status']
            
            ExcelExportService.style_header_row(ws, base + 7, headers)
            
            row = base + 8
            for student_data in report_data['student_marks']:
                student = student_data['student']
                for mark in student_data['marks']:
                    ws.cell(row=row, column=1, value=student.roll_number)
                    ws.cell(row=row, column=2, value=student.name)
                    
                    # Only include Assessment Type column when "All Assessments" is selected
                    if report_data.get('assessment_type') == 'all' or not report_data.get('assessment_type'):
                        ws.cell(row=row, column=3, value=mark['assessment_type'])
                        col_offset = 1
                    else:
                        col_offset = 0
                    
                    ws.cell(row=row, column=3 + col_offset, value=ExcelExportService.format_number(mark['marks_obtained']))
                    ws.cell(row=row, column=4 + col_offset, value=ExcelExportService.format_number(mark['max_marks']))
                    ExcelExportService.set_percentage(ws.cell(row=row, column=5 + col_offset), mark['percentage'], align_left=True)
                    ws.cell(row=row, column=6 + col_offset, value=mark['grade'])
                    ws.cell(row=row, column=7 + col_offset, value=mark['performance_status'])
                    row += 1
            
            ExcelExportService.center_all_cells(ws)
            ExcelExportService.auto_adjust_columns(ws)
            return wb
            
        except Exception as e:
            print(f"Error exporting class marks report: {e}")
            return None
    
    @staticmethod
    def export_class_attendance_report(report_data):
        """Export class attendance report to Excel with clean format (no college headers)"""
        try:
            wb = ExcelExportService.create_workbook()
            ws = wb.active
            ws.title = "Class Attendance Report"
            
            # Clean subject info table
            subject = report_data['subject']
            subject_headers = ['Field', 'Value']
            ExcelExportService.style_header_row(ws, 1, subject_headers)
            
            ws.cell(row=2, column=1, value="Subject")
            ws.cell(row=2, column=2, value=f"{subject['name']} ({subject['code']})")
            ws.cell(row=3, column=1, value="Course")
            course_line = subject.get('course_display') or subject.get('course') or ''
            ws.cell(row=3, column=2, value=course_line)
            ws.cell(row=4, column=1, value="Year/Semester")
            ws.cell(row=4, column=2, value=f"{subject['year']}/{subject['semester']}")
            ws.cell(row=5, column=1, value="Period")
            # Format period as "Month Name - Year" or "Overall"
            if report_data.get('month') == 'overall' or report_data.get('month') == 'Overall':
                period_display = "Overall"
            else:
                month_name = report_data.get('month', '')
                year = report_data.get('year', '')
                period_display = f"{month_name} - {year}"
            ws.cell(row=5, column=2, value=period_display)
            # Optional section line
            if subject.get('section'):
                ws.cell(row=6, column=1, value="Section")
                ws.cell(row=6, column=2, value=str(subject.get('section')).upper())
                next_row2 = 7
            else:
                next_row2 = 6
            if subject.get('lecturers'):
                ws.cell(row=next_row2, column=1, value="Faculty")
                ws.cell(row=next_row2, column=2, value=", ".join(subject.get('lecturers') or []))
            
            # Clean statistics table
            stats = report_data['statistics']
            stats_headers = ['Statistic', 'Value']
            ExcelExportService.style_header_row(ws, next_row2 + 2, stats_headers)
            
            base2 = next_row2 + 3
            ws.cell(row=base2 + 0, column=1, value="Total Students")
            ws.cell(row=base2 + 0, column=2, value=stats['total_students'])
            ws.cell(row=base2 + 1, column=1, value="Total Classes Conducted")
            ws.cell(row=base2 + 1, column=2, value=stats['total_classes_conducted'])
            ws.cell(row=base2 + 2, column=1, value="Class Average Attendance")
            ExcelExportService.set_percentage(ws.cell(row=base2 + 2, column=2), stats['class_average_attendance'], align_left=True)
            ws.cell(row=base2 + 3, column=1, value="Good Attendance (≥75%)")
            ws.cell(row=base2 + 3, column=2, value=stats['students_with_good_attendance'])
            ws.cell(row=base2 + 4, column=1, value="Poor Attendance (<50%)")
            ws.cell(row=base2 + 4, column=2, value=stats['students_with_poor_attendance'])
            
            # Student attendance table
            headers = ['Roll Number', 'Student Name', 'Total Classes', 'Present', 
                      'Absent', 'Attendance %', 'Status']
            ExcelExportService.style_header_row(ws, base2 + 6, headers)
            
            row = base2 + 7
            for student in report_data['student_attendance']:
                ws.cell(row=row, column=1, value=student['roll_number'])
                ws.cell(row=row, column=2, value=student['student_name'])
                ExcelExportService.set_number(ws.cell(row=row, column=3), student['total_classes'], align_right=True)
                ExcelExportService.set_number(ws.cell(row=row, column=4), student['present_classes'], align_right=True)
                ExcelExportService.set_number(ws.cell(row=row, column=5), student['absent_classes'], align_right=True)
                ExcelExportService.set_percentage(ws.cell(row=row, column=6), student['attendance_percentage'], align_left=True)
                ws.cell(row=row, column=7, value=student['status'])
                row += 1
            
            ExcelExportService.center_all_cells(ws)
            ExcelExportService.auto_adjust_columns(ws)
            return wb
            
        except Exception as e:
            print(f"Error exporting class attendance report: {e}")
            return None
    
    @staticmethod
    def export_course_overview_report(report_data):
        """Export course overview report to Excel with clean format (no college headers)"""
        try:
            wb = ExcelExportService.create_workbook()
            ws = wb.active
            ws.title = "Course Overview"
            
            # Clean course info table
            course = report_data['course']
            course_headers = ['Field', 'Value']
            ExcelExportService.style_header_row(ws, 1, course_headers)
            
            ws.cell(row=2, column=1, value="Class")
            ws.cell(row=2, column=2, value=course['name'])
            ws.cell(row=3, column=1, value="Duration")
            ws.cell(row=3, column=2, value=f"{course['duration_years']} years ({course['total_semesters']} semesters)")
            ws.cell(row=4, column=1, value="Total Students")
            ws.cell(row=4, column=2, value=report_data['total_students'])
            ws.cell(row=5, column=1, value="Total Subjects")
            ws.cell(row=5, column=2, value=report_data['total_subjects'])
            
            # Subject-wise report
            headers = ['Subject Name', 'Subject Code', 'Enrolled Students', 'Average Marks %', 'Passing Rate %', 'Average Attendance %']
            ExcelExportService.style_header_row(ws, 7, headers)
            
            row = 8
            for subject in report_data['subjects']:
                ws.cell(row=row, column=1, value=subject['subject_name'])
                ws.cell(row=row, column=2, value=subject['subject_code'])
                ws.cell(row=row, column=3, value=subject['enrolled_students'])
                ExcelExportService.set_percentage(ws.cell(row=row, column=4), subject['marks_statistics']['average_marks'], align_left=True)
                ExcelExportService.set_percentage(ws.cell(row=row, column=5), subject['marks_statistics']['passing_rate'], align_left=True)
                ExcelExportService.set_percentage(ws.cell(row=row, column=6), subject['attendance_statistics']['average_attendance'], align_left=True)
                row += 1
            
            ExcelExportService.center_all_cells(ws)
            ExcelExportService.auto_adjust_columns(ws)
            return wb
            
        except Exception as e:
            print(f"Error exporting course overview report: {e}")
            return None
    
    # ======================== LECTURER SHORTAGE/DEFICIENCY EXPORTS ========================
    @staticmethod
    def export_attendance_shortage(threshold, shortage_data, lecturer_name=None, selected_subject_id=None):
        """Export Attendance Shortage (lecturer view) to Excel.
        shortage_data: [{ 'subject': Subject, 'shortage_students': [ {student, present_classes, total_classes, attendance_percentage}, ... ] }]
        Returns workbook bytes.
        """
        try:
            wb = ExcelExportService.create_workbook()
            ws = wb.active
            ws.title = "Attendance Shortage"

            # Title and meta rows
            ws['A1'] = 'Attendance Shortage Report'
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:F1')
            
            # Lecturer and threshold info
            meta_row = 2
            if lecturer_name:
                ws['A2'] = 'Lecturer'
                ws['B2'] = lecturer_name
                meta_row = 3
            ws[f'A{meta_row}'] = 'Threshold'
            ws[f'B{meta_row}'] = f"{threshold}%"

            # Start row for first subject section
            row = (meta_row + 1)
            # Build a section (header + table) for EACH subject block
            for idx, block in enumerate(shortage_data or []):
                subj = block.get('subject')
                course_name = subj.course.name if getattr(subj, 'course', None) else ''

                # Subject header: Subject, Code, Course (stacked)
                ws[f'A{row}'] = 'Subject'
                ws[f'B{row}'] = subj.name if subj else ''
                row += 1
                ws[f'A{row}'] = 'Code'
                ws[f'B{row}'] = getattr(subj, 'code', '')
                row += 1
                ws[f'A{row}'] = 'Course'
                ws[f'B{row}'] = course_name
                row += 1

                # Section table header (no Subject/Code/Course columns in table)
                headers = ['Student', 'Roll Number', 'Present', 'Total', 'Percent']
                ExcelExportService.style_header_row(ws, row, headers)
                row += 1

                # Sort students by roll number (last 3 digits)
                shortage_students = block.get('shortage_students') or []
                def get_roll_sort_key(rec):
                    roll_number = rec['student'].roll_number
                    if len(roll_number) >= 3:
                        last_three = roll_number[-3:]
                        try:
                            return int(last_three)
                        except ValueError:
                            return 999
                    return 999
                
                sorted_students = sorted(shortage_students, key=get_roll_sort_key)
                
                # Section rows
                for rec in sorted_students:
                    ws.cell(row=row, column=1, value=rec['student'].name)
                    ws.cell(row=row, column=2, value=rec['student'].roll_number)
                    ExcelExportService.set_number(ws.cell(row=row, column=3), rec.get('present_classes') or 0, align_right=True)
                    ExcelExportService.set_number(ws.cell(row=row, column=4), rec.get('total_classes') or 0, align_right=True)
                    ExcelExportService.set_percentage(ws.cell(row=row, column=5), rec.get('attendance_percentage') or 0, align_left=True)
                    row += 1

                # Blank row between sections
                row += 1

            ExcelExportService.center_all_cells(ws)
            ExcelExportService.auto_adjust_columns(ws)

            # Wrap and left-align Student column in each section so long names stay within the cell
            from openpyxl.styles import Alignment
            last_row = ws.max_row or row - 1
            # Student col is the 1st column in our section tables
            for r in range(1, last_row + 1):
                cell = ws.cell(row=r, column=1)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            return ExcelExportService.workbook_to_bytes(wb)
        except Exception as e:
            print(f"Error exporting attendance shortage: {e}")
            return None

    @staticmethod
    def export_marks_deficiency(threshold, deficiency_data, lecturer_name=None, selected_subject_id=None):
        """Export Marks Deficiency (lecturer view) to Excel.
        deficiency_data: [{ 'subject': Subject, 'deficient_students': [ {student, overall_percentage, marks_summary}, ... ] }]
        Returns workbook bytes.
        """
        try:
            wb = ExcelExportService.create_workbook()
            ws = wb.active
            ws.title = "Marks Deficiency"

            # Title and meta rows
            ws['A1'] = 'Marks Deficiency Report'
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:G1')
            
            # Lecturer and threshold info
            meta_row = 2
            if lecturer_name:
                ws['A2'] = 'Lecturer'
                ws['B2'] = lecturer_name
                meta_row = 3
            ws[f'A{meta_row}'] = 'Threshold'
            ws[f'B{meta_row}'] = f"{threshold}%"

            # Create separate tables for each subject
            current_row = meta_row + 1
            
            # Sort deficiency_data by subject name for consistent ordering
            sorted_deficiency_data = sorted(deficiency_data or [], key=lambda x: getattr(x.get('subject'), 'name', '') if x.get('subject') else '')
            
            for block_idx, block in enumerate(sorted_deficiency_data):
                subj = block.get('subject')
                if not subj:
                    continue
                    
                course_name = subj.course.name if getattr(subj, 'course', None) else ''
                
                # Subject details above each table
                if block_idx > 0:
                    current_row += 2  # Add spacing between tables
                
                ws[f'A{current_row}'] = 'Subject'
                ws[f'B{current_row}'] = subj.name
                current_row += 1
                ws[f'A{current_row}'] = 'Code'
                ws[f'B{current_row}'] = subj.code
                current_row += 1
                ws[f'A{current_row}'] = 'Course'
                ws[f'B{current_row}'] = course_name
                current_row += 1
                
                # Decide which mark components to include based on actual data present
                deficient_students = block.get('deficient_students') or []
                include = { 'internal1': False, 'internal2': False, 'assignment': False, 'project': False }
                def _is_updated(val):
                    try:
                        if isinstance(val, dict):
                            obt = val.get('obtained')
                            mx = val.get('max')
                        else:
                            obt = getattr(val, 'obtained', None)
                            mx = getattr(val, 'max', None)
                        # Consider updated only if any value is numeric and > 0
                        obt_num = float(obt) if obt is not None else 0.0
                        mx_num = float(mx) if mx is not None else 0.0
                        return (obt_num > 0.0) or (mx_num > 0.0)
                    except Exception:
                        return False
                for rec in deficient_students:
                    ms = rec.get('marks_summary') or {}
                    for k in include.keys():
                        if _is_updated(ms.get(k)):
                            include[k] = True

                headers = ['Student', 'Roll Number', 'Overall %']
                component_to_header = {
                    'internal1': 'Internal 1',
                    'internal2': 'Internal 2',
                    'assignment': 'Assignment',
                    'project': 'Project',
                }
                ordered_components = [k for k in ['internal1','internal2','assignment','project'] if include.get(k)]
                headers.extend([component_to_header[k] for k in ordered_components])
                ExcelExportService.style_header_row(ws, current_row, headers)
                current_row += 1
                
                # Sort students by roll number (last 3 digits)
                def get_roll_sort_key(rec):
                    roll_number = rec['student'].roll_number
                    if len(roll_number) >= 3:
                        last_three = roll_number[-3:]
                        try:
                            return int(last_three)
                        except ValueError:
                            return 999
                    return 999
                
                sorted_students = sorted(deficient_students, key=get_roll_sort_key)
                
                # Data rows for this subject
                for rec in sorted_students:
                    ms = rec.get('marks_summary') or {}
                    def _fmt(a):
                        try:
                            if isinstance(a, dict):
                                obt = a.get('obtained')
                                mx = a.get('max')
                            else:
                                obt = getattr(a, 'obtained', None)
                                mx = getattr(a, 'max', None)
                            if obt is None and mx is None:
                                return ''
                            f_obt = ExcelExportService.format_number(obt)
                            f_max = ExcelExportService.format_number(mx)
                            if f_obt is None and f_max is None:
                                return ''
                            if f_obt is None:
                                f_obt = ''
                            if f_max is None:
                                f_max = ''
                            return f"{f_obt}/{f_max}"
                        except Exception:
                            return ''

                    col = 1
                    ws.cell(row=current_row, column=col, value=rec['student'].name); col += 1
                    ws.cell(row=current_row, column=col, value=rec['student'].roll_number); col += 1
                    overall_percentage = rec.get('overall_percentage')
                    if overall_percentage is not None:
                        ExcelExportService.set_percentage(ws.cell(row=current_row, column=col), overall_percentage, align_left=True)
                    else:
                        ws.cell(row=current_row, column=col, value="-")
                    col += 1
                    for comp in ordered_components:
                        ws.cell(row=current_row, column=col, value=_fmt(ms.get(comp))); col += 1
                    current_row += 1

            ExcelExportService.center_all_cells(ws)
            ExcelExportService.auto_adjust_columns(ws)

            return ExcelExportService.workbook_to_bytes(wb)
        except Exception as e:
            print(f"Error exporting marks deficiency: {e}")
            return None
    
    @staticmethod
    def workbook_to_bytes(workbook):
        """Convert workbook to bytes for download"""
        try:
            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            return output.getvalue()
        except Exception as e:
            print(f"Error converting workbook to bytes: {e}")
            return None

    @staticmethod
    def export_comprehensive_class_report(report):
        """Export comprehensive class report to Excel with full width utilization"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from io import BytesIO
            
            workbook = Workbook()
            ws = workbook.active
            
            # Set title
            course_name = report['course']['name']
            report_type = report['report_type'].title()
            ws.title = f"{report_type} Report - {course_name}"
            
            # Header
            ws.merge_cells('A1:Z1')
            ws['A1'] = f"Comprehensive {report_type} Report - {course_name}"
            ws['A1'].font = Font(name='Arial', size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Course info
            ws['A2'] = f"Course: {course_name} ({report['course']['code']})"
            ws['A3'] = f"Total Students: {len(report['students'])}"
            ws['A4'] = f"Total Subjects: {len(report['subjects'])}"
            
            # Add Assessment Type for marks reports
            if report['report_type'] == 'marks' and report.get('assessment_type'):
                assessment_display = report['assessment_type'].title().replace('1', ' 1').replace('2', ' 2')
                ws['A5'] = f"Assessment Type: {assessment_display}"
                header_rows = 5
            else:
                header_rows = 4
            
            # Set header styles
            for row in range(1, header_rows + 1):
                ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True)
            
            # Start data from row after headers
            current_row = header_rows + 2
            
            if report['report_type'] == 'attendance':
                # Create attendance report
                headers = ['Roll No', 'Student Name'] + [subj['name'] for subj in report['subjects']]
                
                # Add headers to the worksheet
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                
                # Style headers with blue background and center alignment
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                current_row += 1
                
                # Add student data
                for student in report['students']:
                    row_data = [student['roll_number'], student['name']]
                    
                    for subject in report['subjects']:
                        subject_data = report['data'].get(subject['id'], {})
                        student_attendance = subject_data.get('student_attendance', {}).get(student['id'], {})
                        
                        # Check if student is enrolled in this subject
                        is_enrolled = subject_data.get('enrolled_students', {}).get(student['id'], False)
                        
                        if not is_enrolled:
                            row_data.append("NA")
                        elif student_attendance and student_attendance.get('total_classes', 0) > 0:
                            percentage = student_attendance.get('percentage', 0)
                            row_data.append(f"{percentage}%")
                        else:
                            row_data.append("-")
                    
                    ws.append(row_data)
                    
                    # Style data rows
                    for col in range(1, len(row_data) + 1):
                        cell = ws.cell(row=current_row, column=col)
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        # Color code attendance percentages
                        if col > 2 and row_data[col-1] != "N/A":
                            try:
                                percentage = float(row_data[col-1].split('(')[1].split('%')[0])
                                if percentage >= 75:
                                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                                elif percentage >= 60:
                                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                                else:
                                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                            except:
                                pass
                    
                    current_row += 1
            
            elif report['report_type'] == 'marks':
                # Create marks report
                headers = ['Roll No', 'Student Name'] + [subj['name'] for subj in report['subjects']]
                
                # Add headers to the worksheet
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                
                # Style headers with blue background and center alignment
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                current_row += 1
                
                # Add student data
                for student in report['students']:
                    row_data = [student['roll_number'], student['name']]
                    
                    for subject in report['subjects']:
                        subject_data = report['data'].get(subject['id'], {})
                        student_marks = subject_data.get('student_marks', {}).get(student['id'], {})
                        
                        # Check if student is enrolled in this subject
                        is_enrolled = subject_data.get('enrolled_students', {}).get(student['id'], False)
                        
                        if not is_enrolled:
                            row_data.append("NA")
                        elif student_marks and report.get('assessment_type'):
                            # Specific assessment type selected
                            assessment_data = student_marks.get(report['assessment_type'], {})
                            if assessment_data and assessment_data.get('recorded', False) and assessment_data.get('max', 0) > 0:
                                obtained = assessment_data.get('obtained', 0)
                                max_marks = assessment_data.get('max', 0)
                                # Format marks: remove .0 from whole numbers
                                obtained_str = str(int(obtained)) if obtained == int(obtained) else str(obtained)
                                max_str = str(int(max_marks)) if max_marks == int(max_marks) else str(max_marks)
                                row_data.append(f"{obtained_str}/{max_str}")
                            else:
                                row_data.append("-")
                        else:
                            row_data.append("-")
                    
                    ws.append(row_data)
                    
                    # Style data rows
                    for col in range(1, len(row_data) + 1):
                        cell = ws.cell(row=current_row, column=col)
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        # Color code marks based on obtained/max format
                        if col > 2 and row_data[col-1] not in ["N/A", "NA", "-"]:
                            try:
                                # Parse "obtained/max" format
                                if '/' in row_data[col-1]:
                                    obtained, max_marks = row_data[col-1].split('/')
                                    obtained = float(obtained)
                                    max_marks = float(max_marks)
                                    if max_marks > 0:
                                        percentage = (obtained / max_marks) * 100
                                        if percentage >= 60:
                                            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                                        elif percentage >= 40:
                                            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                                        else:
                                            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                            except:
                                pass
                    
                    current_row += 1
            
            # Auto-adjust column widths for full screen utilization
            ExcelExportService.auto_adjust_columns(ws)
            
            # Convert to bytes
            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            print(f"Error in export_comprehensive_class_report: {e}")
            return None