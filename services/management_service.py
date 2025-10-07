"""
Management service for Moulya College Management System
Business logic for management portal operations
"""

from models.user import Lecturer
from models.academic import Course, Subject, AcademicYear
from models.student import Student, StudentEnrollment
from models.assignments import SubjectAssignment
from database import db
from utils.db_helpers import safe_add_and_commit, bulk_insert
from utils.validators import *
from utils.sorting_helpers import SortingHelpers
from services.auth_service import AuthService
import openpyxl
from io import BytesIO

class ManagementService:
    """Management service class"""
    
    @staticmethod
    def get_dashboard_stats():
        """Get dashboard statistics with sorted ordering"""
        try:
            # Get lecturers with custom sorting
            all_lecturers = Lecturer.query.filter_by(is_active=True).all()
            sorted_lecturers = SortingHelpers.sort_lecturers(all_lecturers)
            recent_lecturers = sorted_lecturers[:5]
            
            # Get students with custom sorting
            all_students = Student.query.filter_by(is_active=True).all()
            sorted_students = SortingHelpers.sort_students(all_students)
            recent_students = sorted_students[:5]
            
            stats = {
                'total_lecturers': len(sorted_lecturers),
                'total_students': len(sorted_students),
                'total_courses': Course.query.filter_by(is_active=True).count(),
                'total_subjects': Subject.query.filter_by(is_active=True).count(),
                'recent_lecturers': recent_lecturers,
                'recent_students': recent_students
            }
            return stats
        except Exception as e:
            return {}
    
    @staticmethod
    def get_lecturers_paginated(page=1, search='', per_page=20):
        """Get paginated lecturers list with sorted ordering"""
        try:
            query = Lecturer.query.filter_by(is_active=True)
            
            if search:
                query = query.filter(
                    db.or_(
                        Lecturer.name.contains(search),
                        Lecturer.lecturer_id.contains(search),
                        Lecturer.username.contains(search)
                    )
                )
            
            # Get all lecturers and sort them using helper
            all_lecturers = query.all()
            sorted_lecturers = SortingHelpers.sort_lecturers(all_lecturers)
            
            # Manual pagination
            start_idx = (page - 1) * per_page
            end_idx = start_idx + per_page
            paginated_lecturers = sorted_lecturers[start_idx:end_idx]
            
            # Create pagination object manually
            total = len(sorted_lecturers)
            has_prev = page > 1
            has_next = end_idx < total
            
            class PaginationObject:
                def __init__(self, items, page, per_page, total, has_prev, has_next):
                    self.items = items
                    self.page = page
                    self.per_page = per_page
                    self.total = total
                    self.pages = (total + per_page - 1) // per_page
                    self.has_prev = has_prev
                    self.has_next = has_next
                    self.prev_num = page - 1 if has_prev else None
                    self.next_num = page + 1 if has_next else None
                
                def iter_pages(self, left_edge=2, right_edge=2, left_current=2, right_current=3):
                    """Generate page numbers for pagination display"""
                    last = self.pages
                    for num in range(1, last + 1):
                        if num <= left_edge or \
                           (num > self.page - left_current - 1 and num < self.page + right_current) or \
                           num > last - right_edge:
                            yield num
            
            pagination = PaginationObject(
                items=paginated_lecturers,
                page=page,
                per_page=per_page,
                total=total,
                has_prev=has_prev,
                has_next=has_next
            )
            
            return {
                'lecturers': paginated_lecturers,
                'pagination': pagination
            }
        except Exception as e:
            return {'lecturers': [], 'pagination': None}
    
    @staticmethod
    def get_students_paginated(page=1, search='', course_id=None, per_page=20):
        """Get paginated students list with sorted ordering by course"""
        try:
            query = Student.query.filter_by(is_active=True)
            
            if search:
                query = query.filter(
                    db.or_(
                        Student.name.contains(search),
                        Student.roll_number.contains(search)
                    )
                )
            
            if course_id:
                query = query.filter_by(course_id=course_id)
            
            # Get all students and sort them using helper
            all_students = query.all()
            sorted_students = SortingHelpers.sort_students(all_students)
            
            # Manual pagination
            start_idx = (page - 1) * per_page
            end_idx = start_idx + per_page
            paginated_students = sorted_students[start_idx:end_idx]
            
            # Create pagination object manually
            total = len(sorted_students)
            has_prev = page > 1
            has_next = end_idx < total
            
            class PaginationObject:
                def __init__(self, items, page, per_page, total, has_prev, has_next):
                    self.items = items
                    self.page = page
                    self.per_page = per_page
                    self.total = total
                    self.pages = (total + per_page - 1) // per_page
                    self.has_prev = has_prev
                    self.has_next = has_next
                    self.prev_num = page - 1 if has_prev else None
                    self.next_num = page + 1 if has_next else None
                
                def iter_pages(self, left_edge=2, right_edge=2, left_current=2, right_current=3):
                    """Generate page numbers for pagination display"""
                    last = self.pages
                    for num in range(1, last + 1):
                        if num <= left_edge or \
                           (num > self.page - left_current - 1 and num < self.page + right_current) or \
                           num > last - right_edge:
                            yield num
            
            pagination = PaginationObject(
                items=paginated_students,
                page=page,
                per_page=per_page,
                total=total,
                has_prev=has_prev,
                has_next=has_next
            )
            
            return {
                'students': paginated_students,
                'pagination': pagination
            }
        except Exception as e:
            return {'students': [], 'pagination': None}
    
    @staticmethod
    def get_subjects_by_course(course_id=None):
        """Get subjects filtered by course"""
        try:
            if course_id:
                return Subject.query.filter_by(course_id=course_id, is_active=True).order_by(Subject.year, Subject.semester).all()
            else:
                return Subject.query.filter_by(is_active=True).order_by(Subject.year, Subject.semester).all()
        except Exception as e:
            return []
    
    @staticmethod
    def add_lecturer(lecturer_data):
        """Add single lecturer"""
        try:
            # Validate data
            is_valid, message = validate_lecturer_id(lecturer_data.get('lecturer_id'))
            if not is_valid:
                return False, message
            
            is_valid, message = validate_name(lecturer_data.get('name'))
            if not is_valid:
                return False, message
            
            # Check if lecturer ID already exists
            # If an ACTIVE lecturer exists, block; if an INACTIVE one exists, remove it to allow re-adding
            existing_active = Lecturer.query.filter_by(lecturer_id=lecturer_data['lecturer_id'], is_active=True).first()
            if existing_active:
                return False, "Lecturer ID already exists"

            existing_inactive = Lecturer.query.filter_by(lecturer_id=lecturer_data['lecturer_id'], is_active=False).first()
            if existing_inactive:
                try:
                    # Remove dependent rows and the inactive lecturer to free unique constraints
                    from models.assignments import SubjectAssignment
                    SubjectAssignment.query.filter_by(lecturer_id=existing_inactive.id).delete(synchronize_session=False)
                    from models.attendance import AttendanceRecord, MonthlyAttendanceSummary
                    from models.marks import StudentMarks
                    AttendanceRecord.query.filter_by(lecturer_id=existing_inactive.id).delete(synchronize_session=False)
                    MonthlyAttendanceSummary.query.filter_by(lecturer_id=existing_inactive.id).delete(synchronize_session=False)
                    StudentMarks.query.filter_by(lecturer_id=existing_inactive.id).delete(synchronize_session=False)
                    db.session.delete(existing_inactive)
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    return False, f"Error replacing inactive lecturer with same ID: {str(e)}"
            
            # Generate credentials
            success, username, password, msg = AuthService.generate_lecturer_credentials(
                lecturer_data['name'], 
                lecturer_data['lecturer_id'],
                lecturer_data.get('username'),
                lecturer_data.get('password')
            )
            
            if not success:
                return False, msg
            
            # Create lecturer
            lecturer = Lecturer(
                lecturer_id=lecturer_data['lecturer_id'],
                name=lecturer_data['name'],
                username=username
            )
            lecturer.set_password(password)
            
            success, message = safe_add_and_commit(lecturer)
            if success:
                # Create subject assignments if subjects were provided
                subject_ids = lecturer_data.get('subject_ids', [])
                if subject_ids:
                    from models.assignments import SubjectAssignment
                    from datetime import datetime
                    
                    current_year = datetime.now().year
                    assignments = []
                    
                    for subject_id in subject_ids:
                        # Check if assignment already exists
                        existing = SubjectAssignment.query.filter_by(
                            lecturer_id=lecturer.id,
                            subject_id=int(subject_id),
                            academic_year=current_year
                        ).first()
                        
                        if not existing:
                            assignment = SubjectAssignment(
                                lecturer_id=lecturer.id,
                                subject_id=int(subject_id),
                                academic_year=current_year
                            )
                            assignments.append(assignment)
                    
                    if assignments:
                        for assignment in assignments:
                            safe_add_and_commit(assignment)
                
                return True, f"Lecturer added successfully. Username: {username}, Password: {password}"
            else:
                return False, message
                
        except Exception as e:
            return False, f"Error adding lecturer: {str(e)}"
    
    @staticmethod
    def bulk_add_lecturers(file_data):
        """Bulk upsert lecturers from Excel file. Improved: auto-detect header, flexible columns, clear errors."""
        try:
            workbook = openpyxl.load_workbook(BytesIO(file_data))
            sheet = workbook.active

            # Detect header row and map columns
            header_row = None
            for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                header_row = [str(cell).strip().lower().replace(' ', '_') if cell else '' for cell in row]
                break
            # Map common header variations
            header_aliases = {
                'lecturer_id': ['lecturer_id', 'lecturerid', 'lecturer id', 'id'],
                'name': ['name', 'full_name', 'fullname', 'full name']
            }
            col_map = {}
            for key, aliases in header_aliases.items():
                for alias in aliases:
                    if alias in header_row:
                        col_map[key] = header_row.index(alias)
                        break
            # Also map subject_codes if present
            for idx, col in enumerate(header_row):
                if col in ['subject_codes', 'subject code', 'subjects', 'subject']:  # Accept variations
                    col_map['subject_codes'] = idx
            required_cols = ['lecturer_id', 'name']
            missing_cols = [col for col in required_cols if col not in col_map]
            if missing_cols:
                return False, f"Missing required columns: {', '.join(missing_cols)}", [], []

            lecturers_to_create = []
            errors = []
            credentials = []
            updates_applied = 0

            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                try:
                    lecturer_id = str(row[col_map['lecturer_id']]).strip() if row[col_map['lecturer_id']] else None
                    name = str(row[col_map['name']]).strip() if row[col_map['name']] else None
                    subject_codes = None
                    if 'subject_codes' in col_map:
                        subject_codes = str(row[col_map['subject_codes']]).strip() if row[col_map['subject_codes']] and str(row[col_map['subject_codes']]).strip().lower() != 'none' else None

                    if not lecturer_id or not name:
                        errors.append(f"Row {row_num}: Lecturer ID and Name are required")
                        continue

                    # Parse subject codes (comma-separated)
                    subject_ids = []
                    if subject_codes:
                        subject_code_list = [code.strip() for code in subject_codes.split(',') if code.strip()]
                        for code in subject_code_list:
                            subject = Subject.query.filter_by(code=code, is_active=True).first()
                            if subject:
                                subject_ids.append(subject.id)
                            else:
                                errors.append(f"Row {row_num}: Subject code '{code}' not found")

                    # Validate data
                    is_valid, message = validate_lecturer_id(lecturer_id)
                    if not is_valid:
                        errors.append(f"Row {row_num}: Lecturer ID '{lecturer_id}' - {message}")
                        continue

                    is_valid, message = validate_name(name)
                    if not is_valid:
                        errors.append(f"Row {row_num}: Name '{name}' - {message}")
                        continue

                    # Upsert by lecturer_id
                    existing_active_lecturer = Lecturer.query.filter_by(lecturer_id=lecturer_id, is_active=True).first()
                    if existing_active_lecturer:
                        if existing_active_lecturer.name != name:
                            existing_active_lecturer.name = name
                            updates_applied += 1
                        if subject_ids:
                            from models.assignments import SubjectAssignment
                            from datetime import datetime
                            current_year = datetime.now().year
                            existing_assignments = SubjectAssignment.query.filter_by(
                                lecturer_id=existing_active_lecturer.id,
                                academic_year=current_year
                            ).all()
                            existing_subject_ids = {a.subject_id for a in existing_assignments if a.is_active}
                            desired_subject_ids = set(int(sid) for sid in subject_ids)
                            to_add = desired_subject_ids - existing_subject_ids
                            for sid in to_add:
                                assignment = SubjectAssignment(
                                    lecturer_id=existing_active_lecturer.id,
                                    subject_id=int(sid),
                                    academic_year=current_year
                                )
                                db.session.add(assignment)
                            to_deactivate = existing_subject_ids - desired_subject_ids
                            if to_deactivate:
                                for a in existing_assignments:
                                    if a.subject_id in to_deactivate and a.is_active:
                                        a.is_active = False
                    else:
                        existing_inactive_lecturer = Lecturer.query.filter_by(lecturer_id=lecturer_id, is_active=False).first()
                        if existing_inactive_lecturer:
                            from models.assignments import SubjectAssignment
                            SubjectAssignment.query.filter_by(lecturer_id=existing_inactive_lecturer.id).delete()
                            db.session.delete(existing_inactive_lecturer)
                        success, username, password, msg = AuthService.generate_lecturer_credentials(name, lecturer_id)
                        if not success:
                            errors.append(f"Row {row_num}: {msg}")
                            continue
                        lecturer = Lecturer(
                            lecturer_id=lecturer_id,
                            name=name,
                            username=username
                        )
                        lecturer.set_password(password)
                        lecturer._temp_subject_ids = subject_ids
                        lecturers_to_create.append(lecturer)
                        credentials.append({
                            'lecturer_id': lecturer_id,
                            'name': name,
                            'username': username,
                            'password': password
                        })
                except Exception as e:
                    errors.append(f"Row {row_num}: Error processing data - {str(e)}")

            if lecturers_to_create or updates_applied > 0:
                try:
                    if updates_applied > 0:
                        db.session.commit()
                    for lecturer in lecturers_to_create:
                        db.session.add(lecturer)
                    if lecturers_to_create:
                        db.session.commit()
                    message = []
                    if lecturers_to_create:
                        message.append(f"added {len(lecturers_to_create)}")
                    if updates_applied:
                        message.append(f"updated {updates_applied}")
                    message = "Successfully " + ", ".join(message) + " lecturer(s)"
                    from models.assignments import SubjectAssignment
                    from datetime import datetime
                    current_year = datetime.now().year
                    assignments = []
                    for lecturer in lecturers_to_create:
                        if hasattr(lecturer, '_temp_subject_ids') and lecturer._temp_subject_ids:
                            for subject_id in lecturer._temp_subject_ids:
                                existing = SubjectAssignment.query.filter_by(
                                    lecturer_id=lecturer.id,
                                    subject_id=subject_id,
                                    academic_year=current_year
                                ).first()
                                if not existing:
                                    assignment = SubjectAssignment(
                                        lecturer_id=lecturer.id,
                                        subject_id=subject_id,
                                        academic_year=current_year
                                    )
                                    assignments.append(assignment)
                    if assignments:
                        success_assign, message_assign = bulk_insert(assignments)
                    return True, message, credentials, errors
                except Exception as e:
                    db.session.rollback()
                    return False, f"Database error: {str(e)}", [], errors
            else:
                return False, "No valid lecturer changes found", [], errors
        except Exception as e:
            return False, f"Error processing Excel file: {str(e)}", [], []
    
    @staticmethod
    def add_student(student_data):
        """Add single student"""
        try:
            # Validate data
            is_valid, message = validate_roll_number(student_data.get('roll_number'))
            if not is_valid:
                return False, message
            
            is_valid, message = validate_name(student_data.get('name'))
            if not is_valid:
                return False, message
            
            is_valid, message = validate_academic_year(student_data.get('academic_year'))
            if not is_valid:
                return False, message
            
            # Check if roll number already exists
            existing_active = Student.query.filter_by(roll_number=student_data['roll_number'], is_active=True).first()
            if existing_active:
                return False, "Roll number already exists"

            # Remove inactive duplicate if present
            existing_inactive = Student.query.filter_by(roll_number=student_data['roll_number'], is_active=False).first()
            if existing_inactive:
                ok, msg = ManagementService.delete_student_permanently(existing_inactive.id)
                if not ok:
                    return False, msg
            
            # Validate course exists
            course = Course.query.get(student_data.get('course_id'))
            if not course:
                return False, "Invalid course selected"
            
            # Create student
            student = Student(
                roll_number=student_data['roll_number'],
                name=student_data['name'],
                course_id=student_data['course_id'],
                academic_year=student_data['academic_year'],
                current_semester=student_data.get('current_semester', 1),
                email=student_data.get('email')
            )
            
            success, message = safe_add_and_commit(student)
            return success, message
                
        except Exception as e:
            return False, f"Error adding student: {str(e)}"
    
    @staticmethod
    def bulk_add_students(file_data):
        """Bulk upsert students from Excel file.

        Behavior:
        - Match by `roll_number`.
        - If student exists (active), update fields: name, course, academic_year, email.
        - If student does not exist, create a new one.
        - If a course code is new (e.g., prefix-based), auto-create minimal course if needed.
        """
        try:
            workbook = openpyxl.load_workbook(BytesIO(file_data))
            sheet = workbook.active

            students_to_create = []
            errors = []
            updates_applied = 0

            # Read header row and map columns by header name (case-insensitive)
            header_row = [str(c.value).strip() if c.value is not None else '' for c in next(sheet.iter_rows(min_row=1, max_row=1))]
            header_to_index = {h.lower(): i for i, h in enumerate(header_row) if h}

            def find_col(*possible_names):
                for name in possible_names:
                    idx = header_to_index.get(name.lower())
                    if idx is not None:
                        return idx
                return None

            col_roll = find_col('roll number', 'roll_number', 'rollno', 'roll', 'roll no')
            col_name = find_col('name', 'full name', 'student name')
            col_course_code = find_col('course code', 'course', 'course_code', 'course name', 'course short code')
            col_ac_year = find_col('academic year', 'year', 'year level', 'class year')
            col_semester = find_col('semester', 'current semester', 'sem', 'current_semester')
            # Try exact-name matches first
            col_class = find_col('class', 'class name', 'year/class', 'year & course', 'class section', 'class & section')
            # If still not found, choose first header containing the word 'class'
            if col_class is None:
                for h, idx in header_to_index.items():
                    if 'class' in h:
                        col_class = idx
                        break
            col_email = find_col('email', 'email id', 'e-mail')

            # If headers are missing, fall back to legacy fixed positions (A..E)
            if col_roll is None:
                col_roll = 0
            if col_name is None:
                col_name = 1
            if col_course_code is None:
                col_course_code = 2
            if col_ac_year is None:
                col_ac_year = 3
            if col_email is None:
                col_email = 4

            import re

            def parse_year(value):
                if value is None:
                    return None
                try:
                    # Try numeric directly
                    return int(value)
                except Exception:
                    # Extract first digit 1-3 from strings like "1", "1st", "Year 1"
                    match = re.search(r"([1-3])", str(value))
                    return int(match.group(1)) if match else None

            ROMAN_TO_INT = { 'I': 1, 'II': 2, 'III': 3 }
            
            # Semester defaulting mapping based on academic year
            SEMESTER_DEFAULT_MAPPING = {1: 1, 2: 3, 3: 5}

            def parse_class(value):
                """Parse strings like 'II BCA B' → (course_code='BCA', academic_year=2).
                Section (e.g., 'B') is currently ignored.
                """
                if not value:
                    return None, None
                text = str(value).strip()
                # Try split by spaces, handle roman numerals or digits
                parts = text.split()
                year_val = None
                course_val = None
                if parts:
                    first = parts[0].upper()
                    if first in ROMAN_TO_INT:
                        year_val = ROMAN_TO_INT[first]
                        if len(parts) > 1:
                            course_val = re.sub(r"[^A-Za-z]", "", parts[1]).upper() or None
                    else:
                        # Try digits anywhere
                        y = parse_year(first)
                        if y:
                            year_val = y
                            if len(parts) > 1:
                                course_val = re.sub(r"[^A-Za-z]", "", parts[1]).upper() or None
                        else:
                            # Maybe 'BCA II B' format → find alpha block and roman/digit
                            m_course = re.search(r"([A-Za-z]+)", text)
                            m_year = re.search(r"\b(I{1,3}|[1-3])\b", text, re.IGNORECASE)
                            if m_year:
                                yr = m_year.group(1).upper()
                                year_val = ROMAN_TO_INT.get(yr, parse_year(yr))
                            if m_course:
                                course_val = m_course.group(1).upper()
                return course_val, year_val



            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue

                try:
                    roll_number = str(row[col_roll]).strip() if (col_roll is not None and col_roll < len(row) and row[col_roll]) else None
                    name = str(row[col_name]).strip() if col_name < len(row) and row[col_name] else None
                    course_code_raw = str(row[col_course_code]).strip() if (col_course_code is not None and col_course_code < len(row) and row[col_course_code]) else None
                    # Use course code exactly as it appears in Excel file
                    course_code = course_code_raw if course_code_raw else None
                    academic_year = parse_year(row[col_ac_year] if (col_ac_year is not None and col_ac_year < len(row)) else None)
                    # If missing, try parsing from combined class column
                    if (course_code is None or academic_year is None) and col_class is not None and col_class < len(row):
                        cc_from_class, year_from_class = parse_class(row[col_class])
                        if course_code is None:
                            course_code = cc_from_class
                        if academic_year is None:
                            academic_year = year_from_class
                    
                    # Parse semester from Excel file
                    semester = None
                    if col_semester is not None and col_semester < len(row) and row[col_semester] is not None:
                        try:
                            semester = int(row[col_semester])
                        except (ValueError, TypeError):
                            # Try to extract number from string
                            match = re.search(r"([1-6])", str(row[col_semester]))
                            if match:
                                semester = int(match.group(1))
                    
                    # Apply semester defaulting logic if semester not specified
                    if semester is None and academic_year is not None:
                        semester = SEMESTER_DEFAULT_MAPPING.get(academic_year, 1)

                    # Course code must be provided in Excel file - no fallback extraction

                    # If academic year is still None, default to 1 (so initial bulk upload doesn't drop rows)
                    if academic_year is None:
                        academic_year = 1
                    email_val = row[col_email] if col_email < len(row) else None
                    email = str(email_val).strip() if email_val and str(email_val).strip().lower() not in ['none', 'na', 'n/a'] else None

                    if not all([roll_number, name, course_code, academic_year]):
                        errors.append(f"Row {row_num}: Roll number, name, course code, and academic year are required")
                        continue

                    # Validate data
                    is_valid, message = validate_roll_number(roll_number)
                    if not is_valid:
                        errors.append(f"Row {row_num}: {message}")
                        continue

                    is_valid, message = validate_name(name)
                    if not is_valid:
                        errors.append(f"Row {row_num}: {message}")
                        continue

                    is_valid, message = validate_academic_year(academic_year)
                    if not is_valid:
                        errors.append(f"Row {row_num}: {message}")
                        continue

                    # Find course by normalized code; auto-create if not found
                    course = Course.query.filter(
                        db.func.lower(Course.code) == course_code.lower(),
                        Course.is_active == True
                    ).first()
                    if not course and course_code:
                        try:
                            new_course = Course(
                                name=course_code,  # This will be "I BCOM A" format
                                code=course_code,  # This will be "I BCOM A" format
                                description=f"Auto-created from bulk upload",
                                duration_years=3,
                                total_semesters=6,
                                is_active=True
                            )
                            db.session.add(new_course)
                            db.session.commit()
                            course = new_course
                        except Exception:
                            db.session.rollback()
                    if not course:
                        errors.append(f"Row {row_num}: Course code {course_code} not found")
                        continue

                    existing = Student.query.filter_by(roll_number=roll_number, is_active=True).first()
                    if existing:
                        # Update existing fields if changed
                        changed = False
                        if existing.name != name:
                            existing.name = name
                            changed = True
                        if existing.course_id != course.id:
                            existing.course_id = course.id
                            changed = True
                        if existing.academic_year != academic_year:
                            existing.academic_year = academic_year
                            changed = True
                        if existing.current_semester != (semester or 1):
                            existing.current_semester = semester or 1
                            changed = True
                        if (email or None) != (existing.email or None):
                            existing.email = email
                            changed = True
                        if changed:
                            updates_applied += 1
                    else:
                        # If inactive exists, hard-delete to free unique key, then create
                        existing_inactive = Student.query.filter_by(roll_number=roll_number, is_active=False).first()
                        if existing_inactive:
                            ok, msg = ManagementService.delete_student_permanently(existing_inactive.id)
                            if not ok:
                                errors.append(f"Row {row_num}: {msg}")
                                continue

                        student = Student(
                            roll_number=roll_number,
                            name=name,
                            course_id=course.id,
                            academic_year=academic_year,
                            current_semester=semester or 1,
                            email=email
                        )

                        students_to_create.append(student)

                except Exception as e:
                    errors.append(f"Row {row_num}: Error processing data - {str(e)}")
            if students_to_create or updates_applied > 0:
                try:
                    # Commit updates to existing students
                    if updates_applied > 0:
                        db.session.commit()
                    
                    # Add new students
                    if students_to_create:
                        success, message = bulk_insert(students_to_create)
                        if not success:
                            return False, message, errors
                    
                    response_msg = []
                    if students_to_create:
                        response_msg.append(f"added {len(students_to_create)}")
                    if updates_applied:
                        response_msg.append(f"updated {updates_applied}")
                    return True, f"Successfully {', '.join(response_msg)} student(s)", errors
                except Exception as e:
                    db.session.rollback()
                    return False, f"Database error: {str(e)}", errors
            else:
                return False, "No valid student changes found", errors

        except Exception as e:
            return False, f"Error processing Excel file: {str(e)}", []
    
    @staticmethod
    def create_course(course_data):
        """Create new course"""
        try:
            # Validate data
            is_valid, message = validate_course_code(course_data.get('code'))
            if not is_valid:
                return False, message
            
            is_valid, message = validate_name(course_data.get('name'), "Course name")
            if not is_valid:
                return False, message
            
            # Check if course code already exists
            existing_active = Course.query.filter_by(code=course_data['code'], is_active=True).first()
            if existing_active:
                return False, "Course code already exists"

            # Remove inactive duplicate course (same code)
            existing_inactive = Course.query.filter_by(code=course_data['code'], is_active=False).first()
            if existing_inactive:
                ok, msg = ManagementService.delete_course_permanently(existing_inactive.id)
                if not ok:
                    return False, msg
            
            course = Course(
                name=course_data['name'],
                code=course_data['code'],
                description=course_data.get('description'),
                duration_years=course_data.get('duration_years', 3),
                total_semesters=course_data.get('total_semesters', 6)
            )
            
            success, message = safe_add_and_commit(course)
            return success, message
            
        except Exception as e:
            return False, f"Error creating course: {str(e)}"
    
    @staticmethod
    def create_subject(subject_data):
        """Create new subject"""
        try:
            # Validate data
            is_valid, message = validate_subject_code(subject_data.get('code'))
            if not is_valid:
                return False, message
            
            is_valid, message = validate_name(subject_data.get('name'), "Subject name")
            if not is_valid:
                return False, message
            
            is_valid, message = validate_academic_year(subject_data.get('year'))
            if not is_valid:
                return False, message
            
            is_valid, message = validate_semester(subject_data.get('semester'))
            if not is_valid:
                return False, message
            
            # Validate course exists
            course = Course.query.get(subject_data.get('course_id'))
            if not course:
                return False, "Invalid course selected"
            
            # Check if subject code already exists for this course
            existing_active = Subject.query.filter_by(
                code=subject_data['code'], 
                course_id=subject_data['course_id'],
                is_active=True
            ).first()
            if existing_active:
                return False, "Subject code already exists for this course"

            # Remove inactive duplicate if present for same course/code
            existing_inactive = Subject.query.filter_by(
                code=subject_data['code'],
                course_id=subject_data['course_id'],
                is_active=False
            ).first()
            if existing_inactive:
                ok, msg = ManagementService.delete_subject_permanently(existing_inactive.id)
                if not ok:
                    return False, msg
            
            subject = Subject(
                name=subject_data['name'],
                code=subject_data['code'],
                course_id=subject_data['course_id'],
                semester=subject_data['semester'],
                year=subject_data['year'],
                description=subject_data.get('description')
            )
            
            success, message = safe_add_and_commit(subject)
            return success, message
            
        except Exception as e:
            return False, f"Error creating subject: {str(e)}"
    
    @staticmethod
    def assign_subjects_to_lecturer(lecturer_id, subject_ids):
        """Assign subjects to an existing lecturer"""
        try:
            lecturer = Lecturer.query.get(lecturer_id)
            if not lecturer:
                return False, "Lecturer not found"
            
            from models.assignments import SubjectAssignment
            from datetime import datetime
            
            current_year = datetime.now().year
            assignments_created = 0
            
            for subject_id in subject_ids:
                # 1) If an active assignment exists, skip
                existing_active = SubjectAssignment.query.filter_by(
                    lecturer_id=lecturer_id,
                    subject_id=int(subject_id),
                    academic_year=current_year,
                    is_active=True
                ).first()

                if existing_active:
                    continue

                # 2) If an inactive assignment exists for this year, reactivate it
                existing_any = SubjectAssignment.query.filter_by(
                    lecturer_id=lecturer_id,
                    subject_id=int(subject_id),
                    academic_year=current_year
                ).first()

                if existing_any:
                    existing_any.is_active = True
                    try:
                        db.session.commit()
                        assignments_created += 1
                    except Exception as e:
                        db.session.rollback()
                        return False, f"Error reactivating assignment for subject {subject_id}: {str(e)}"
                    continue

                # 3) Otherwise create a fresh assignment
                assignment = SubjectAssignment(
                    lecturer_id=lecturer_id,
                    subject_id=int(subject_id),
                    academic_year=current_year
                )
                success, message = safe_add_and_commit(assignment)
                if success:
                    assignments_created += 1
                else:
                    return False, f"Error creating assignment for subject {subject_id}: {message}"
            
            if assignments_created > 0:
                return True, f"Successfully assigned {assignments_created} subject(s) to {lecturer.name}"
            else:
                return True, f"No new subjects assigned to {lecturer.name} (subjects were already assigned)"
                
        except Exception as e:
            return False, f"Error assigning subjects: {str(e)}"
    
    @staticmethod
    def unassign_subject_from_lecturer(lecturer_id, subject_id):
        """Unassign (deactivate) a subject from a lecturer for current academic year"""
        try:
            lecturer = Lecturer.query.get(lecturer_id)
            if not lecturer:
                return False, "Lecturer not found"

            from datetime import datetime
            current_year = datetime.now().year

            assignment = SubjectAssignment.query.filter_by(
                lecturer_id=lecturer_id,
                subject_id=int(subject_id),
                academic_year=current_year,
                is_active=True
            ).first()

            if not assignment:
                # Check if it exists but already inactive
                existing_inactive = SubjectAssignment.query.filter_by(
                    lecturer_id=lecturer_id,
                    subject_id=int(subject_id),
                    academic_year=current_year,
                    is_active=False
                ).first()
                if existing_inactive:
                    return True, "Subject already unassigned for this year"
                return False, "Assignment not found for this year"

            assignment.is_active = False
            db.session.commit()
            return True, f"Unassigned subject successfully from {lecturer.name}"

        except Exception as e:
            db.session.rollback()
            return False, f"Error unassigning subject: {str(e)}"


    @staticmethod
    def delete_lecturer_permanently(lecturer_id):
        """Hard-delete a lecturer and all dependent records referencing the lecturer.

        Removes:
        - Subject assignments
        - Attendance records and monthly summaries
        - Student marks
        Finally deletes the lecturer row itself.
        """
        try:
            lecturer = Lecturer.query.get(lecturer_id)
            if not lecturer:
                return False, "Lecturer not found"

            # Delete dependent rows explicitly to satisfy FK constraints
            from models.assignments import SubjectAssignment
            from models.attendance import AttendanceRecord, MonthlyAttendanceSummary
            from models.marks import StudentMarks

            SubjectAssignment.query.filter_by(lecturer_id=lecturer.id).delete(synchronize_session=False)
            AttendanceRecord.query.filter_by(lecturer_id=lecturer.id).delete(synchronize_session=False)
            MonthlyAttendanceSummary.query.filter_by(lecturer_id=lecturer.id).delete(synchronize_session=False)
            StudentMarks.query.filter_by(lecturer_id=lecturer.id).delete(synchronize_session=False)

            db.session.delete(lecturer)
            db.session.commit()
            return True, "Lecturer permanently deleted"
        except Exception as e:
            db.session.rollback()
            return False, f"Error deleting lecturer: {str(e)}"

    @staticmethod
    def delete_student_permanently(student_id):
        """Hard-delete a student and dependent rows (enrollments, attendance, marks)."""
        try:
            student = Student.query.get(student_id)
            if not student:
                return False, "Student not found"

            from models.student import StudentEnrollment
            from models.attendance import AttendanceRecord, MonthlyStudentAttendance
            from models.marks import StudentMarks

            StudentEnrollment.query.filter_by(student_id=student.id).delete(synchronize_session=False)
            AttendanceRecord.query.filter_by(student_id=student.id).delete(synchronize_session=False)
            # Also delete monthly per-student attendance aggregates
            MonthlyStudentAttendance.query.filter_by(student_id=student.id).delete(synchronize_session=False)
            StudentMarks.query.filter_by(student_id=student.id).delete(synchronize_session=False)

            db.session.delete(student)
            db.session.commit()
            return True, "Student permanently deleted"
        except Exception as e:
            db.session.rollback()
            return False, f"Error deleting student: {str(e)}"

    @staticmethod
    def delete_subject_permanently(subject_id):
        """Hard-delete a subject and all dependent rows (assignments, enrollments, attendance, summaries, marks)."""
        try:
            subject = Subject.query.get(subject_id)
            if not subject:
                return False, "Subject not found"

            from models.assignments import SubjectAssignment
            from models.student import StudentEnrollment
            from models.attendance import AttendanceRecord, MonthlyAttendanceSummary
            from models.marks import StudentMarks

            SubjectAssignment.query.filter_by(subject_id=subject.id).delete(synchronize_session=False)
            StudentEnrollment.query.filter_by(subject_id=subject.id).delete(synchronize_session=False)
            AttendanceRecord.query.filter_by(subject_id=subject.id).delete(synchronize_session=False)
            MonthlyAttendanceSummary.query.filter_by(subject_id=subject.id).delete(synchronize_session=False)
            StudentMarks.query.filter_by(subject_id=subject.id).delete(synchronize_session=False)

            db.session.delete(subject)
            db.session.commit()
            return True, "Subject permanently deleted"
        except Exception as e:
            db.session.rollback()
            return False, f"Error deleting subject: {str(e)}"

    @staticmethod
    def delete_course_permanently(course_id):
        """Hard-delete a course by first deleting all students, subjects and associated data, then the course itself."""
        try:
            course = Course.query.get(course_id)
            if not course:
                return False, "Course not found"

            # First, delete all students enrolled in this course
            students = Student.query.filter_by(course_id=course.id).all()
            for student in students:
                ok, msg = ManagementService.delete_student_permanently(student.id)
                if not ok:
                    return False, msg

            # Then delete all subjects under the course (and their dependencies)
            subjects = Subject.query.filter_by(course_id=course.id).all()
            for subj in subjects:
                ok, msg = ManagementService.delete_subject_permanently(subj.id)
                if not ok:
                    return False, msg

            # Finally delete the course itself
            db.session.delete(course)
            db.session.commit()
            return True, "Course permanently deleted"
        except Exception as e:
            db.session.rollback()
            return False, f"Error deleting course: {str(e)}"

    # ============================ TRACKING QUERIES ============================
    @staticmethod
    def _course_and_class_from_subject(subject):
        """Best-effort course short code and class/year display from subject/course details."""
        course_code = None
        class_display = None
        try:
            if subject and subject.course:
                course_code = subject.course.code
                if getattr(subject, 'year', None):
                    class_display = f"Year {subject.year}"
        except Exception:
            pass
        return course_code, class_display

    @staticmethod
    def get_marks_tracking(assessment_type=None, course_id=None, subject_id=None, lecturer_id=None):
        """Compute lecturers who updated vs pending marks for given filters.

        Returns:
            { 'updated': [...], 'pending': [...] }
        """
        from models.assignments import SubjectAssignment
        from models.marks import StudentMarks
        from datetime import datetime

        current_year = datetime.now().year
        q = SubjectAssignment.query.filter_by(is_active=True, academic_year=current_year)
        if course_id:
            q = q.join(Subject).filter(Subject.course_id == course_id)
        if subject_id:
            q = q.filter(SubjectAssignment.subject_id == subject_id)
        if lecturer_id:
            q = q.filter(SubjectAssignment.lecturer_id == lecturer_id)

        assignments = q.all()
        updated, pending = [], []

        for a in assignments:
            subj = a.subject
            lect = a.lecturer
            if not subj or not lect:
                continue

            marks_q = StudentMarks.query.filter(StudentMarks.subject_id == subj.id)
            if assessment_type:
                marks_q = marks_q.filter(StudentMarks.assessment_type == assessment_type)
            has_any = marks_q.first() is not None

            course_code, class_display = ManagementService._course_and_class_from_subject(subj)
            base_item = {
                'lecturer_name': lect.name,
                'subject_name': subj.name,
                'subject_code': subj.code,
                'course_code': course_code,
                'class_display': class_display,
                'assessment_type': assessment_type or 'Any',
            }

            if has_any:
                updated.append(base_item)
            else:
                try:
                    pending_count = subj.get_enrolled_students_count()
                except Exception:
                    pending_count = None
                pending.append({**base_item, 'pending_count': pending_count})

        return {'updated': updated, 'pending': pending}

    @staticmethod
    def get_attendance_tracking(month, year, course_id=None, subject_id=None, lecturer_id=None):
        """Compute lecturers who updated vs pending attendance summaries for a given month/year."""
        from models.assignments import SubjectAssignment
        from models.attendance import MonthlyAttendanceSummary
        from datetime import datetime

        today = datetime.now()
        m = month or today.month
        y = year or today.year

        # Filter assignments for the same calendar year as a best-effort for academic year
        q = SubjectAssignment.query.filter_by(is_active=True, academic_year=y)
        if course_id:
            q = q.join(Subject).filter(Subject.course_id == course_id)
        if subject_id:
            q = q.filter(SubjectAssignment.subject_id == subject_id)
        if lecturer_id:
            q = q.filter(SubjectAssignment.lecturer_id == lecturer_id)

        assignments = q.all()
        updated, pending = [], []

        for a in assignments:
            subj = a.subject
            lect = a.lecturer
            if not subj or not lect:
                continue

            summary = MonthlyAttendanceSummary.query.filter_by(
                subject_id=subj.id, lecturer_id=lect.id, month=m, year=y
            ).first()

            course_code, class_display = ManagementService._course_and_class_from_subject(subj)
            item = {
                'lecturer_name': lect.name,
                'subject_name': subj.name,
                'subject_code': subj.code,
                'course_code': course_code,
                'class_display': class_display,
                'month': m,
                'year': y,
                'has_summary': bool(summary)
            }
            if summary:
                updated.append(item)
            else:
                pending.append(item)

        return {'updated': updated, 'pending': pending}